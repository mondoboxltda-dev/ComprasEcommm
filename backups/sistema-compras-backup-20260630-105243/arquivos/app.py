from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
import re
import unicodedata
from xml.etree import ElementTree as ET

import mysql.connector
from flask import Flask, flash, jsonify, redirect, render_template, request, send_file, url_for
from openpyxl import Workbook, load_workbook

from config import Config


app = Flask(__name__)
app.config.from_object(Config)


DEFAULT_UNITS = (
    ("UN", "Unidade"),
    ("CX", "Caixa"),
    ("PCT", "Pacote"),
    ("KG", "Quilograma"),
    ("G", "Grama"),
    ("L", "Litro"),
    ("ML", "Mililitro"),
    ("M", "Metro"),
    ("ROLO", "Rolo"),
    ("PALLET", "Pallet"),
    ("FARDO", "Fardo"),
    ("KIT", "Kit"),
    ("PAR", "Par"),
    ("M2", "Metro Quadrado (M²)"),
    ("M3", "Metro Cúbico (M³)"),
)


def get_db():
    return mysql.connector.connect(
        host=app.config["MYSQL_HOST"],
        port=app.config["MYSQL_PORT"],
        user=app.config["MYSQL_USER"],
        password=app.config["MYSQL_PASSWORD"],
        database=app.config["MYSQL_DATABASE"],
    )


def query(sql, params=(), one=False):
    connection = get_db()
    cursor = connection.cursor(dictionary=True)
    cursor.execute(sql, params)
    result = cursor.fetchone() if one else cursor.fetchall()
    cursor.close()
    connection.close()
    return result


def execute(sql, params=()):
    connection = get_db()
    cursor = connection.cursor()
    try:
        cursor.execute(sql, params)
        connection.commit()
        return cursor.lastrowid
    except Exception:
        connection.rollback()
        raise
    finally:
        cursor.close()
        connection.close()


def decimal_value(value):
    return Decimal(str(value or 0))


def money_value(value):
    if value is None or value == "":
        return Decimal("0")
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))
    text = str(value).strip().replace("R$", "").replace(" ", "")
    if "," in text:
        text = text.replace(".", "").replace(",", ".")
    return Decimal(text)


def infer_purchase_conversion(product):
    code = (product.get("codigo") or "").upper().replace(" ", "")
    description = (product.get("descricao") or "").upper()
    unit = (product.get("unidade") or "").upper()
    if unit not in ("PALLET", "CX", "CAIXA", "ROLO"):
        return None
    match = re.search(r"(\d+(?:[,.]\d+)?)\s*(?:CHAPA|CHAPAS|PECA|PECAS|P[EÇ]A|P[EÇ]AS)", description)
    if match:
        factor = money_value(match.group(1))
        base_unit = "CHAPAS" if "CHAPA" in match.group(0) else "PECAS"
        return factor, base_unit
    if "MDF4MM" in code:
        return Decimal("126"), "CHAPAS"
    if "MDF2MM" in code:
        return Decimal("168"), "CHAPAS"
    return None


def normalize_purchase_unit(unit):
    value = (unit or "").strip().upper()
    aliases = {
        "PT": "PALLET",
        "PL": "PALLET",
        "PAL": "PALLET",
        "CH": "CHAPAS",
        "CHAPA": "CHAPAS",
        "PC": "PECAS",
        "PECA": "PECAS",
        "PEÇA": "PECAS",
        "UNID": "UN",
        "UNIDADE": "UN",
    }
    return aliases.get(value, value or "UN")


def conversion_for_receipt_item(product, item_unit):
    product_unit = normalize_purchase_unit(product.get("unidade"))
    purchase_unit = normalize_purchase_unit(item_unit or product.get("unidade"))
    product_factor = decimal_value(product.get("quantidade_por_unidade_compra") or 1)
    if product_factor <= 0:
        product_factor = Decimal("1")
    if purchase_unit == product_unit:
        return purchase_unit, product_factor
    if purchase_unit in ("PALLET", "CX", "CAIXA", "ROLO"):
        return purchase_unit, product_factor
    return purchase_unit, Decimal("1")


def import_cell_value(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).strip()


def normalize_header(value):
    text = unicodedata.normalize("NFKD", import_cell_value(value))
    text = "".join(char for char in text if not unicodedata.combining(char))
    return re.sub(r"\s+", " ", text).strip().lower()


def normalize_cnpj(value):
    digits = re.sub(r"\D", "", value or "")
    if not digits:
        return ""
    if len(digits) == 14:
        return f"{digits[:2]}.{digits[2:5]}.{digits[5:8]}/{digits[8:12]}-{digits[12:]}"
    return value.strip()


def is_valid_cnpj(value):
    digits = re.sub(r"\D", "", value or "")
    if len(digits) != 14 or digits == digits[0] * 14:
        return False

    def check_digit(base, weights):
        total = sum(int(number) * weight for number, weight in zip(base, weights))
        remainder = total % 11
        return "0" if remainder < 2 else str(11 - remainder)

    first = check_digit(digits[:12], [5, 4, 3, 2, 9, 8, 7, 6, 5, 4, 3, 2])
    second = check_digit(digits[:13], [6, 5, 4, 3, 2, 9, 8, 7, 6, 5, 4, 3, 2])
    return digits[-2:] == first + second


def ensure_supplier_schema():
    column = query(
        """
        SELECT IS_NULLABLE
        FROM information_schema.COLUMNS
        WHERE TABLE_SCHEMA=%s AND TABLE_NAME='fornecedores' AND COLUMN_NAME='cnpj'
        """,
        (app.config["MYSQL_DATABASE"],),
        one=True,
    )
    if column and column["IS_NULLABLE"] == "NO":
        execute("ALTER TABLE fornecedores MODIFY cnpj VARCHAR(18) NULL")


def ensure_nfe_schema():
    execute(
        """
        CREATE TABLE IF NOT EXISTS produto_fornecedor_relacionamentos (
          id INT AUTO_INCREMENT PRIMARY KEY,
          fornecedor VARCHAR(150) NOT NULL,
          cnpj VARCHAR(18),
          codigo_fornecedor VARCHAR(80) NOT NULL,
          descricao_fornecedor VARCHAR(255) NOT NULL,
          produto_id INT NOT NULL,
          criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
          atualizado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
          CONSTRAINT fk_rel_produto FOREIGN KEY (produto_id) REFERENCES produtos(id),
          UNIQUE KEY uk_rel_cnpj_codigo (cnpj,codigo_fornecedor),
          KEY idx_rel_fornecedor_codigo (fornecedor,codigo_fornecedor)
        )
        """
    )
    execute(
        """
        CREATE TABLE IF NOT EXISTS nfe_importacoes (
          id INT AUTO_INCREMENT PRIMARY KEY,
          ordem_id INT NULL,
          fornecedor VARCHAR(150) NOT NULL,
          fornecedor_cnpj VARCHAR(18) NOT NULL,
          numero_nf VARCHAR(30) NOT NULL,
          serie VARCHAR(10),
          chave_nfe VARCHAR(60) NOT NULL UNIQUE,
          data_emissao DATE,
          data_entrada DATE NOT NULL,
          valor_total DECIMAL(14,2) NOT NULL DEFAULT 0,
          criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
          CONSTRAINT fk_nfe_ordem FOREIGN KEY (ordem_id) REFERENCES ordens_compra(id),
          UNIQUE KEY uk_nfe_fornecedor_numero_serie (fornecedor_cnpj,numero_nf,serie)
        )
        """
    )
    if not column_exists("nfe_importacoes", "ordem_id"):
        execute("ALTER TABLE nfe_importacoes ADD COLUMN ordem_id INT NULL AFTER id")
        constraint = query(
            """
            SELECT CONSTRAINT_NAME
            FROM information_schema.TABLE_CONSTRAINTS
            WHERE TABLE_SCHEMA=%s AND TABLE_NAME='nfe_importacoes' AND CONSTRAINT_NAME='fk_nfe_ordem'
            """,
            (app.config["MYSQL_DATABASE"],),
            one=True,
        )
        if not constraint:
            execute("ALTER TABLE nfe_importacoes ADD CONSTRAINT fk_nfe_ordem FOREIGN KEY (ordem_id) REFERENCES ordens_compra(id)")
    nfe_columns = (
        ("origem", "ALTER TABLE nfe_importacoes ADD COLUMN origem ENUM('XML','Manual') NOT NULL DEFAULT 'XML' AFTER ordem_id"),
        ("tipo_documento", "ALTER TABLE nfe_importacoes ADD COLUMN tipo_documento VARCHAR(40) NOT NULL DEFAULT 'Nota Fiscal' AFTER numero_nf"),
        ("condicao_pagamento", "ALTER TABLE nfe_importacoes ADD COLUMN condicao_pagamento VARCHAR(120) NULL AFTER valor_total"),
        ("observacoes", "ALTER TABLE nfe_importacoes ADD COLUMN observacoes TEXT NULL AFTER condicao_pagamento"),
        ("usuario_responsavel", "ALTER TABLE nfe_importacoes ADD COLUMN usuario_responsavel VARCHAR(120) NOT NULL DEFAULT 'Sistema' AFTER observacoes"),
        ("status_recebimento", "ALTER TABLE nfe_importacoes ADD COLUMN status_recebimento VARCHAR(60) NOT NULL DEFAULT 'Pendente' AFTER usuario_responsavel"),
        ("confirmado_em", "ALTER TABLE nfe_importacoes ADD COLUMN confirmado_em DATETIME NULL AFTER status_recebimento"),
    )
    for column, statement in nfe_columns:
        if not column_exists("nfe_importacoes", column):
            execute(statement)
    execute(
        """
        CREATE TABLE IF NOT EXISTS nfe_importacao_itens (
          id INT AUTO_INCREMENT PRIMARY KEY,
          nfe_importacao_id INT NOT NULL,
          produto_id INT NULL,
          codigo_fornecedor VARCHAR(80),
          descricao VARCHAR(255) NOT NULL,
          ncm VARCHAR(20),
          cfop VARCHAR(10),
          cest VARCHAR(20),
          unidade VARCHAR(20),
          quantidade DECIMAL(14,4) NOT NULL DEFAULT 0,
          valor_unitario DECIMAL(14,4) NOT NULL DEFAULT 0,
          valor_total DECIMAL(14,2) NOT NULL DEFAULT 0,
          ean VARCHAR(30),
          status ENUM('vinculado','pendente') NOT NULL DEFAULT 'pendente',
          CONSTRAINT fk_nfe_item_importacao FOREIGN KEY (nfe_importacao_id) REFERENCES nfe_importacoes(id) ON DELETE CASCADE,
          CONSTRAINT fk_nfe_item_produto FOREIGN KEY (produto_id) REFERENCES produtos(id)
        )
        """
    )
    if not column_exists("nfe_importacao_itens", "desconto"):
        execute("ALTER TABLE nfe_importacao_itens ADD COLUMN desconto DECIMAL(14,2) NOT NULL DEFAULT 0 AFTER valor_unitario")
    execute(
        """
        CREATE TABLE IF NOT EXISTS movimentacoes_estoque (
          id INT AUTO_INCREMENT PRIMARY KEY,
          produto_id INT NOT NULL,
          recebimento_id INT NOT NULL,
          recebimento_item_id INT NOT NULL,
          data_movimentacao DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
          tipo VARCHAR(30) NOT NULL,
          quantidade DECIMAL(14,4) NOT NULL,
          valor_unitario DECIMAL(14,4) NOT NULL DEFAULT 0,
          valor_total DECIMAL(14,2) NOT NULL DEFAULT 0,
          origem VARCHAR(20) NOT NULL,
          usuario VARCHAR(120) NOT NULL DEFAULT 'Sistema',
          UNIQUE KEY uk_mov_item (recebimento_item_id,tipo),
          CONSTRAINT fk_mov_produto FOREIGN KEY (produto_id) REFERENCES produtos(id),
          CONSTRAINT fk_mov_recebimento FOREIGN KEY (recebimento_id) REFERENCES nfe_importacoes(id),
          CONSTRAINT fk_mov_item FOREIGN KEY (recebimento_item_id) REFERENCES nfe_importacao_itens(id) ON DELETE CASCADE
        )
        """
    )
    execute(
        """
        CREATE TABLE IF NOT EXISTS historico_custos (
          id INT AUTO_INCREMENT PRIMARY KEY,
          produto_id INT NOT NULL,
          fornecedor VARCHAR(150) NOT NULL,
          documento VARCHAR(60),
          data_entrada DATE NOT NULL,
          quantidade DECIMAL(14,4) NOT NULL,
          valor_unitario DECIMAL(14,4) NOT NULL,
          valor_total DECIMAL(14,2) NOT NULL,
          origem VARCHAR(20) NOT NULL,
          recebimento_id INT NOT NULL,
          recebimento_item_id INT NOT NULL,
          criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
          UNIQUE KEY uk_custo_item (recebimento_item_id,origem),
          CONSTRAINT fk_custo_produto FOREIGN KEY (produto_id) REFERENCES produtos(id),
          CONSTRAINT fk_custo_recebimento FOREIGN KEY (recebimento_id) REFERENCES nfe_importacoes(id),
          CONSTRAINT fk_custo_item FOREIGN KEY (recebimento_item_id) REFERENCES nfe_importacao_itens(id) ON DELETE CASCADE
        )
        """
    )
    if table_exists("historico_custos"):
        if not column_exists("historico_custos", "unidade_compra"):
            execute("ALTER TABLE historico_custos ADD COLUMN unidade_compra VARCHAR(30) NULL AFTER quantidade")
        if not column_exists("historico_custos", "quantidade_por_unidade_compra"):
            execute("ALTER TABLE historico_custos ADD COLUMN quantidade_por_unidade_compra DECIMAL(14,4) NOT NULL DEFAULT 1 AFTER unidade_compra")
        if not column_exists("historico_custos", "unidade_base"):
            execute("ALTER TABLE historico_custos ADD COLUMN unidade_base VARCHAR(30) NULL AFTER quantidade_por_unidade_compra")
        if not column_exists("historico_custos", "quantidade_total_base"):
            execute("ALTER TABLE historico_custos ADD COLUMN quantidade_total_base DECIMAL(14,4) NOT NULL DEFAULT 0 AFTER unidade_base")
        if not column_exists("historico_custos", "valor_unitario_base"):
            execute("ALTER TABLE historico_custos ADD COLUMN valor_unitario_base DECIMAL(14,4) NOT NULL DEFAULT 0 AFTER valor_unitario")
    execute(
        """
        CREATE TABLE IF NOT EXISTS recebimento_auditoria (
          id INT AUTO_INCREMENT PRIMARY KEY,
          recebimento_id INT NOT NULL,
          usuario VARCHAR(120) NOT NULL DEFAULT 'Sistema',
          acao VARCHAR(80) NOT NULL,
          origem VARCHAR(20) NOT NULL,
          detalhes TEXT,
          criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
          CONSTRAINT fk_audit_recebimento FOREIGN KEY (recebimento_id) REFERENCES nfe_importacoes(id) ON DELETE CASCADE
        )
        """
    )


def nfe_text(parent, path, namespaces=None, default=""):
    node = parent.find(path, namespaces or {})
    return (node.text or "").strip() if node is not None and node.text is not None else default


def nfe_decimal(value):
    return Decimal(str(value or "0").replace(",", "."))


def parse_nfe_xml(file_storage):
    try:
        tree = ET.parse(file_storage)
    except ET.ParseError as error:
        raise ValueError(f"XML invalido: {error}") from error

    root = tree.getroot()
    namespace = root.tag.split("}")[0].strip("{") if root.tag.startswith("{") else ""
    ns = {"n": namespace} if namespace else {}
    prefix = "n:" if namespace else ""
    inf = root.find(f".//{prefix}infNFe", ns)
    if inf is None:
        raise ValueError("O arquivo XML nao parece ser uma NF-e.")

    ide = inf.find(f"{prefix}ide", ns)
    emit = inf.find(f"{prefix}emit", ns)
    total = inf.find(f"{prefix}total/{prefix}ICMSTot", ns)
    if ide is None or emit is None:
        raise ValueError("Nao foi possivel ler cabecalho da NF-e.")

    chave = (inf.attrib.get("Id") or "").replace("NFe", "")
    numero_nf = nfe_text(ide, f"{prefix}nNF", ns)
    serie = nfe_text(ide, f"{prefix}serie", ns)
    emissao_raw = nfe_text(ide, f"{prefix}dhEmi", ns) or nfe_text(ide, f"{prefix}dEmi", ns)
    data_emissao = emissao_raw[:10] if emissao_raw else None
    fornecedor = nfe_text(emit, f"{prefix}xNome", ns)
    cnpj = normalize_cnpj(nfe_text(emit, f"{prefix}CNPJ", ns))
    telefone = nfe_text(emit, f"{prefix}enderEmit/{prefix}fone", ns)
    valor_total = nfe_decimal(nfe_text(total, f"{prefix}vNF", ns) if total is not None else "0")

    if not chave or not numero_nf or not fornecedor or not cnpj:
        raise ValueError("NF-e sem chave, numero, fornecedor ou CNPJ.")

    items = []
    for det in inf.findall(f"{prefix}det", ns):
        prod = det.find(f"{prefix}prod", ns)
        if prod is None:
            continue
        items.append(
            {
                "codigo_fornecedor": nfe_text(prod, f"{prefix}cProd", ns),
                "descricao": nfe_text(prod, f"{prefix}xProd", ns),
                "ncm": nfe_text(prod, f"{prefix}NCM", ns),
                "cfop": nfe_text(prod, f"{prefix}CFOP", ns),
                "cest": nfe_text(prod, f"{prefix}CEST", ns),
                "unidade": nfe_text(prod, f"{prefix}uCom", ns),
                "quantidade": nfe_decimal(nfe_text(prod, f"{prefix}qCom", ns)),
                "valor_unitario": nfe_decimal(nfe_text(prod, f"{prefix}vUnCom", ns)),
                "valor_total": nfe_decimal(nfe_text(prod, f"{prefix}vProd", ns)),
                "ean": nfe_text(prod, f"{prefix}cEAN", ns),
            }
        )
    if not items:
        raise ValueError("NF-e sem produtos para importar.")

    return {
        "fornecedor": fornecedor,
        "fornecedor_cnpj": cnpj,
        "fornecedor_telefone": telefone,
        "numero_nf": numero_nf,
        "serie": serie,
        "chave_nfe": chave,
        "data_emissao": data_emissao,
        "data_entrada": date.today().isoformat(),
        "valor_total": valor_total,
        "items": items,
    }


def audit_receipt(recebimento_id, acao, detalhes="", usuario="Sistema", origem=None):
    receipt = query("SELECT origem FROM nfe_importacoes WHERE id=%s", (recebimento_id,), one=True)
    execute(
        """
        INSERT INTO recebimento_auditoria (recebimento_id,usuario,acao,origem,detalhes)
        VALUES (%s,%s,%s,%s,%s)
        """,
        (recebimento_id, usuario or "Sistema", acao, origem or (receipt["origem"] if receipt else "Manual"), detalhes),
    )


def apply_receipt_effects(recebimento_id):
    ensure_product_schema()
    ensure_nfe_schema()
    receipt = query("SELECT * FROM nfe_importacoes WHERE id=%s", (recebimento_id,), one=True)
    if not receipt:
        return {"processed": 0}
    items = query(
        """
        SELECT i.*
        FROM nfe_importacao_itens i
        WHERE i.nfe_importacao_id=%s AND i.produto_id IS NOT NULL
        ORDER BY i.id
        """,
        (recebimento_id,),
    )
    processed = 0
    for item in items:
        exists = query(
            "SELECT id FROM movimentacoes_estoque WHERE recebimento_item_id=%s AND tipo='entrada'",
            (item["id"],),
            one=True,
        )
        if exists:
            continue
        product = query(
            """
            SELECT estoque_atual,custo_atual,unidade,quantidade_por_unidade_compra,unidade_base
            FROM produtos
            WHERE id=%s
            """,
            (item["produto_id"],),
            one=True,
        )
        if not product:
            continue
        qty = decimal_value(item["quantidade"])
        unit_cost = decimal_value(item["valor_unitario"])
        total_value = decimal_value(item["valor_total"])
        purchase_unit, purchase_factor = conversion_for_receipt_item(product, item.get("unidade"))
        base_qty = qty * purchase_factor
        base_unit_cost = total_value / base_qty if base_qty > 0 else unit_cost
        current_qty = decimal_value(product["estoque_atual"])
        current_cost = decimal_value(product["custo_atual"])
        new_qty = current_qty + qty
        new_cost = ((current_qty * current_cost) + total_value) / new_qty if new_qty > 0 else unit_cost
        execute(
            "UPDATE produtos SET estoque_atual=%s,custo_atual=%s WHERE id=%s",
            (new_qty, new_cost, item["produto_id"]),
        )
        execute(
            """
            INSERT INTO movimentacoes_estoque
            (produto_id,recebimento_id,recebimento_item_id,tipo,quantidade,valor_unitario,valor_total,origem,usuario)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """,
            (
                item["produto_id"],
                recebimento_id,
                item["id"],
                "entrada",
                qty,
                unit_cost,
                total_value,
                receipt["origem"],
                receipt.get("usuario_responsavel") or "Sistema",
            ),
        )
        execute(
            """
            INSERT INTO historico_custos
            (produto_id,fornecedor,documento,data_entrada,quantidade,unidade_compra,quantidade_por_unidade_compra,
             unidade_base,quantidade_total_base,valor_unitario,valor_unitario_base,valor_total,origem,recebimento_id,recebimento_item_id)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """,
            (
                item["produto_id"],
                receipt["fornecedor"],
                receipt["numero_nf"],
                receipt["data_entrada"],
                qty,
                purchase_unit,
                purchase_factor,
                product.get("unidade_base") or product.get("unidade") or item.get("unidade") or "UN",
                base_qty,
                unit_cost,
                base_unit_cost,
                total_value,
                receipt["origem"],
                recebimento_id,
                item["id"],
            ),
        )
        processed += 1
    audit_receipt(recebimento_id, "Processamento de estoque/custo", f"{processed} item(ns) processado(s).", receipt.get("usuario_responsavel") or "Sistema", receipt["origem"])
    return {"processed": processed}


def calculate_receipt_status(recebimento_id):
    receipt = query(
        "SELECT id,origem,ordem_id,status_recebimento FROM nfe_importacoes WHERE id=%s",
        (recebimento_id,),
        one=True,
    )
    if not receipt:
        return None
    pending = query(
        """
        SELECT COUNT(1) total
        FROM nfe_importacao_itens
        WHERE nfe_importacao_id=%s AND status='pendente'
        """,
        (recebimento_id,),
        one=True,
    )["total"]
    if pending:
        return "Produto Pendente de Cadastro"
    if receipt["status_recebimento"] == "Recebido Parcialmente":
        return "Recebido Parcialmente"
    if receipt["origem"] == "XML" and not receipt["ordem_id"]:
        return "Aguardando Ordem de Compra"
    return "Recebido"


def refresh_receipt_status(recebimento_id):
    status = calculate_receipt_status(recebimento_id)
    if status:
        execute("UPDATE nfe_importacoes SET status_recebimento=%s WHERE id=%s", (status, recebimento_id))
    return status


def refresh_receipt_statuses():
    for receipt in query("SELECT id FROM nfe_importacoes"):
        refresh_receipt_status(receipt["id"])


def serialize_row(row):
    serialized = {}
    for key, value in row.items():
        if isinstance(value, Decimal):
            serialized[key] = float(value)
        elif isinstance(value, (date, datetime)):
            serialized[key] = value.isoformat()
        else:
            serialized[key] = value
    return serialized


def purchase_history_where(filters):
    where = []
    params = []
    if filters.get("codigo"):
        where.append("p.codigo LIKE %s")
        params.append(f"%{filters['codigo']}%")
    if filters.get("produto"):
        where.append("p.descricao LIKE %s")
        params.append(f"%{filters['produto']}%")
    if filters.get("fornecedor"):
        where.append("h.fornecedor LIKE %s")
        params.append(f"%{filters['fornecedor']}%")
    if filters.get("categoria"):
        where.append("p.categoria=%s")
        params.append(filters["categoria"])
    if filters.get("nf"):
        where.append("h.documento LIKE %s")
        params.append(f"%{filters['nf']}%")
    if filters.get("data_inicio"):
        where.append("h.data_entrada >= %s")
        params.append(filters["data_inicio"])
    if filters.get("data_fim"):
        where.append("h.data_entrada <= %s")
        params.append(filters["data_fim"])
    if filters.get("unidade"):
        where.append("p.unidade=%s")
        params.append(filters["unidade"])
    if filters.get("preco_min"):
        where.append("h.valor_unitario >= %s")
        params.append(money_value(filters["preco_min"]))
    if filters.get("preco_max"):
        where.append("h.valor_unitario <= %s")
        params.append(money_value(filters["preco_max"]))
    if filters.get("origem"):
        where.append("h.origem=%s")
        params.append(filters["origem"])
    if filters.get("status"):
        where.append("n.status_recebimento=%s")
        params.append(filters["status"])
    return ("WHERE " + " AND ".join(where)) if where else "", params


def request_purchase_history_filters():
    return {
        "codigo": request.args.get("codigo") or "",
        "produto": request.args.get("produto") or "",
        "fornecedor": request.args.get("fornecedor") or "",
        "categoria": request.args.get("categoria") or "",
        "nf": request.args.get("nf") or "",
        "data_inicio": request.args.get("data_inicio") or "",
        "data_fim": request.args.get("data_fim") or "",
        "unidade": request.args.get("unidade") or "",
        "preco_min": request.args.get("preco_min") or "",
        "preco_max": request.args.get("preco_max") or "",
        "origem": request.args.get("origem") or "",
        "status": request.args.get("status") or "",
    }


def identify_nfe_product(fornecedor, cnpj, codigo_fornecedor, descricao):
    rel = query(
        """
        SELECT p.id,p.codigo
        FROM produto_fornecedor_relacionamentos r
        JOIN produtos p ON p.id=r.produto_id
        WHERE (r.cnpj=%s OR LOWER(r.fornecedor)=LOWER(%s)) AND r.codigo_fornecedor=%s
        ORDER BY r.id DESC LIMIT 1
        """,
        (cnpj, fornecedor, codigo_fornecedor),
        one=True,
    )
    if rel:
        return rel

    product = query(
        """
        SELECT p.id,p.codigo
        FROM produtos p
        LEFT JOIN fornecedores f ON f.id=p.fornecedor_id
        WHERE p.status <> 'inativo'
          AND (f.cnpj=%s OR LOWER(f.nome)=LOWER(%s))
          AND (p.codigo=%s OR LOWER(p.descricao)=LOWER(%s))
        ORDER BY (p.codigo=%s) DESC
        LIMIT 1
        """,
        (cnpj, fornecedor, codigo_fornecedor, descricao, codigo_fornecedor),
        one=True,
    )
    if product:
        return product
    return None


def get_or_create_nfe_supplier(fornecedor, cnpj, telefone=""):
    supplier = query("SELECT id,nome FROM fornecedores WHERE cnpj=%s", (cnpj,), one=True) if cnpj else None
    if supplier:
        return {"id": supplier["id"], "nome": supplier["nome"], "created": False}
    supplier = query("SELECT id,nome FROM fornecedores WHERE LOWER(nome)=LOWER(%s)", (fornecedor,), one=True)
    if supplier:
        return {"id": supplier["id"], "nome": supplier["nome"], "created": False}
    supplier_id = execute(
        "INSERT INTO fornecedores (nome,cnpj,contato,telefone,email,status) VALUES (%s,%s,%s,%s,%s,%s)",
        (
            fornecedor or "Não informado",
            cnpj or None,
            "Não informado",
            telefone or "Não informado",
            "Não informado",
            "ativo",
        ),
    )
    return {"id": supplier_id, "nome": fornecedor or "Não informado", "created": True}


def create_pending_product_from_nfe_item(item, fornecedor_id):
    code = "Não registrado"
    descricao = item["descricao"] or "Produto importado da NF-e"
    produto_id = execute(
        """
        INSERT INTO produtos
        (codigo,descricao,categoria,unidade,quantidade_por_unidade_compra,unidade_base,fornecedor_id,estoque_seguranca,status)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """,
        (
            code,
            descricao,
            "Importado via NF-e",
            item["unidade"] or "UN",
            1,
            item["unidade"] or "UN",
            fornecedor_id,
            0,
            "pendente revisao",
        ),
    )
    return {"id": produto_id, "codigo": code, "descricao": descricao, "created": True}


def register_product_supplier_link(fornecedor, cnpj, item, produto_id):
    if not item.get("codigo_fornecedor"):
        return
    execute(
        """
        INSERT INTO produto_fornecedor_relacionamentos
        (fornecedor,cnpj,codigo_fornecedor,descricao_fornecedor,produto_id)
        VALUES (%s,%s,%s,%s,%s)
        ON DUPLICATE KEY UPDATE
          descricao_fornecedor=VALUES(descricao_fornecedor),
          produto_id=VALUES(produto_id),
          atualizado_em=CURRENT_TIMESTAMP
        """,
        (
            fornecedor,
            cnpj,
            item["codigo_fornecedor"],
            item["descricao"],
            produto_id,
        ),
    )


def column_exists(table_name, column_name):
    return query(
        """
        SELECT COLUMN_NAME
        FROM information_schema.COLUMNS
        WHERE TABLE_SCHEMA=%s AND TABLE_NAME=%s AND COLUMN_NAME=%s
        """,
        (app.config["MYSQL_DATABASE"], table_name, column_name),
        one=True,
    )


def table_exists(table_name):
    return query(
        """
        SELECT TABLE_NAME
        FROM information_schema.TABLES
        WHERE TABLE_SCHEMA=%s AND TABLE_NAME=%s
        """,
        (app.config["MYSQL_DATABASE"], table_name),
        one=True,
    )


def ensure_auxiliary_registries():
    execute(
        """
        CREATE TABLE IF NOT EXISTS cadastro_categorias (
          id INT AUTO_INCREMENT PRIMARY KEY,
          nome VARCHAR(120) NOT NULL UNIQUE,
          status ENUM('ativo','inativo') NOT NULL DEFAULT 'ativo',
          criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
          atualizado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
        )
        """
    )
    execute(
        """
        CREATE TABLE IF NOT EXISTS cadastro_tipos_pagamento (
          id INT AUTO_INCREMENT PRIMARY KEY,
          nome VARCHAR(120) NOT NULL UNIQUE,
          status ENUM('ativo','inativo') NOT NULL DEFAULT 'ativo',
          criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
          atualizado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
        )
        """
    )
    for row in query(
        """
        SELECT categoria nome FROM produtos WHERE categoria IS NOT NULL AND categoria<>''
        UNION
        SELECT categoria nome FROM orcamentos WHERE categoria IS NOT NULL AND categoria<>''
        """
    ):
        execute(
            "INSERT IGNORE INTO cadastro_categorias (nome,status) VALUES (%s,'ativo')",
            (row["nome"],),
        )
    execute(
        "INSERT IGNORE INTO cadastro_categorias (nome,status) VALUES (%s,'ativo')",
        ("Importado via NF-e",),
    )
    for nome in ("PIX", "Boleto", "Cartao de Credito", "Cartao de Debito", "Dinheiro", "Transferencia"):
        execute(
            "INSERT IGNORE INTO cadastro_tipos_pagamento (nome,status) VALUES (%s,'ativo')",
            (nome,),
        )
    for row in query("SELECT DISTINCT metodo_pagamento nome FROM ordens_compra WHERE metodo_pagamento IS NOT NULL AND metodo_pagamento<>''"):
        execute(
            "INSERT IGNORE INTO cadastro_tipos_pagamento (nome,status) VALUES (%s,'ativo')",
            (row["nome"],),
        )


def ensure_product_schema():
    execute(
        """
        CREATE TABLE IF NOT EXISTS unidades_medida (
          id INT AUTO_INCREMENT PRIMARY KEY,
          codigo VARCHAR(20) NOT NULL UNIQUE,
          nome VARCHAR(120) NOT NULL,
          status ENUM('ativo','inativo') NOT NULL DEFAULT 'ativo',
          criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        """
    )
    for codigo, nome in DEFAULT_UNITS:
        execute(
            """
            INSERT INTO unidades_medida (codigo,nome,status)
            VALUES (%s,%s,'ativo')
            ON DUPLICATE KEY UPDATE nome=VALUES(nome)
            """,
            (codigo, nome),
        )
    if not column_exists("produtos", "fornecedor_id"):
        execute("ALTER TABLE produtos ADD COLUMN fornecedor_id INT NULL AFTER unidade")
    if not column_exists("produtos", "estoque_seguranca"):
        execute("ALTER TABLE produtos ADD COLUMN estoque_seguranca DECIMAL(12,2) NOT NULL DEFAULT 0 AFTER fornecedor_id")
    if not column_exists("produtos", "quantidade_por_unidade_compra"):
        execute("ALTER TABLE produtos ADD COLUMN quantidade_por_unidade_compra DECIMAL(14,4) NOT NULL DEFAULT 1 AFTER unidade")
    if not column_exists("produtos", "unidade_base"):
        execute("ALTER TABLE produtos ADD COLUMN unidade_base VARCHAR(30) NOT NULL DEFAULT 'UN' AFTER quantidade_por_unidade_compra")
    if not column_exists("produtos", "estoque_atual"):
        execute("ALTER TABLE produtos ADD COLUMN estoque_atual DECIMAL(14,4) NOT NULL DEFAULT 0 AFTER estoque_seguranca")
    if not column_exists("produtos", "custo_atual"):
        execute("ALTER TABLE produtos ADD COLUMN custo_atual DECIMAL(14,4) NOT NULL DEFAULT 0 AFTER estoque_atual")
    if not column_exists("historico_custos", "unidade_compra"):
        execute("ALTER TABLE historico_custos ADD COLUMN unidade_compra VARCHAR(30) NULL AFTER quantidade")
    if not column_exists("historico_custos", "quantidade_por_unidade_compra"):
        execute("ALTER TABLE historico_custos ADD COLUMN quantidade_por_unidade_compra DECIMAL(14,4) NOT NULL DEFAULT 1 AFTER unidade_compra")
    if not column_exists("historico_custos", "unidade_base"):
        execute("ALTER TABLE historico_custos ADD COLUMN unidade_base VARCHAR(30) NULL AFTER quantidade_por_unidade_compra")
    if not column_exists("historico_custos", "quantidade_total_base"):
        execute("ALTER TABLE historico_custos ADD COLUMN quantidade_total_base DECIMAL(14,4) NOT NULL DEFAULT 0 AFTER unidade_base")
    if not column_exists("historico_custos", "valor_unitario_base"):
        execute("ALTER TABLE historico_custos ADD COLUMN valor_unitario_base DECIMAL(14,4) NOT NULL DEFAULT 0 AFTER valor_unitario")
    execute("ALTER TABLE produtos MODIFY codigo VARCHAR(30) NOT NULL")
    unique_code_indexes = query(
        """
        SELECT INDEX_NAME
        FROM information_schema.STATISTICS
        WHERE TABLE_SCHEMA=%s
          AND TABLE_NAME='produtos'
          AND COLUMN_NAME='codigo'
          AND NON_UNIQUE=0
          AND INDEX_NAME <> 'PRIMARY'
        """,
        (app.config["MYSQL_DATABASE"],),
    )
    for index in unique_code_indexes:
        index_name = index["INDEX_NAME"].replace("`", "``")
        execute(f"ALTER TABLE produtos DROP INDEX `{index_name}`")
    execute(
        """
        ALTER TABLE produtos
        MODIFY status ENUM('ativo','inativo','pendente revisao') NOT NULL DEFAULT 'ativo'
        """
    )
    execute("UPDATE produtos SET unidade='UN' WHERE LOWER(unidade) IN ('un','unidade')")
    execute("UPDATE produtos SET unidade='L' WHERE unidade='GL'")
    execute("UPDATE produtos SET quantidade_por_unidade_compra=1 WHERE quantidade_por_unidade_compra IS NULL OR quantidade_por_unidade_compra<=0")
    execute("UPDATE produtos SET unidade_base=unidade WHERE unidade_base IS NULL OR unidade_base=''")
    for product in query(
        """
        SELECT id,codigo,descricao,unidade,quantidade_por_unidade_compra,unidade_base
        FROM produtos
        WHERE (
            UPPER(unidade) IN ('PALLET','CX','CAIXA','ROLO')
            AND (
              COALESCE(quantidade_por_unidade_compra,1)<=1
              OR UPPER(COALESCE(unidade_base,'')) IN ('','UN')
            )
          )
          OR (
            UPPER(codigo) LIKE 'MDF%%'
            AND UPPER(COALESCE(unidade_base,'')) IN ('','UN')
          )
        """
    ):
        inferred = infer_purchase_conversion(product)
        if inferred:
            execute(
                "UPDATE produtos SET quantidade_por_unidade_compra=%s,unidade_base=%s WHERE id=%s",
                (inferred[0], inferred[1], product["id"]),
            )
        elif (product.get("codigo") or "").upper().replace(" ", "").startswith("MDF"):
            execute("UPDATE produtos SET unidade_base=%s WHERE id=%s", ("CHAPAS", product["id"]))
    if table_exists("historico_custos"):
        execute(
            """
            UPDATE historico_custos h
            JOIN produtos p ON p.id=h.produto_id
            LEFT JOIN nfe_importacao_itens i ON i.id=h.recebimento_item_id
            SET h.unidade_compra=CASE UPPER(COALESCE(NULLIF(i.unidade,''),p.unidade))
                    WHEN 'PT' THEN 'PALLET'
                    WHEN 'PL' THEN 'PALLET'
                    WHEN 'PAL' THEN 'PALLET'
                    WHEN 'CH' THEN 'CHAPAS'
                    WHEN 'CHAPA' THEN 'CHAPAS'
                    WHEN 'PC' THEN 'PECAS'
                    WHEN 'PECA' THEN 'PECAS'
                    WHEN 'PEÇA' THEN 'PECAS'
                    WHEN 'UNID' THEN 'UN'
                    WHEN 'UNIDADE' THEN 'UN'
                    ELSE UPPER(COALESCE(NULLIF(i.unidade,''),p.unidade,'UN'))
                END,
                h.quantidade_por_unidade_compra=CASE
                    WHEN (
                        CASE UPPER(COALESCE(NULLIF(i.unidade,''),p.unidade))
                            WHEN 'PT' THEN 'PALLET'
                            WHEN 'PL' THEN 'PALLET'
                            WHEN 'PAL' THEN 'PALLET'
                            WHEN 'CH' THEN 'CHAPAS'
                            WHEN 'CHAPA' THEN 'CHAPAS'
                            WHEN 'PC' THEN 'PECAS'
                            WHEN 'PECA' THEN 'PECAS'
                            WHEN 'PEÇA' THEN 'PECAS'
                            WHEN 'UNID' THEN 'UN'
                            WHEN 'UNIDADE' THEN 'UN'
                            ELSE UPPER(COALESCE(NULLIF(i.unidade,''),p.unidade,'UN'))
                        END
                    ) = (
                        CASE UPPER(COALESCE(p.unidade,'UN'))
                            WHEN 'PT' THEN 'PALLET'
                            WHEN 'PL' THEN 'PALLET'
                            WHEN 'PAL' THEN 'PALLET'
                            WHEN 'CH' THEN 'CHAPAS'
                            WHEN 'CHAPA' THEN 'CHAPAS'
                            WHEN 'PC' THEN 'PECAS'
                            WHEN 'PECA' THEN 'PECAS'
                            WHEN 'PEÇA' THEN 'PECAS'
                            WHEN 'UNID' THEN 'UN'
                            WHEN 'UNIDADE' THEN 'UN'
                            ELSE UPPER(COALESCE(p.unidade,'UN'))
                        END
                    )
                    OR (
                        CASE UPPER(COALESCE(NULLIF(i.unidade,''),p.unidade))
                            WHEN 'PT' THEN 'PALLET'
                            WHEN 'PL' THEN 'PALLET'
                            WHEN 'PAL' THEN 'PALLET'
                            ELSE UPPER(COALESCE(NULLIF(i.unidade,''),p.unidade,'UN'))
                        END
                    ) IN ('PALLET','CX','CAIXA','ROLO')
                    THEN COALESCE(NULLIF(p.quantidade_por_unidade_compra,0),1)
                    ELSE 1
                END,
                h.unidade_base=COALESCE(NULLIF(p.unidade_base,''),p.unidade),
                h.quantidade_total_base=h.quantidade * h.quantidade_por_unidade_compra,
                h.valor_unitario_base=h.valor_total / NULLIF(h.quantidade * h.quantidade_por_unidade_compra,0)
            WHERE h.produto_id=p.id
            """
        )
        execute(
            """
            UPDATE historico_custos
            SET quantidade_total_base=quantidade * COALESCE(NULLIF(quantidade_por_unidade_compra,0),1),
                valor_unitario_base=valor_total / NULLIF(quantidade * COALESCE(NULLIF(quantidade_por_unidade_compra,0),1),0)
            """
        )
    execute("UPDATE produtos SET estoque_seguranca=COALESCE(NULLIF(estoque_seguranca,0),estoque_semanal,0)")
    execute(
        """
        UPDATE produtos p
        SET fornecedor_id = (
            SELECT o.fornecedor_id
            FROM ordens_compra o
            WHERE o.produto_id=p.id
            ORDER BY o.data_preenchimento DESC, o.id DESC
            LIMIT 1
        )
        WHERE p.fornecedor_id IS NULL
        """
    )
    execute(
        """
        UPDATE produtos p
        SET fornecedor_id = (SELECT f.id FROM fornecedores f WHERE f.status='ativo' ORDER BY f.nome LIMIT 1)
        WHERE p.fornecedor_id IS NULL
        """
    )
    constraint = query(
        """
        SELECT CONSTRAINT_NAME
        FROM information_schema.TABLE_CONSTRAINTS
        WHERE TABLE_SCHEMA=%s AND TABLE_NAME='produtos' AND CONSTRAINT_NAME='fk_produto_fornecedor'
        """,
        (app.config["MYSQL_DATABASE"],),
        one=True,
    )
    if not constraint:
        execute("ALTER TABLE produtos ADD CONSTRAINT fk_produto_fornecedor FOREIGN KEY (fornecedor_id) REFERENCES fornecedores(id)")


def ensure_order_schema():
    if not column_exists("ordens_compra", "parcelas"):
        execute("ALTER TABLE ordens_compra ADD COLUMN parcelas TINYINT NOT NULL DEFAULT 1 AFTER metodo_pagamento")
    if not column_exists("ordens_compra", "prazos_parcelas"):
        execute("ALTER TABLE ordens_compra ADD COLUMN prazos_parcelas VARCHAR(255) NULL AFTER parcelas")


@app.template_filter("moeda")
def moeda(value):
    number = decimal_value(value)
    return f"R$ {number:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


@app.context_processor
def inject_globals():
    return {"hoje": date.today(), "ano": date.today().year}


@app.route("/")
def dashboard():
    indicadores = query(
        """
        SELECT
          COALESCE(SUM(CASE WHEN YEAR(data_preenchimento)=YEAR(CURDATE())
                            AND MONTH(data_preenchimento)=MONTH(CURDATE()) THEN valor_total END), 0) total_mes,
          SUM(status_compra='aguardando autorizacao') aguardando,
          SUM(data_entrega < CURDATE() AND status_compra NOT IN ('recebida','cancelada')) atrasadas
        FROM ordens_compra
        """,
        one=True,
    )
    categorias = query(
        """SELECT p.categoria, COALESCE(SUM(o.valor_total),0) total
           FROM produtos p LEFT JOIN ordens_compra o ON o.produto_id=p.id
           GROUP BY p.categoria ORDER BY total DESC LIMIT 6"""
    )
    ultimas = query(
        """SELECT o.*, f.nome fornecedor, p.descricao produto
           FROM ordens_compra o JOIN fornecedores f ON f.id=o.fornecedor_id
           JOIN produtos p ON p.id=o.produto_id ORDER BY o.id DESC LIMIT 8"""
    )
    return render_template("dashboard.html", indicadores=indicadores, categorias=categorias, ultimas=ultimas)


def registry_usage(module, name):
    if module == "categorias":
        usages = []
        product_count = query("SELECT COUNT(*) total FROM produtos WHERE categoria=%s", (name,), one=True)["total"]
        ensure_monthly_budget_schema()
        budget_count = query(
            """
            SELECT COUNT(*) total
            FROM orcamentos o
            JOIN cadastro_categorias c ON c.id=o.categoria_id
            WHERE c.nome=%s
            """,
            (name,),
            one=True,
        )["total"]
        if product_count:
            usages.append(f"Produtos ({product_count})")
        if budget_count:
            usages.append(f"Orcamentos ({budget_count})")
        return usages
    if module == "pagamentos":
        count = query("SELECT COUNT(*) total FROM ordens_compra WHERE metodo_pagamento=%s", (name,), one=True)["total"]
        return [f"Ordens de compra ({count})"] if count else []
    return []


def ensure_monthly_budget_schema():
    execute(
        """
        CREATE TABLE IF NOT EXISTS orcamentos (
          id INT AUTO_INCREMENT PRIMARY KEY,
          mes TINYINT NOT NULL,
          ano SMALLINT NOT NULL,
          categoria_id INT NULL,
          categoria VARCHAR(80) NULL,
          valor_orcamento DECIMAL(14,2) NULL,
          orcamento_previsto DECIMAL(14,2) NULL,
          data_importacao TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
          usuario_importacao VARCHAR(120) NULL
        )
        """
    )
    if not column_exists("orcamentos", "categoria_id"):
        execute("ALTER TABLE orcamentos ADD COLUMN categoria_id INT NULL AFTER ano")
    if not column_exists("orcamentos", "valor_orcamento"):
        execute("ALTER TABLE orcamentos ADD COLUMN valor_orcamento DECIMAL(14,2) NULL AFTER categoria_id")
    if not column_exists("orcamentos", "data_importacao"):
        execute("ALTER TABLE orcamentos ADD COLUMN data_importacao TIMESTAMP DEFAULT CURRENT_TIMESTAMP AFTER valor_orcamento")
    if not column_exists("orcamentos", "usuario_importacao"):
        execute("ALTER TABLE orcamentos ADD COLUMN usuario_importacao VARCHAR(120) NULL AFTER data_importacao")
    if column_exists("orcamentos", "categoria"):
        execute("ALTER TABLE orcamentos MODIFY categoria VARCHAR(80) NULL")
    execute(
        """
        UPDATE orcamentos o
        JOIN cadastro_categorias c ON LOWER(c.nome)=LOWER(o.categoria)
        SET o.categoria_id=c.id
        WHERE o.categoria_id IS NULL AND o.categoria IS NOT NULL
        """
    )
    execute("UPDATE orcamentos SET valor_orcamento=orcamento_previsto WHERE valor_orcamento IS NULL AND orcamento_previsto IS NOT NULL")
    execute("UPDATE orcamentos SET orcamento_previsto=valor_orcamento WHERE orcamento_previsto IS NULL AND valor_orcamento IS NOT NULL")
    unique_indexes = query(
        """
        SELECT INDEX_NAME
        FROM information_schema.STATISTICS
        WHERE TABLE_SCHEMA=%s AND TABLE_NAME='orcamentos' AND NON_UNIQUE=0 AND INDEX_NAME<>'PRIMARY'
        GROUP BY INDEX_NAME
        """,
        (app.config["MYSQL_DATABASE"],),
    )
    for index in unique_indexes:
        index_name = index["INDEX_NAME"].replace("`", "``")
        execute(f"ALTER TABLE orcamentos DROP INDEX `{index_name}`")
    index_exists = query(
        """
        SELECT INDEX_NAME
        FROM information_schema.STATISTICS
        WHERE TABLE_SCHEMA=%s AND TABLE_NAME='orcamentos' AND INDEX_NAME='uk_orcamento_categoria_competencia'
        """,
        (app.config["MYSQL_DATABASE"],),
        one=True,
    )
    if not index_exists:
        execute("ALTER TABLE orcamentos ADD UNIQUE KEY uk_orcamento_categoria_competencia (categoria_id,ano,mes)")


def budget_spent_sql():
    return """
        SELECT
          c.id categoria_id,
          YEAR(n.data_entrada) ano,
          MONTH(n.data_entrada) mes,
          COALESCE(SUM(i.valor_total),0) total_gasto
        FROM nfe_importacao_itens i
        JOIN nfe_importacoes n ON n.id=i.nfe_importacao_id
        JOIN produtos p ON p.id=i.produto_id
        JOIN cadastro_categorias c ON LOWER(c.nome)=LOWER(p.categoria)
        GROUP BY c.id,YEAR(n.data_entrada),MONTH(n.data_entrada)
    """


def budget_status(valor_orcamento, valor_gasto):
    budget = decimal_value(valor_orcamento)
    spent = decimal_value(valor_gasto)
    if budget == 0 and spent == 0:
        return "SEM ORCAMENTO"
    if (budget == 0 and spent > 0) or spent > budget:
        return "ORCAMENTO ESTOURADO"
    return "DENTRO DO ORCAMENTO"


def build_budget_context():
    ensure_auxiliary_registries()
    ensure_product_schema()
    ensure_nfe_schema()
    ensure_monthly_budget_schema()

    mes = request.args.get("mes") or str(date.today().month)
    ano_filter = request.args.get("ano") or str(date.today().year)
    categoria_id = request.args.get("categoria_id") or ""
    month = int(mes)
    year = int(ano_filter)
    params = [year, month, year, month]
    filters = ["(b.id IS NOT NULL OR COALESCE(g.total_gasto,0) > 0)"]
    if categoria_id:
        filters.append("c.id=%s")
        params.append(int(categoria_id))
    where = " AND ".join(filters)
    rows = query(
        f"""
        SELECT
          b.id,
          %s mes,
          %s ano,
          c.id categoria_id,
          COALESCE(b.valor_orcamento,0) valor_orcamento,
          b.data_importacao,
          b.usuario_importacao,
          c.nome categoria,
          COALESCE(g.total_gasto,0) valor_gasto,
          (COALESCE(b.valor_orcamento,0)-COALESCE(g.total_gasto,0)) saldo,
          CASE
            WHEN COALESCE(b.valor_orcamento,0) > 0 THEN (COALESCE(g.total_gasto,0)/b.valor_orcamento)*100
            WHEN COALESCE(g.total_gasto,0) > 0 THEN 100
            ELSE 0
          END percentual
        FROM cadastro_categorias c
        LEFT JOIN orcamentos b ON b.categoria_id=c.id AND b.ano=%s AND b.mes=%s
        LEFT JOIN ({budget_spent_sql()}) g
          ON g.categoria_id=c.id AND g.ano=%s AND g.mes=%s
        WHERE {where}
        ORDER BY c.nome
        """,
        (month, year, *tuple(params)),
    )
    for row in rows:
        row["status"] = budget_status(row["valor_orcamento"], row["valor_gasto"])
    total_budget = sum(decimal_value(row["valor_orcamento"]) for row in rows)
    total_spent = sum(decimal_value(row["valor_gasto"]) for row in rows)
    summary = {
        "orcamento": total_budget,
        "gasto": total_spent,
        "saldo": total_budget - total_spent,
        "percentual": (total_spent / total_budget * 100) if total_budget else (Decimal("100") if total_spent > 0 else Decimal("0")),
        "status": budget_status(total_budget, total_spent),
    }
    return {
        "budget_rows": rows,
        "budget_summary": summary,
        "budget_filters": {"mes": int(mes), "ano": int(ano_filter), "categoria_id": categoria_id},
        "budget_categories": query("SELECT id,nome FROM cadastro_categorias WHERE status='ativo' ORDER BY nome"),
        "budget_years": query("SELECT DISTINCT ano FROM orcamentos ORDER BY ano DESC"),
    }


@app.route("/cadastros", methods=["GET", "POST"])
def cadastros():
    ensure_auxiliary_registries()
    ensure_monthly_budget_schema()
    module = request.args.get("modulo", "categorias")
    if module not in ("categorias", "pagamentos", "orcamentos"):
        module = "categorias"
    if module == "orcamentos":
        context = build_budget_context()
        counts = {
            "categorias": query("SELECT COUNT(*) total FROM cadastro_categorias", one=True)["total"],
            "pagamentos": query("SELECT COUNT(*) total FROM cadastro_tipos_pagamento", one=True)["total"],
            "orcamentos": query("SELECT COUNT(*) total FROM orcamentos", one=True)["total"],
        }
        return render_template("cadastros.html", module=module, counts=counts, **context)
    configs = {
        "categorias": {
            "table": "cadastro_categorias",
            "title": "Categorias de Produto",
            "field": "Nome da categoria",
            "modal": "categoriaForm",
        },
        "pagamentos": {
            "table": "cadastro_tipos_pagamento",
            "title": "Tipos de Pagamento",
            "field": "Nome do tipo de pagamento",
            "modal": "pagamentoForm",
        },
    }
    config = configs[module]

    if request.method == "POST":
        form = request.form
        form_module = form.get("modulo", module)
        config = configs.get(form_module, config)
        item_id = form.get("id")
        nome = (form.get("nome") or "").strip()
        status = form.get("status") or "ativo"
        if not nome:
            flash("Informe o nome do cadastro.", "danger")
            return redirect(url_for("cadastros", modulo=form_module))
        duplicate = query(
            f"SELECT id FROM {config['table']} WHERE LOWER(nome)=LOWER(%s) AND (%s='' OR id<>%s)",
            (nome, item_id or "", item_id or 0),
            one=True,
        )
        if duplicate:
            flash("Ja existe um cadastro com este nome.", "danger")
            return redirect(url_for("cadastros", modulo=form_module))
        if item_id:
            old = query(f"SELECT nome FROM {config['table']} WHERE id=%s", (item_id,), one=True)
            usages = registry_usage(form_module, old["nome"]) if old and old["nome"] != nome else []
            if usages:
                flash(f"Nao e possivel renomear. Cadastro em uso em: {', '.join(usages)}.", "danger")
                return redirect(url_for("cadastros", modulo=form_module))
            execute(f"UPDATE {config['table']} SET nome=%s,status=%s WHERE id=%s", (nome, status, item_id))
        else:
            execute(f"INSERT INTO {config['table']} (nome,status) VALUES (%s,%s)", (nome, status))
        flash("Cadastro salvo com sucesso.", "success")
        return redirect(url_for("cadastros", modulo=form_module))

    search = (request.args.get("q") or "").strip()
    sort = request.args.get("sort", "nome")
    direction = request.args.get("dir", "asc")
    page = max(int(request.args.get("page", 1) or 1), 1)
    per_page = 10
    allowed_sort = {"nome", "status", "criado_em", "atualizado_em"}
    if sort not in allowed_sort:
        sort = "nome"
    direction = "desc" if direction == "desc" else "asc"
    where = "WHERE nome LIKE %s" if search else ""
    params = (f"%{search}%",) if search else ()
    total = query(f"SELECT COUNT(*) total FROM {config['table']} {where}", params, one=True)["total"]
    offset = (page - 1) * per_page
    rows = query(
        f"SELECT * FROM {config['table']} {where} ORDER BY {sort} {direction} LIMIT %s OFFSET %s",
        params + (per_page, offset),
    )
    total_pages = max((total + per_page - 1) // per_page, 1)
    counts = {
        "categorias": query("SELECT COUNT(*) total FROM cadastro_categorias", one=True)["total"],
        "pagamentos": query("SELECT COUNT(*) total FROM cadastro_tipos_pagamento", one=True)["total"],
        "orcamentos": query("SELECT COUNT(*) total FROM orcamentos", one=True)["total"],
    }
    return render_template(
        "cadastros.html",
        module=module,
        config=config,
        rows=rows,
        search=search,
        sort=sort,
        direction=direction,
        page=page,
        total_pages=total_pages,
        total=total,
        counts=counts,
    )


@app.post("/cadastros/<module>/<int:item_id>/excluir")
def excluir_cadastro(module, item_id):
    ensure_auxiliary_registries()
    configs = {
        "categorias": "cadastro_categorias",
        "pagamentos": "cadastro_tipos_pagamento",
    }
    table = configs.get(module)
    if not table:
        flash("Cadastro invalido.", "danger")
        return redirect(url_for("cadastros"))
    item = query(f"SELECT nome FROM {table} WHERE id=%s", (item_id,), one=True)
    if not item:
        flash("Cadastro nao encontrado.", "danger")
        return redirect(url_for("cadastros", modulo=module))
    usages = registry_usage(module, item["nome"])
    if usages:
        flash(f"Nao e possivel excluir. Cadastro em uso em: {', '.join(usages)}.", "danger")
        return redirect(url_for("cadastros", modulo=module))
    execute(f"DELETE FROM {table} WHERE id=%s", (item_id,))
    flash("Cadastro excluido.", "success")
    return redirect(url_for("cadastros", modulo=module))


@app.post("/cadastros/orcamentos/salvar")
def salvar_orcamento_mensal():
    ensure_monthly_budget_schema()
    form = request.form
    budget_id = form.get("id")
    mes = int(form.get("mes") or 0)
    ano_budget = int(form.get("ano") or 0)
    categoria_id = int(form.get("categoria_id") or 0)
    valor = money_value(form.get("valor_orcamento"))
    if mes < 1 or mes > 12 or ano_budget <= 0 or not categoria_id or valor < 0:
        flash("Informe mes, ano, categoria e valor validos.", "danger")
        return redirect(url_for("cadastros", modulo="orcamentos"))
    categoria = query("SELECT id FROM cadastro_categorias WHERE id=%s", (categoria_id,), one=True)
    if not categoria:
        flash("Categoria nao encontrada.", "danger")
        return redirect(url_for("cadastros", modulo="orcamentos"))
    duplicate = query(
        """
        SELECT id FROM orcamentos
        WHERE categoria_id=%s AND ano=%s AND mes=%s AND (%s='' OR id<>%s)
        """,
        (categoria_id, ano_budget, mes, budget_id or "", budget_id or 0),
        one=True,
    )
    if duplicate:
        execute(
            "UPDATE orcamentos SET valor_orcamento=%s,orcamento_previsto=%s,data_importacao=CURRENT_TIMESTAMP WHERE id=%s",
            (valor, valor, duplicate["id"]),
        )
        flash("Orcamento existente atualizado.", "success")
    elif budget_id:
        execute(
            """
            UPDATE orcamentos
            SET mes=%s,ano=%s,categoria_id=%s,valor_orcamento=%s,orcamento_previsto=%s,data_importacao=CURRENT_TIMESTAMP
            WHERE id=%s
            """,
            (mes, ano_budget, categoria_id, valor, valor, budget_id),
        )
        flash("Orcamento atualizado.", "success")
    else:
        execute(
            """
            INSERT INTO orcamentos (mes,ano,categoria_id,valor_orcamento,orcamento_previsto,usuario_importacao)
            VALUES (%s,%s,%s,%s,%s,%s)
            """,
            (mes, ano_budget, categoria_id, valor, valor, "Sistema"),
        )
        flash("Orcamento cadastrado.", "success")
    return redirect(url_for("cadastros", modulo="orcamentos", mes=mes, ano=ano_budget))


@app.post("/cadastros/orcamentos/<int:item_id>/excluir")
def excluir_orcamento_mensal(item_id):
    ensure_monthly_budget_schema()
    row = query("SELECT mes,ano FROM orcamentos WHERE id=%s", (item_id,), one=True)
    execute("DELETE FROM orcamentos WHERE id=%s", (item_id,))
    flash("Orcamento excluido.", "success")
    if row:
        return redirect(url_for("cadastros", modulo="orcamentos", mes=row["mes"], ano=row["ano"]))
    return redirect(url_for("cadastros", modulo="orcamentos"))


@app.get("/cadastros/orcamentos/modelo")
def baixar_modelo_orcamentos():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Orcamentos Mensais"
    sheet.append(["Mês", "Ano", "Categoria", "Valor de Compra"])
    sheet.append([date.today().month, date.today().year, "Categoria Exemplo", 50000])
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name="modelo_orcamentos_mensais.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.post("/cadastros/orcamentos/importar")
def importar_orcamentos_mensais():
    ensure_auxiliary_registries()
    ensure_monthly_budget_schema()
    file = request.files.get("arquivo")
    if not file or not file.filename:
        flash("Selecione um arquivo XLSX.", "danger")
        return redirect(url_for("cadastros", modulo="orcamentos"))
    if not file.filename.lower().endswith(".xlsx"):
        flash("O arquivo deve estar no formato .xlsx.", "danger")
        return redirect(url_for("cadastros", modulo="orcamentos"))

    categories = {
        row["nome"].strip().lower(): row["id"]
        for row in query("SELECT id,nome FROM cadastro_categorias WHERE status='ativo'")
    }
    total = imported = updated = errors = 0
    error_messages = []
    try:
        workbook = load_workbook(file, data_only=True)
        sheet = workbook.active
        header = [normalize_header(cell.value) for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        required = ["mes", "ano", "categoria", "valor de compra"]
        if header != required:
            flash("O XLSX deve conter exatamente as colunas: Mes, Ano, Categoria, Valor de Compra.", "danger")
            workbook.close()
            return redirect(url_for("cadastros", modulo="orcamentos"))
        for line_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if all(value is None or str(value).strip() == "" for value in row):
                continue
            total += 1
            row_errors = []
            mes_raw, ano_raw, categoria_raw, valor_raw = row[:4]
            try:
                mes = int(mes_raw)
            except Exception:
                mes = 0
                row_errors.append("mes invalido")
            try:
                ano_budget = int(ano_raw)
            except Exception:
                ano_budget = 0
                row_errors.append("ano invalido")
            categoria_nome = import_cell_value(categoria_raw)
            categoria_id = categories.get(categoria_nome.lower())
            if not categoria_nome:
                row_errors.append("categoria obrigatoria")
            elif not categoria_id:
                row_errors.append("categoria nao cadastrada")
            try:
                valor = money_value(valor_raw)
            except Exception:
                valor = Decimal("-1")
                row_errors.append("valor invalido")
            if mes < 1 or mes > 12:
                row_errors.append("mes fora de 1 a 12")
            if ano_budget <= 0:
                row_errors.append("ano obrigatorio")
            if valor < 0:
                row_errors.append("valor negativo")
            if row_errors:
                errors += 1
                error_messages.append(f"Linha {line_number}: {', '.join(row_errors)}")
                continue
            existing = query(
                "SELECT id FROM orcamentos WHERE categoria_id=%s AND ano=%s AND mes=%s",
                (categoria_id, ano_budget, mes),
                one=True,
            )
            if existing:
                execute(
                    "UPDATE orcamentos SET valor_orcamento=%s,orcamento_previsto=%s,data_importacao=CURRENT_TIMESTAMP,usuario_importacao=%s WHERE id=%s",
                    (valor, valor, "Importacao XLSX", existing["id"]),
                )
                updated += 1
            else:
                execute(
                    """
                    INSERT INTO orcamentos (mes,ano,categoria_id,valor_orcamento,orcamento_previsto,usuario_importacao)
                    VALUES (%s,%s,%s,%s,%s,%s)
                    """,
                    (mes, ano_budget, categoria_id, valor, valor, "Importacao XLSX"),
                )
                imported += 1
        workbook.close()
        summary = (
            f"Importacao concluida. Linhas lidas: {total}. "
            f"Novos orcamentos: {imported}. Atualizados: {updated}. Erros: {errors}."
        )
        if error_messages:
            summary += " " + " | ".join(error_messages[:6])
            if len(error_messages) > 6:
                summary += f" | Mais {len(error_messages) - 6} erro(s)."
        flash(summary, "success" if imported or updated else "warning")
    except Exception as error:
        flash(f"Nao foi possivel importar o orcamento: {error}", "danger")
    return redirect(url_for("cadastros", modulo="orcamentos"))


@app.route("/fornecedores", methods=["GET", "POST"])
def fornecedores():
    ensure_supplier_schema()
    if request.method == "POST":
        form = request.form
        fornecedor_id = form.get("id")
        params = (form["nome"], form["cnpj"], form.get("contato"), form.get("telefone"), form.get("email"), form["status"])
        try:
            if fornecedor_id:
                execute("UPDATE fornecedores SET nome=%s,cnpj=%s,contato=%s,telefone=%s,email=%s,status=%s WHERE id=%s", params + (fornecedor_id,))
            else:
                execute("INSERT INTO fornecedores (nome,cnpj,contato,telefone,email,status) VALUES (%s,%s,%s,%s,%s,%s)", params)
        except mysql.connector.IntegrityError:
            flash("Ja existe um fornecedor cadastrado com este CNPJ.", "danger")
            return redirect(url_for("fornecedores"))
        flash("Fornecedor salvo com sucesso.", "success")
        return redirect(url_for("fornecedores"))
    return render_template("fornecedores.html", fornecedores=query("SELECT * FROM fornecedores ORDER BY nome"))


@app.get("/fornecedores/modelo")
def baixar_modelo_fornecedores():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Fornecedores"
    sheet.append(["Fornecedor", "CNPJ", "Contato", "Telefone", "Email"])
    sheet.append(["Fornecedor Exemplo", "00.000.000/0000-00", "João Silva", "(11) 99999-9999", "joao@exemplo.com"])
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name="modelo_fornecedores.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.post("/fornecedores/<int:item_id>/excluir")
def excluir_fornecedor(item_id):
    try:
        execute("DELETE FROM fornecedores WHERE id=%s", (item_id,))
        flash("Fornecedor excluido.", "success")
    except mysql.connector.Error:
        flash("Fornecedor possui ordens vinculadas e nao pode ser excluido.", "danger")
    return redirect(url_for("fornecedores"))


@app.post("/fornecedores/importar")
def importar_fornecedores():
    ensure_supplier_schema()
    file = request.files.get("arquivo")
    if not file or not file.filename:
        flash("Selecione um arquivo XLSX para importar.", "danger")
        return redirect(url_for("fornecedores"))
    if not file.filename.lower().endswith(".xlsx"):
        flash("Envie um arquivo no formato .xlsx.", "danger")
        return redirect(url_for("fornecedores"))

    try:
        workbook = load_workbook(file, read_only=True, data_only=True)
        sheet = workbook.active
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        expected_headers = ["Fornecedor", "CNPJ", "Contato", "Telefone", "Email"]
        headers = [import_cell_value(value) for value in (header_row or []) if import_cell_value(value)]
        if headers != expected_headers:
            workbook.close()
            flash("O arquivo deve conter exatamente as colunas: Fornecedor, CNPJ, Contato, Telefone, Email.", "danger")
            return redirect(url_for("fornecedores"))

        total = imported = ignored = errors = 0
        error_messages = []
        seen_cnpjs = set()
        seen_names = set()
        for row in sheet.iter_rows(min_row=2, values_only=True):
            values = [import_cell_value(value) for value in row[:5]]
            if not any(values):
                continue

            total += 1
            nome, cnpj_raw, contato, telefone, email = values
            cnpj = normalize_cnpj(cnpj_raw)
            row_errors = []

            if not nome:
                row_errors.append("fornecedor sem nome")
            if cnpj and not is_valid_cnpj(cnpj):
                row_errors.append("CNPJ invalido")

            name_key = nome.strip().lower()
            cnpj_key = re.sub(r"\D", "", cnpj)
            if cnpj_key:
                if cnpj_key in seen_cnpjs or query("SELECT id FROM fornecedores WHERE REPLACE(REPLACE(REPLACE(cnpj,'.',''),'/',''),'-','')=%s", (cnpj_key,), one=True):
                    row_errors.append("CNPJ duplicado")
            elif name_key and (name_key in seen_names or query("SELECT id FROM fornecedores WHERE LOWER(nome)=%s", (name_key,), one=True)):
                row_errors.append("fornecedor duplicado por nome")

            if row_errors:
                errors += len(row_errors)
                ignored += 1
                error_messages.append(f"Linha {total + 1}: {', '.join(row_errors)}")
                continue

            execute(
                "INSERT INTO fornecedores (nome,cnpj,contato,telefone,email,status) VALUES (%s,%s,%s,%s,%s,%s)",
                (nome, cnpj or None, contato or None, telefone or None, email or None, "ativo"),
            )
            imported += 1
            if cnpj_key:
                seen_cnpjs.add(cnpj_key)
            else:
                seen_names.add(name_key)

        workbook.close()
        summary = (
            f"Importacao concluida. Total de linhas lidas: {total}. "
            f"Fornecedores importados: {imported}. Fornecedores ignorados: {ignored}. "
            f"Erros encontrados: {errors}."
        )
        if error_messages:
            summary += " " + " | ".join(error_messages[:5])
            if len(error_messages) > 5:
                summary += f" | Mais {len(error_messages) - 5} erro(s)."
        flash(summary, "success" if imported else "warning")
    except Exception as error:
        flash(f"Nao foi possivel importar o arquivo: {error}", "danger")
    return redirect(url_for("fornecedores"))


@app.route("/produtos", methods=["GET", "POST"])
def produtos():
    ensure_auxiliary_registries()
    ensure_product_schema()
    if request.method == "POST":
        form = request.form
        product_id = form.get("id")
        estoque_seguranca = decimal_value(form.get("estoque_seguranca"))
        if estoque_seguranca <= 0:
            flash("Estoque de seguranca deve ser um numero positivo.", "danger")
            return redirect(url_for("produtos"))
        quantidade_por_unidade = decimal_value(form.get("quantidade_por_unidade_compra") or 1)
        if quantidade_por_unidade <= 0:
            flash("Quantidade por unidade de compra deve ser um numero positivo.", "danger")
            return redirect(url_for("produtos"))
        codigo = form["codigo"].strip()
        params = (
            codigo,
            form["descricao"],
            form["categoria"],
            form["unidade"],
            quantidade_por_unidade,
            form.get("unidade_base") or form["unidade"],
            form["fornecedor_id"],
            estoque_seguranca,
            form["status"],
        )
        try:
            if product_id:
                execute("UPDATE produtos SET codigo=%s,descricao=%s,categoria=%s,unidade=%s,quantidade_por_unidade_compra=%s,unidade_base=%s,fornecedor_id=%s,estoque_seguranca=%s,status=%s WHERE id=%s", params + (product_id,))
            else:
                execute("INSERT INTO produtos (codigo,descricao,categoria,unidade,quantidade_por_unidade_compra,unidade_base,fornecedor_id,estoque_seguranca,status) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)", params)
        except mysql.connector.IntegrityError:
            flash("Nao foi possivel salvar o produto. Verifique os campos vinculados.", "danger")
            return redirect(url_for("produtos"))
        flash("Produto salvo com sucesso.", "success")
        return redirect(url_for("produtos"))
    produtos_rows = query(
        """
        SELECT p.*, f.nome fornecedor_nome, u.nome unidade_nome,
          (SELECT h.valor_unitario FROM historico_custos h WHERE h.produto_id=p.id ORDER BY h.data_entrada DESC,h.id DESC LIMIT 1) ultimo_preco,
          (SELECT h.valor_unitario_base FROM historico_custos h WHERE h.produto_id=p.id ORDER BY h.data_entrada DESC,h.id DESC LIMIT 1) ultimo_preco_base,
          (SELECT h.data_entrada FROM historico_custos h WHERE h.produto_id=p.id ORDER BY h.data_entrada DESC,h.id DESC LIMIT 1) ultima_compra,
          (SELECT h.fornecedor FROM historico_custos h WHERE h.produto_id=p.id ORDER BY h.data_entrada DESC,h.id DESC LIMIT 1) ultimo_fornecedor,
          (SELECT AVG(h.valor_unitario) FROM historico_custos h WHERE h.produto_id=p.id) preco_medio_historico,
          (SELECT AVG(h.valor_unitario_base) FROM historico_custos h WHERE h.produto_id=p.id) preco_medio_base_historico
        FROM produtos p
        LEFT JOIN fornecedores f ON f.id=p.fornecedor_id
        LEFT JOIN unidades_medida u ON u.codigo=p.unidade
        ORDER BY p.descricao
        """
    )
    categorias = query("SELECT nome categoria FROM cadastro_categorias WHERE status='ativo' ORDER BY nome")
    unidades = query("SELECT codigo,nome FROM unidades_medida WHERE status='ativo' ORDER BY nome")
    fornecedores_ativos = query("SELECT id,nome FROM fornecedores WHERE status='ativo' ORDER BY nome")
    return render_template(
        "produtos.html",
        produtos=produtos_rows,
        categorias=categorias,
        unidades=unidades,
        fornecedores=fornecedores_ativos,
    )


@app.post("/produtos/<int:item_id>/excluir")
def excluir_produto(item_id):
    try:
        execute("DELETE FROM produtos WHERE id=%s", (item_id,))
        flash("Produto excluido.", "success")
    except mysql.connector.Error:
        flash("Produto possui ordens vinculadas e nao pode ser excluido.", "danger")
    return redirect(url_for("produtos"))


@app.route("/ordens")
def ordens():
    rows = query("""SELECT o.*, f.nome fornecedor, p.descricao produto,p.unidade,p.quantidade_por_unidade_compra,p.unidade_base FROM ordens_compra o
                    JOIN fornecedores f ON f.id=o.fornecedor_id JOIN produtos p ON p.id=o.produto_id
                    ORDER BY o.id DESC""")
    return render_template("ordens.html", ordens=rows)


@app.route("/ordens/nova", methods=["GET", "POST"])
@app.route("/ordens/<int:item_id>/editar", methods=["GET", "POST"])
def nova_ordem(item_id=None):
    ensure_auxiliary_registries()
    ensure_order_schema()
    ordem = query("SELECT * FROM ordens_compra WHERE id=%s", (item_id,), one=True) if item_id else None
    if request.method == "POST":
        form = request.form
        quantidade = decimal_value(form["quantidade"])
        preco = decimal_value(form["preco_negociado"])
        frete = decimal_value(form.get("frete"))
        parcelas = int(form.get("parcelas") or 0)
        prazos = [int(value) for value in form.getlist("prazos_parcelas[]") if str(value).strip() != ""]
        if quantidade <= 0 or preco <= 0 or parcelas <= 0 or len(prazos) != parcelas:
            flash("Informe quantidade, preco, quantidade de parcelas e prazo de cada parcela.", "danger")
            return redirect(request.url)
        product_supplier = query(
            "SELECT id,categoria FROM produtos WHERE id=%s AND fornecedor_id=%s",
            (form["produto_id"], form["fornecedor_id"]),
            one=True,
        )
        if not product_supplier:
            flash("Este produto nao pertence ao fornecedor selecionado.", "danger")
            return redirect(request.url)
        total = quantidade * preco + frete
        status_compra = form.get("status_compra") if item_id else "aguardando autorizacao"
        if status_compra in ("aprovada", "recebida"):
            status_compra = ordem["status_compra"] if ordem else "aguardando autorizacao"
        params = (
            form["data_preenchimento"],
            product_supplier["categoria"] if product_supplier else "Produto",
            form["fornecedor_id"],
            form["produto_id"],
            quantidade,
            preco,
            frete,
            total,
            form["data_entrega"],
            form["metodo_pagamento"],
            parcelas,
            ",".join(str(value) for value in prazos),
            prazos[0],
            form.get("nota_fiscal") or None,
            status_compra,
        )
        if item_id:
            execute("""UPDATE ordens_compra SET data_preenchimento=%s,tipo_material=%s,fornecedor_id=%s,produto_id=%s,
                       quantidade=%s,preco_negociado=%s,frete=%s,valor_total=%s,data_entrega=%s,metodo_pagamento=%s,
                       parcelas=%s,prazos_parcelas=%s,prazo_dias=%s,nota_fiscal=%s,status_compra=%s WHERE id=%s""", params + (item_id,))
        else:
            connection = get_db()
            cursor = connection.cursor()
            try:
                cursor.execute("SELECT COALESCE(MAX(id),0)+1 FROM ordens_compra")
                next_id = cursor.fetchone()[0]
                numero = f"OC-{date.today().year}-{next_id:05d}"
                cursor.execute("""INSERT INTO ordens_compra
                    (numero_oc,data_preenchimento,tipo_material,fornecedor_id,produto_id,quantidade,preco_negociado,frete,
                     valor_total,data_entrega,metodo_pagamento,parcelas,prazos_parcelas,prazo_dias,nota_fiscal,status_compra)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""", (numero,) + params)
                ordem_id = cursor.lastrowid
                cursor.execute("INSERT INTO autorizacoes (ordem_id,status) VALUES (%s,'pendente')", (ordem_id,))
                cursor.execute("INSERT INTO recebimentos (ordem_id,data_prevista,status) VALUES (%s,%s,'aguardando')", (ordem_id, form["data_entrega"]))
                connection.commit()
            except Exception:
                connection.rollback()
                raise
            finally:
                cursor.close()
                connection.close()
        flash("Ordem de compra salva com sucesso.", "success")
        return redirect(url_for("ordens"))
    return render_template(
        "nova_ordem.html",
        ordem=ordem,
        fornecedores=query("SELECT id,nome FROM fornecedores WHERE status='ativo' ORDER BY nome"),
        produtos=query("SELECT id,codigo,descricao,fornecedor_id FROM produtos WHERE status='ativo' ORDER BY descricao"),
        tipos_pagamento=query("SELECT nome FROM cadastro_tipos_pagamento WHERE status='ativo' ORDER BY nome"),
    )


@app.post("/ordens/<int:item_id>/excluir")
def excluir_ordem(item_id):
    execute("DELETE FROM ordens_compra WHERE id=%s", (item_id,))
    flash("Ordem excluida.", "success")
    return redirect(url_for("ordens"))


@app.get("/api/fornecedores/<int:fornecedor_id>/produtos")
def produtos_por_fornecedor(fornecedor_id):
    rows = query(
        """
        SELECT id,codigo,descricao,unidade,categoria,quantidade_por_unidade_compra,unidade_base
        FROM produtos
        WHERE fornecedor_id=%s AND status='ativo'
        ORDER BY descricao
        """,
        (fornecedor_id,),
    )
    return jsonify([serialize_row(row) for row in rows])


@app.get("/api/produtos/<int:produto_id>/consumo_categoria")
def consumo_categoria_produto(produto_id):
    product = query("SELECT id,categoria,unidade FROM produtos WHERE id=%s", (produto_id,), one=True)
    if not product:
        return jsonify({"error": "Produto nao encontrado"}), 404
    ref_date = date.today()
    if request.args.get("data_ref"):
        try:
            ref_date = datetime.strptime(request.args.get("data_ref"), "%Y-%m-%d").date()
        except ValueError:
            ref_date = date.today()
    year = ref_date.year
    month = ref_date.month
    budget = query(
        """
        SELECT o.valor_orcamento
        FROM orcamentos o
        JOIN cadastro_categorias c ON c.id=o.categoria_id
        WHERE c.nome=%s AND o.ano=%s AND o.mes=%s
        LIMIT 1
        """,
        (product["categoria"], year, month),
        one=True,
    )
    received = query(
        """
        SELECT COALESCE(SUM(h.valor_total),0) total
        FROM historico_custos h
        JOIN produtos p ON p.id=h.produto_id
        WHERE p.categoria=%s
          AND YEAR(h.data_entrada)=%s
          AND MONTH(h.data_entrada)=%s
        """,
        (product["categoria"], year, month),
        one=True,
    )
    committed_orders = query(
        """
        SELECT COALESCE(SUM(o.valor_total),0) total
        FROM ordens_compra o
        JOIN produtos p ON p.id=o.produto_id
        WHERE p.categoria=%s
          AND YEAR(o.data_preenchimento)=%s
          AND MONTH(o.data_preenchimento)=%s
          AND o.status_compra IN ('aguardando autorizacao','aprovada','em compra')
        """,
        (product["categoria"], year, month),
        one=True,
    )
    received_quantities = query(
        """
        SELECT COALESCE(NULLIF(TRIM(p.unidade),''),'UN') unidade, COALESCE(SUM(h.quantidade),0) quantidade
        FROM historico_custos h
        JOIN produtos p ON p.id=h.produto_id
        WHERE p.categoria=%s
          AND YEAR(h.data_entrada)=%s
          AND MONTH(h.data_entrada)=%s
        GROUP BY COALESCE(NULLIF(TRIM(p.unidade),''),'UN')
        """,
        (product["categoria"], year, month),
    )
    order_quantities = query(
        """
        SELECT COALESCE(NULLIF(TRIM(p.unidade),''),'UN') unidade, COALESCE(SUM(o.quantidade),0) quantidade
        FROM ordens_compra o
        JOIN produtos p ON p.id=o.produto_id
        WHERE p.categoria=%s
          AND YEAR(o.data_preenchimento)=%s
          AND MONTH(o.data_preenchimento)=%s
          AND o.status_compra IN ('aguardando autorizacao','aprovada','em compra')
        GROUP BY COALESCE(NULLIF(TRIM(p.unidade),''),'UN')
        """,
        (product["categoria"], year, month),
    )
    quantities = {}
    for row in received_quantities + order_quantities:
        unit = row["unidade"] or "UN"
        quantities[unit] = quantities.get(unit, Decimal("0")) + decimal_value(row["quantidade"])
    received_total = decimal_value(received["total"] if received else 0)
    committed_total = decimal_value(committed_orders["total"] if committed_orders else 0)
    return jsonify(
        {
            "categoria": product["categoria"],
            "referencia": f"{month:02d}/{year}",
            "unidade": product["unidade"] or "UN",
            "tem_orcamento": bool(budget),
            "orcamento": float(decimal_value(budget["valor_orcamento"]) if budget else Decimal("0")),
            "consumido": float(received_total + committed_total),
            "consumido_recebimentos": float(received_total),
            "consumido_ordens": float(committed_total),
            "quantidades": [
                {"unidade": unit, "quantidade": float(quantity)}
                for unit, quantity in sorted(quantities.items())
            ],
        }
    )


@app.get("/api/produtos/<int:produto_id>/historico")
def historico_produto(produto_id):
    product = query("SELECT * FROM produtos WHERE id=%s", (produto_id,), one=True)
    if not product:
        return jsonify({"error": "Produto nao encontrado"}), 404
    fornecedor_id = request.args.get("fornecedor_id")
    supplier = query("SELECT nome FROM fornecedores WHERE id=%s", (fornecedor_id,), one=True) if fornecedor_id else None
    supplier_name = supplier["nome"] if supplier else None
    supplier_filter = "AND h.fornecedor=%s" if supplier_name else ""
    supplier_params = (supplier_name,) if supplier_name else ()
    summary = query(
        f"""
        SELECT
          COUNT(*) total_compras,
          COALESCE(SUM(h.quantidade),0) quantidade_total,
          COALESCE(SUM(h.quantidade_total_base),0) quantidade_total_base,
          COALESCE(SUM(h.valor_total),0) valor_total_comprado,
          AVG(h.valor_unitario) preco_medio,
          AVG(h.valor_unitario_base) preco_medio_base,
          MIN(h.valor_unitario) menor_preco,
          MAX(h.valor_unitario) maior_preco,
          CASE WHEN SUM(h.quantidade) > 0 THEN SUM(h.valor_total)/SUM(h.quantidade) ELSE 0 END custo_medio_ponderado,
          CASE WHEN SUM(h.quantidade_total_base) > 0 THEN SUM(h.valor_total)/SUM(h.quantidade_total_base) ELSE 0 END custo_medio_base_ponderado
        FROM historico_custos h
        WHERE h.produto_id=%s {supplier_filter}
        """,
        (produto_id,) + supplier_params,
        one=True,
    )
    latest = query(
        f"""
        SELECT h.*
        FROM historico_custos h
        WHERE h.produto_id=%s {supplier_filter}
        ORDER BY h.data_entrada DESC,h.id DESC
        LIMIT 1
        """,
        (produto_id,) + supplier_params,
        one=True,
    )
    purchases = query(
        """
        SELECT h.*,n.status_recebimento
        FROM historico_custos h
        LEFT JOIN nfe_importacoes n ON n.id=h.recebimento_id
        WHERE h.produto_id=%s
        ORDER BY h.data_entrada DESC,h.id DESC
        LIMIT 8
        """,
        (produto_id,),
    )
    chart = query(
        """
        SELECT data_entrada,valor_unitario,valor_unitario_base,fornecedor,documento
        FROM historico_custos
        WHERE produto_id=%s
        ORDER BY data_entrada,id
        LIMIT 30
        """,
        (produto_id,),
    )
    best_supplier = query(
        """
        SELECT fornecedor,AVG(valor_unitario) preco_medio,AVG(valor_unitario_base) preco_medio_base,COUNT(*) total_compras
        FROM historico_custos
        WHERE produto_id=%s
        GROUP BY fornecedor
        ORDER BY preco_medio ASC,total_compras DESC
        LIMIT 1
        """,
        (produto_id,),
        one=True,
    )
    general = query(
        """
        SELECT AVG(valor_unitario) preco_medio_geral,AVG(valor_unitario_base) preco_medio_base_geral
        FROM historico_custos
        WHERE produto_id=%s
        """,
        (produto_id,),
        one=True,
    )
    avg_price = decimal_value(summary["preco_medio"]) if summary and summary["preco_medio"] is not None else Decimal("0")
    last_price = decimal_value(latest["valor_unitario"]) if latest else Decimal("0")
    alert_percent = ((last_price - avg_price) / avg_price * 100) if avg_price and last_price > avg_price else Decimal("0")
    return jsonify(
        {
            "produto": serialize_row(product),
            "resumo": serialize_row(summary),
            "ultima_compra": serialize_row(latest) if latest else None,
            "ultimas_compras": [serialize_row(row) for row in purchases],
            "grafico": [serialize_row(row) for row in chart],
            "melhor_fornecedor": serialize_row(best_supplier) if best_supplier else None,
            "preco_medio_geral": float(decimal_value(general["preco_medio_geral"]) if general and general["preco_medio_geral"] is not None else Decimal("0")),
            "preco_medio_base_geral": float(decimal_value(general["preco_medio_base_geral"]) if general and general["preco_medio_base_geral"] is not None else Decimal("0")),
            "alerta_percentual": float(alert_percent),
        }
    )


@app.route("/historico-compras")
def historico_compras():
    ensure_nfe_schema()
    filters = request_purchase_history_filters()
    where_sql, params = purchase_history_where(filters)
    rows = query(
        f"""
        SELECT
          h.id,h.data_entrada,h.fornecedor,h.documento,h.quantidade,h.unidade_compra,
          h.quantidade_por_unidade_compra,h.unidade_base,h.quantidade_total_base,
          h.valor_unitario,h.valor_unitario_base,h.valor_total,h.origem,
          p.id produto_id,p.codigo,p.descricao produto,p.categoria,p.unidade,
          n.status_recebimento
        FROM historico_custos h
        JOIN produtos p ON p.id=h.produto_id
        LEFT JOIN nfe_importacoes n ON n.id=h.recebimento_id
        {where_sql}
        ORDER BY h.data_entrada DESC,h.id DESC
        LIMIT 500
        """,
        tuple(params),
    )
    summary = query(
        f"""
        SELECT
          COALESCE(SUM(h.valor_total),0) total_comprado,
          COALESCE(SUM(h.quantidade),0) quantidade_total,
          COALESCE(SUM(h.quantidade_total_base),0) quantidade_total_base,
          AVG(h.valor_unitario) preco_medio_geral,
          AVG(h.valor_unitario_base) preco_medio_base_geral,
          COUNT(*) total_registros
        FROM historico_custos h
        JOIN produtos p ON p.id=h.produto_id
        LEFT JOIN nfe_importacoes n ON n.id=h.recebimento_id
        {where_sql}
        """,
        tuple(params),
        one=True,
    )
    supplier_most = query(
        f"""
        SELECT h.fornecedor,COUNT(*) total
        FROM historico_custos h
        JOIN produtos p ON p.id=h.produto_id
        LEFT JOIN nfe_importacoes n ON n.id=h.recebimento_id
        {where_sql}
        GROUP BY h.fornecedor
        ORDER BY total DESC
        LIMIT 1
        """,
        tuple(params),
        one=True,
    )
    product_increase = query(
        f"""
        SELECT p.codigo,p.descricao produto,
               MAX(h.valor_unitario)-AVG(h.valor_unitario) aumento
        FROM historico_custos h
        JOIN produtos p ON p.id=h.produto_id
        LEFT JOIN nfe_importacoes n ON n.id=h.recebimento_id
        {where_sql}
        GROUP BY p.id
        HAVING COUNT(*) > 1
        ORDER BY aumento DESC
        LIMIT 1
        """,
        tuple(params),
        one=True,
    )
    grouped = {}
    for entry in rows:
        code = entry["codigo"] or "Nao registrado"
        group = grouped.setdefault(
            code,
            {
                "codigo": code,
                "produto": entry["produto"],
                "categoria": entry["categoria"],
                "unidade": entry["unidade"],
                "unidade_base": entry["unidade_base"] or entry["unidade"],
                "quantidade_por_unidade_compra": decimal_value(entry["quantidade_por_unidade_compra"] or 1),
                "quantidade_total": Decimal("0"),
                "quantidades_compra": {},
                "quantidade_total_base": Decimal("0"),
                "valor_total_comprado": Decimal("0"),
                "entradas_count": 0,
                "ultimo_preco": decimal_value(entry["valor_unitario"]),
                "ultimo_preco_base": decimal_value(entry["valor_unitario_base"]),
                "ultima_compra": entry["data_entrada"],
                "ultimo_fornecedor": entry["fornecedor"],
                "entries": [],
                "suppliers": {},
            },
        )
        qty = decimal_value(entry["quantidade"])
        base_qty = decimal_value(entry["quantidade_total_base"])
        value = decimal_value(entry["valor_total"])
        purchase_unit = entry["unidade_compra"] or entry["unidade"] or "UN"
        group["quantidade_total"] += qty
        group["quantidades_compra"][purchase_unit] = group["quantidades_compra"].get(purchase_unit, Decimal("0")) + qty
        group["quantidade_total_base"] += base_qty
        group["valor_total_comprado"] += value
        group["entradas_count"] += 1
        group["entries"].append(entry)
        supplier = group["suppliers"].setdefault(
            entry["fornecedor"],
            {
                "fornecedor": entry["fornecedor"],
                "quantidade_total": Decimal("0"),
                "quantidades_compra": {},
                "quantidade_total_base": Decimal("0"),
                "valor_total_comprado": Decimal("0"),
                "entradas_count": 0,
                "ultimo_preco": decimal_value(entry["valor_unitario"]),
                "ultimo_preco_base": decimal_value(entry["valor_unitario_base"]),
                "ultima_compra": entry["data_entrada"],
                "menor_preco": decimal_value(entry["valor_unitario"]),
            },
        )
        supplier["quantidade_total"] += qty
        supplier["quantidades_compra"][purchase_unit] = supplier["quantidades_compra"].get(purchase_unit, Decimal("0")) + qty
        supplier["quantidade_total_base"] += base_qty
        supplier["valor_total_comprado"] += value
        supplier["entradas_count"] += 1
        supplier["menor_preco"] = min(supplier["menor_preco"], decimal_value(entry["valor_unitario"]))
    grouped_rows = []
    for group in grouped.values():
        group["quantidades_compra_lista"] = [
            {"unidade": unit, "quantidade": quantity}
            for unit, quantity in sorted(group["quantidades_compra"].items())
        ]
        group["preco_medio_historico"] = (
            group["valor_total_comprado"] / group["quantidade_total"]
            if group["quantidade_total"] > 0
            else Decimal("0")
        )
        group["preco_medio_base_historico"] = (
            group["valor_total_comprado"] / group["quantidade_total_base"]
            if group["quantidade_total_base"] > 0
            else Decimal("0")
        )
        avg_price = group["preco_medio_historico"]
        last_price = group["ultimo_preco"]
        group["variacao_percentual"] = ((last_price - avg_price) / avg_price * 100) if avg_price else Decimal("0")
        supplier_rows = []
        for supplier in group["suppliers"].values():
            supplier["quantidades_compra_lista"] = [
                {"unidade": unit, "quantidade": quantity}
                for unit, quantity in sorted(supplier["quantidades_compra"].items())
            ]
            supplier["preco_medio"] = (
                supplier["valor_total_comprado"] / supplier["quantidade_total"]
                if supplier["quantidade_total"] > 0
                else Decimal("0")
            )
            supplier["preco_medio_base"] = (
                supplier["valor_total_comprado"] / supplier["quantidade_total_base"]
                if supplier["quantidade_total_base"] > 0
                else Decimal("0")
            )
            supplier_rows.append(supplier)
        group["supplier_summary"] = sorted(supplier_rows, key=lambda item: item["preco_medio"])
        grouped_rows.append(group)
    grouped_rows.sort(key=lambda item: (item["ultima_compra"], item["codigo"]), reverse=True)
    product_increase = None
    if grouped_rows:
        product_increase = max(grouped_rows, key=lambda item: item["variacao_percentual"])
        if product_increase["variacao_percentual"] <= 0:
            product_increase = None
    summary["total_codigos"] = len(grouped_rows)
    categorias = query("SELECT nome FROM cadastro_categorias WHERE status='ativo' ORDER BY nome")
    unidades = query("SELECT codigo,nome FROM unidades_medida WHERE status='ativo' ORDER BY nome")
    filter_options = {
        "codigos": query(
            """
            SELECT DISTINCT p.codigo valor
            FROM historico_custos h
            JOIN produtos p ON p.id=h.produto_id
            WHERE TRIM(COALESCE(p.codigo,'')) <> ''
            ORDER BY p.codigo
            """
        ),
        "produtos": query(
            """
            SELECT DISTINCT p.descricao valor
            FROM historico_custos h
            JOIN produtos p ON p.id=h.produto_id
            WHERE TRIM(COALESCE(p.descricao,'')) <> ''
            ORDER BY p.descricao
            """
        ),
        "fornecedores": query(
            """
            SELECT DISTINCT h.fornecedor valor
            FROM historico_custos h
            WHERE TRIM(COALESCE(h.fornecedor,'')) <> ''
            ORDER BY h.fornecedor
            """
        ),
        "documentos": query(
            """
            SELECT DISTINCT h.documento valor
            FROM historico_custos h
            WHERE TRIM(COALESCE(h.documento,'')) <> ''
            ORDER BY h.documento
            """
        ),
        "origens": ["XML", "Manual", "Importacao XLSX"],
        "status": ["Recebido", "Recebido Parcialmente", "Pendente", "Produto Pendente de Cadastro", "Aguardando Ordem de Compra"],
    }
    return render_template(
        "historico_compras.html",
        rows=rows,
        grouped_rows=grouped_rows,
        summary=summary,
        supplier_most=supplier_most,
        product_increase=product_increase,
        filtros=filters,
        categorias=categorias,
        unidades=unidades,
        filter_options=filter_options,
    )


@app.get("/historico-compras/exportar")
def exportar_historico_compras():
    ensure_nfe_schema()
    filters = request_purchase_history_filters()
    where_sql, params = purchase_history_where(filters)
    rows = query(
        f"""
        SELECT
          h.id,h.data_entrada,h.fornecedor,h.documento,h.quantidade,h.unidade_compra,
          h.quantidade_por_unidade_compra,h.unidade_base,h.quantidade_total_base,
          h.valor_unitario,h.valor_unitario_base,h.valor_total,h.origem,
          p.id produto_id,p.codigo,p.descricao produto,p.categoria,p.unidade,
          n.status_recebimento
        FROM historico_custos h
        JOIN produtos p ON p.id=h.produto_id
        LEFT JOIN nfe_importacoes n ON n.id=h.recebimento_id
        {where_sql}
        ORDER BY h.data_entrada DESC,h.id DESC
        """,
        tuple(params),
    )
    grouped = {}
    for entry in rows:
        code = entry["codigo"] or "Nao registrado"
        group = grouped.setdefault(
            code,
            {
                "codigo": code,
                "produto": entry["produto"],
                "categoria": entry["categoria"],
                "unidade_base": entry["unidade_base"] or entry["unidade"],
                "quantidades_compra": {},
                "quantidade_total": Decimal("0"),
                "quantidade_total_base": Decimal("0"),
                "valor_total_comprado": Decimal("0"),
                "entradas_count": 0,
                "ultimo_preco": decimal_value(entry["valor_unitario"]),
                "ultimo_preco_base": decimal_value(entry["valor_unitario_base"]),
                "ultima_compra": entry["data_entrada"],
                "ultimo_fornecedor": entry["fornecedor"],
            },
        )
        qty = decimal_value(entry["quantidade"])
        base_qty = decimal_value(entry["quantidade_total_base"])
        value = decimal_value(entry["valor_total"])
        purchase_unit = entry["unidade_compra"] or entry["unidade"] or "UN"
        group["quantidade_total"] += qty
        group["quantidade_total_base"] += base_qty
        group["valor_total_comprado"] += value
        group["entradas_count"] += 1
        group["quantidades_compra"][purchase_unit] = group["quantidades_compra"].get(purchase_unit, Decimal("0")) + qty
    grouped_rows = []
    for group in grouped.values():
        group["qtd_compra"] = "\n".join(
            f"{quantity} {unit}" for unit, quantity in sorted(group["quantidades_compra"].items())
        )
        group["preco_medio_compra"] = (
            group["valor_total_comprado"] / group["quantidade_total"]
            if group["quantidade_total"] > 0
            else Decimal("0")
        )
        group["preco_medio_base"] = (
            group["valor_total_comprado"] / group["quantidade_total_base"]
            if group["quantidade_total_base"] > 0
            else Decimal("0")
        )
        grouped_rows.append(group)
    grouped_rows.sort(key=lambda item: (item["ultima_compra"], item["codigo"]), reverse=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Consolidado"
    sheet.append([
        "Codigo interno", "Produto", "Categoria", "Qtd. compra", "Qtd. base",
        "Preco medio compra", "Preco medio base", "Ultimo preco compra",
        "Ultimo preco base", "Ultima compra", "Ultimo fornecedor", "Valor total", "Entradas"
    ])
    for row in grouped_rows:
        sheet.append([
            row["codigo"],
            row["produto"],
            row["categoria"],
            row["qtd_compra"],
            float(row["quantidade_total_base"]),
            float(row["preco_medio_compra"]),
            float(row["preco_medio_base"]),
            float(row["ultimo_preco"]),
            float(row["ultimo_preco_base"]),
            row["ultima_compra"],
            row["ultimo_fornecedor"],
            float(row["valor_total_comprado"]),
            row["entradas_count"],
        ])

    detail = workbook.create_sheet("Entradas")
    detail.append([
        "Data", "Codigo interno", "Produto", "Categoria", "Fornecedor", "NF/Documento",
        "Qtd. compra", "Unidade compra", "Qtd. por unidade", "Unidade base",
        "Qtd. base", "Valor unidade compra", "Valor base", "Valor total", "Origem", "Status"
    ])
    for row in rows:
        detail.append([
            row["data_entrada"],
            row["codigo"] or "Nao registrado",
            row["produto"],
            row["categoria"],
            row["fornecedor"],
            row["documento"],
            float(decimal_value(row["quantidade"])),
            row["unidade_compra"] or row["unidade"],
            float(decimal_value(row["quantidade_por_unidade_compra"] or 1)),
            row["unidade_base"] or row["unidade"],
            float(decimal_value(row["quantidade_total_base"])),
            float(decimal_value(row["valor_unitario"])),
            float(decimal_value(row["valor_unitario_base"])),
            float(decimal_value(row["valor_total"])),
            row["origem"],
            row["status_recebimento"] or "Recebido",
        ])

    for ws in workbook.worksheets:
        ws.freeze_panes = "A2"
        for column_cells in ws.columns:
            letter = column_cells[0].column_letter
            max_length = max(len(str(cell.value or "")) for cell in column_cells[:100])
            ws.column_dimensions[letter].width = min(max(max_length + 2, 12), 42)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=f"historico_compras_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/autorizacoes", methods=["GET", "POST"])
def autorizacoes():
    if request.method == "POST":
        form = request.form
        approval_date = form.get("data_aprovacao") or None
        execute("UPDATE autorizacoes SET autorizado_por=%s,data_aprovacao=%s,status=%s,observacao=%s WHERE id=%s", (form.get("autorizado_por"), approval_date, form["status"], form.get("observacao"), form["id"]))
        status_ordem = {"aprovada": "aprovada", "reprovada": "reprovada", "pendente": "aguardando autorizacao"}[form["status"]]
        execute("UPDATE ordens_compra o JOIN autorizacoes a ON a.ordem_id=o.id SET o.status_compra=%s WHERE a.id=%s", (status_ordem, form["id"]))
        flash("Autorizacao atualizada.", "success")
        return redirect(url_for("autorizacoes"))
    rows = query("""SELECT a.*,o.numero_oc,o.valor_total,o.quantidade,p.unidade,p.quantidade_por_unidade_compra,p.unidade_base
                    FROM autorizacoes a
                    JOIN ordens_compra o ON o.id=a.ordem_id
                    JOIN produtos p ON p.id=o.produto_id
                    ORDER BY a.id DESC""")
    return render_template("autorizacoes.html", autorizacoes=rows)


@app.route("/recebimentos", methods=["GET", "POST"])
def recebimentos():
    ensure_auxiliary_registries()
    ensure_product_schema()
    ensure_nfe_schema()
    if request.method == "POST":
        form = request.form
        real = form.get("data_real") or None
        status = form["status"]
        execute("UPDATE recebimentos SET data_real=%s,status=%s WHERE id=%s", (real, status, form["id"]))
        if status == "entregue completo":
            execute("UPDATE ordens_compra o JOIN recebimentos r ON r.ordem_id=o.id SET o.status_compra='recebida' WHERE r.id=%s", (form["id"],))
        flash("Recebimento atualizado.", "success")
        return redirect(url_for("recebimentos"))
    rows = query("""SELECT r.*,o.numero_oc,
        CASE WHEN r.data_real IS NOT NULL THEN GREATEST(DATEDIFF(r.data_real,r.data_prevista),0)
             WHEN r.data_prevista<CURDATE() AND r.status NOT IN ('entregue parcial','entregue completo') THEN DATEDIFF(CURDATE(),r.data_prevista) ELSE 0 END dias_atraso,
        CASE WHEN r.data_prevista<CURDATE() AND r.data_real IS NULL AND r.status='aguardando' THEN 'atrasado' ELSE r.status END status_exibicao
        FROM recebimentos r JOIN ordens_compra o ON o.id=r.ordem_id ORDER BY r.data_prevista""")
    filters = {
        "origem": request.args.get("origem") or "",
        "fornecedor": request.args.get("fornecedor") or "",
        "produto": request.args.get("produto") or "",
        "documento": request.args.get("documento") or "",
        "status": request.args.get("status") or "",
        "data_inicio": request.args.get("data_inicio") or "",
        "data_fim": request.args.get("data_fim") or "",
    }
    refresh_receipt_statuses()
    where = []
    params = []
    if filters["origem"]:
        where.append("n.origem=%s")
        params.append(filters["origem"])
    if filters["fornecedor"]:
        where.append("n.fornecedor LIKE %s")
        params.append(f"%{filters['fornecedor']}%")
    if filters["documento"]:
        where.append("n.numero_nf LIKE %s")
        params.append(f"%{filters['documento']}%")
    if filters["status"]:
        where.append("n.status_recebimento=%s")
        params.append(filters["status"])
    if filters["data_inicio"]:
        where.append("n.data_entrada >= %s")
        params.append(filters["data_inicio"])
    if filters["data_fim"]:
        where.append("n.data_entrada <= %s")
        params.append(filters["data_fim"])
    if filters["produto"]:
        where.append(
            """
            EXISTS (
              SELECT 1 FROM nfe_importacao_itens ix
              LEFT JOIN produtos px ON px.id=ix.produto_id
              WHERE ix.nfe_importacao_id=n.id
                AND (ix.descricao LIKE %s OR px.descricao LIKE %s OR px.codigo LIKE %s)
            )
            """
        )
        term = f"%{filters['produto']}%"
        params.extend([term, term, term])
    where_sql = "WHERE " + " AND ".join(where) if where else ""
    nfe_rows = query(
        f"""
        SELECT n.*,o.numero_oc,
          (SELECT COUNT(*) FROM nfe_importacao_itens i WHERE i.nfe_importacao_id=n.id) total_itens,
          (SELECT COUNT(*) FROM nfe_importacao_itens i WHERE i.nfe_importacao_id=n.id AND i.status='pendente') itens_pendentes
        FROM nfe_importacoes n
        LEFT JOIN ordens_compra o ON o.id=n.ordem_id
        {where_sql}
        ORDER BY n.criado_em DESC, n.id DESC
        """,
        tuple(params),
    )
    produtos_ativos = query("SELECT id,codigo,descricao FROM produtos WHERE status='ativo' ORDER BY descricao")
    ordens_disponiveis = query(
        """
        SELECT o.id,o.numero_oc,o.data_entrega,o.quantidade,o.preco_negociado,o.produto_id,
               f.id fornecedor_id,f.nome fornecedor,
               p.codigo produto_codigo,p.descricao produto_descricao,p.unidade produto_unidade
        FROM ordens_compra o
        JOIN fornecedores f ON f.id=o.fornecedor_id
        JOIN produtos p ON p.id=o.produto_id
        ORDER BY o.id DESC
        """
    )
    fornecedores_ativos = query("SELECT id,nome,cnpj FROM fornecedores WHERE status='ativo' ORDER BY nome")
    unidades = query("SELECT codigo,nome FROM unidades_medida WHERE status='ativo' ORDER BY nome")
    categorias = query("SELECT nome FROM cadastro_categorias WHERE status='ativo' ORDER BY nome")
    tipos_pagamento = query("SELECT nome FROM cadastro_tipos_pagamento WHERE status='ativo' ORDER BY nome")
    manual_products = query("SELECT id,codigo,descricao,unidade,quantidade_por_unidade_compra,unidade_base,fornecedor_id FROM produtos WHERE status<>'inativo' ORDER BY descricao")
    nfe_items_by_id = {}
    for nfe in nfe_rows:
        nfe_items_by_id[nfe["id"]] = query(
            """
            SELECT i.*,p.codigo codigo_interno,p.descricao descricao_interna,p.status produto_status,
                   p.quantidade_por_unidade_compra,p.unidade_base,n.fornecedor,n.numero_nf,n.data_entrada
            FROM nfe_importacao_itens i
            JOIN nfe_importacoes n ON n.id=i.nfe_importacao_id
            LEFT JOIN produtos p ON p.id=i.produto_id
            WHERE i.nfe_importacao_id=%s
            ORDER BY i.id
            """,
            (nfe["id"],),
        )
        for item in nfe_items_by_id[nfe["id"]]:
            factor = decimal_value(item.get("quantidade_por_unidade_compra") or 1)
            if factor <= 0:
                factor = Decimal("1")
            base_qty = decimal_value(item["quantidade"]) * factor
            item["quantidade_total_base"] = base_qty
            item["valor_unitario_base"] = decimal_value(item["valor_total"]) / base_qty if base_qty > 0 else decimal_value(item["valor_unitario"])
            item["unidade_base"] = item.get("unidade_base") or item.get("unidade") or "UN"
            purchase_unit = (item.get("unidade") or "").upper()
            item["alerta_conversao"] = item.get("produto_id") and factor == Decimal("1") and purchase_unit in ("PALLET", "CX", "CAIXA", "ROLO")
    return render_template(
        "recebimentos.html",
        recebimentos=rows,
        nfe_rows=nfe_rows,
        nfe_items_by_id=nfe_items_by_id,
        produtos_ativos=produtos_ativos,
        ordens_disponiveis=ordens_disponiveis,
        fornecedores=fornecedores_ativos,
        unidades=unidades,
        categorias=categorias,
        tipos_pagamento=tipos_pagamento,
        manual_products=manual_products,
        filtros=filters,
    )


def import_nfe_file(file_storage):
    ensure_auxiliary_registries()
    ensure_product_schema()
    ensure_nfe_schema()
    data = parse_nfe_xml(file_storage)
    exists = query("SELECT id FROM nfe_importacoes WHERE chave_nfe=%s", (data["chave_nfe"],), one=True)
    if exists:
        return {"status": "duplicado", "nfe_id": exists["id"], "pending": 0, "products_created": [], "supplier_created": None}

    supplier = get_or_create_nfe_supplier(
        data["fornecedor"],
        data["fornecedor_cnpj"],
        data.get("fornecedor_telefone"),
    )
    fornecedor_id = supplier["id"]
    identified_items = []
    pending = 0
    created_products = []
    for item in data["items"]:
        product = identify_nfe_product(
            data["fornecedor"],
            data["fornecedor_cnpj"],
            item["codigo_fornecedor"],
            item["descricao"],
        )
        if not product:
            product = create_pending_product_from_nfe_item(item, fornecedor_id)
            pending += 1
            created_products.append(product)
        register_product_supplier_link(data["fornecedor"], data["fornecedor_cnpj"], item, product["id"])
        identified_items.append((item, product, product.get("created", False)))

    connection = get_db()
    cursor = connection.cursor()
    try:
        cursor.execute(
            """
            INSERT INTO nfe_importacoes
            (ordem_id,origem,fornecedor,fornecedor_cnpj,numero_nf,tipo_documento,serie,chave_nfe,data_emissao,data_entrada,valor_total,usuario_responsavel,status_recebimento,confirmado_em)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,NOW())
            """,
            (
                None,
                "XML",
                data["fornecedor"],
                data["fornecedor_cnpj"],
                data["numero_nf"],
                "Nota Fiscal",
                data["serie"],
                data["chave_nfe"],
                data["data_emissao"],
                data["data_entrada"],
                data["valor_total"],
                "Sistema",
                "Produto Pendente de Cadastro" if pending else "Aguardando Ordem de Compra",
            ),
        )
        nfe_id = cursor.lastrowid
        for item, product, product_created in identified_items:
            cursor.execute(
                """
                INSERT INTO nfe_importacao_itens
                (nfe_importacao_id,produto_id,codigo_fornecedor,descricao,ncm,cfop,cest,unidade,quantidade,valor_unitario,desconto,valor_total,ean,status)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """,
                (
                    nfe_id,
                    product["id"] if product else None,
                    item["codigo_fornecedor"],
                    item["descricao"],
                    item["ncm"],
                    item["cfop"],
                    item["cest"],
                    item["unidade"],
                    item["quantidade"],
                    item["valor_unitario"],
                    0,
                    item["valor_total"],
                    item["ean"],
                    "pendente" if product_created else "vinculado",
                ),
            )
        connection.commit()
        audit_receipt(nfe_id, "Importacao XML", f"Documento {data['numero_nf']} importado.", "Sistema", "XML")
        apply_receipt_effects(nfe_id)
        return {
            "status": "importado",
            "nfe_id": nfe_id,
            "pending": pending,
            "products_created": created_products,
            "supplier_created": supplier if supplier["created"] else None,
        }
    except Exception:
        connection.rollback()
        raise
    finally:
        cursor.close()
        connection.close()


@app.post("/recebimentos/importar_xml")
def importar_xml_nfe():
    ensure_nfe_schema()
    files = [file for file in request.files.getlist("xmls") + request.files.getlist("xml") if file and file.filename]
    if not files:
        flash("Selecione ao menos um arquivo XML de NF-e.", "danger")
        return redirect(url_for("recebimentos"))

    imported = errors = duplicates = pending_products = 0
    auto_products = []
    auto_suppliers = []
    error_messages = []
    last_nfe_id = None
    for file in files:
        if not file.filename.lower().endswith(".xml"):
            errors += 1
            error_messages.append(f"{file.filename}: formato invalido")
            continue
        try:
            result = import_nfe_file(file)
            if result["status"] == "duplicado":
                duplicates += 1
                last_nfe_id = result["nfe_id"]
            else:
                imported += 1
                pending_products += result["pending"]
                auto_products.extend(result.get("products_created", []))
                if result.get("supplier_created"):
                    auto_suppliers.append(result["supplier_created"])
                last_nfe_id = result["nfe_id"]
        except Exception as error:
            errors += 1
            error_messages.append(f"{file.filename}: {error}")

    summary = (
        f"Importacao em massa concluida. XMLs importados com sucesso: {imported}. "
        f"XMLs com erro: {errors}. Notas cadastradas: {imported}. "
        f"Produtos cadastrados automaticamente para revisao: {len(auto_products)}. "
        f"Fornecedores cadastrados automaticamente: {len(auto_suppliers)}. "
        f"Duplicadas ignoradas: {duplicates}."
    )
    if auto_suppliers:
        names = ", ".join(sorted({supplier["nome"] for supplier in auto_suppliers})[:5])
        summary += f" Fornecedores criados: {names}."
    if auto_products:
        names = ", ".join(product["descricao"] for product in auto_products[:5])
        summary += f" Produtos pendentes: {names}."
    if error_messages:
        summary += " " + " | ".join(error_messages[:5])
        if len(error_messages) > 5:
            summary += f" | Mais {len(error_messages) - 5} erro(s)."
    flash(summary, "success" if imported else "warning")
    if last_nfe_id:
        return redirect(url_for("recebimentos"))
    return redirect(url_for("recebimentos"))


@app.post("/recebimentos/manual")
def salvar_recebimento_manual():
    ensure_auxiliary_registries()
    ensure_product_schema()
    ensure_nfe_schema()
    form = request.form
    fornecedor_id = form.get("fornecedor_id")
    supplier = query("SELECT * FROM fornecedores WHERE id=%s", (fornecedor_id,), one=True)
    if not supplier:
        flash("Fornecedor obrigatorio para entrada manual.", "danger")
        return redirect(url_for("recebimentos"))
    numero_doc = (form.get("numero_nf") or "").strip() or f"MANUAL-{datetime.now().strftime('%Y%m%d%H%M%S')}"
    tipo_documento = form.get("tipo_documento") or "Sem Documento"
    data_compra = form.get("data_compra") or form.get("data_entrada") or date.today().isoformat()
    data_entrada = form.get("data_entrada") or date.today().isoformat()
    ordem_id = form.get("ordem_id") or None
    codigo_fornecedor_list = form.getlist("codigo_fornecedor[]")
    descricao_list = form.getlist("descricao[]")
    produto_id_list = form.getlist("produto_id[]")
    codigo_interno_list = form.getlist("codigo_interno[]")
    unidade_list = form.getlist("unidade[]")
    quantidade_list = form.getlist("quantidade[]")
    valor_unitario_list = form.getlist("valor_unitario[]")
    desconto_list = form.getlist("desconto[]")
    items = []
    for index, descricao in enumerate(descricao_list):
        descricao = (descricao or "").strip()
        qty = decimal_value(quantidade_list[index] if index < len(quantidade_list) else 0)
        unit = money_value(valor_unitario_list[index] if index < len(valor_unitario_list) else 0)
        desconto = money_value(desconto_list[index] if index < len(desconto_list) else 0)
        if not descricao and qty == 0:
            continue
        if not descricao or qty <= 0 or unit < 0 or desconto < 0:
            flash("Revise os produtos da entrada manual. Descricao, quantidade e valores devem ser validos.", "danger")
            return redirect(url_for("recebimentos"))
        total = (qty * unit) - desconto
        if total < 0:
            flash("O desconto nao pode ser maior que o total do item.", "danger")
            return redirect(url_for("recebimentos"))
        product = None
        product_id = produto_id_list[index] if index < len(produto_id_list) else ""
        codigo_interno = (codigo_interno_list[index] if index < len(codigo_interno_list) else "").strip()
        if product_id:
            product = query("SELECT id,codigo FROM produtos WHERE id=%s", (product_id,), one=True)
        if not product and codigo_interno:
            product = query("SELECT id,codigo FROM produtos WHERE codigo=%s AND status<>'inativo' ORDER BY id DESC LIMIT 1", (codigo_interno,), one=True)
        product_created = False
        item_data = {
            "codigo_fornecedor": (codigo_fornecedor_list[index] if index < len(codigo_fornecedor_list) else "").strip(),
            "descricao": descricao,
            "unidade": unidade_list[index] if index < len(unidade_list) else "UN",
            "quantidade": qty,
            "valor_unitario": unit,
            "desconto": desconto,
            "valor_total": total,
        }
        if not product:
            product = create_pending_product_from_nfe_item(item_data, supplier["id"])
            product_created = True
        register_product_supplier_link(supplier["nome"], supplier.get("cnpj"), item_data, product["id"])
        items.append((item_data, product, product_created))
    if not items:
        flash("Adicione ao menos um produto para salvar a entrada manual.", "danger")
        return redirect(url_for("recebimentos"))

    valor_total = sum(item[0]["valor_total"] for item in items)
    connection = get_db()
    cursor = connection.cursor()
    try:
        cursor.execute(
            """
            INSERT INTO nfe_importacoes
            (ordem_id,origem,fornecedor,fornecedor_cnpj,numero_nf,tipo_documento,serie,chave_nfe,data_emissao,data_entrada,valor_total,
             condicao_pagamento,observacoes,usuario_responsavel,status_recebimento,confirmado_em)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,NOW())
            """,
            (
                ordem_id,
                "Manual",
                supplier["nome"],
                supplier.get("cnpj") or "Nao informado",
                numero_doc,
                tipo_documento,
                "",
                f"MANUAL-{datetime.now().strftime('%Y%m%d%H%M%S%f')}",
                data_compra,
                data_entrada,
                valor_total,
                form.get("condicao_pagamento"),
                form.get("observacoes"),
                "Sistema",
                "Produto Pendente de Cadastro" if any(product_created for _, _, product_created in items) else ("Recebido" if not ordem_id else "Recebido Parcialmente"),
            ),
        )
        recebimento_id = cursor.lastrowid
        for item_data, product, product_created in items:
            cursor.execute(
                """
                INSERT INTO nfe_importacao_itens
                (nfe_importacao_id,produto_id,codigo_fornecedor,descricao,unidade,quantidade,valor_unitario,desconto,valor_total,status)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """,
                (
                    recebimento_id,
                    product["id"],
                    item_data["codigo_fornecedor"],
                    item_data["descricao"],
                    item_data["unidade"],
                    item_data["quantidade"],
                    item_data["valor_unitario"],
                    item_data["desconto"],
                    item_data["valor_total"],
                    "pendente" if product_created else "vinculado",
                ),
            )
        connection.commit()
    except Exception:
        connection.rollback()
        raise
    finally:
        cursor.close()
        connection.close()

    audit_receipt(recebimento_id, "Entrada manual", f"Documento {numero_doc} registrado manualmente.", "Sistema", "Manual")
    apply_receipt_effects(recebimento_id)
    if ordem_id:
        ordem = query("SELECT quantidade FROM ordens_compra WHERE id=%s", (ordem_id,), one=True)
        received_qty = sum(item[0]["quantidade"] for item in items)
        if ordem and received_qty >= decimal_value(ordem["quantidade"]):
            execute("UPDATE ordens_compra SET status_compra='recebida' WHERE id=%s", (ordem_id,))
            execute("UPDATE recebimentos SET data_real=%s,status='entregue completo' WHERE ordem_id=%s", (data_entrada, ordem_id))
        else:
            execute("UPDATE ordens_compra SET status_compra='em compra' WHERE id=%s", (ordem_id,))
            execute("UPDATE recebimentos SET data_real=%s,status='entregue parcial' WHERE ordem_id=%s", (data_entrada, ordem_id))
    flash("Entrada manual registrada e processada com sucesso.", "success")
    return redirect(url_for("recebimentos"))


@app.post("/recebimentos/produto_rapido")
def cadastrar_produto_rapido_recebimento():
    ensure_auxiliary_registries()
    ensure_product_schema()
    form = request.form
    fornecedor_id = form.get("fornecedor_id")
    supplier = query("SELECT * FROM fornecedores WHERE id=%s", (fornecedor_id,), one=True)
    if not supplier:
        flash("Fornecedor obrigatorio para cadastro rapido.", "danger")
        return redirect(url_for("recebimentos"))
    quantidade_por_unidade = decimal_value(form.get("quantidade_por_unidade_compra") or 1)
    if quantidade_por_unidade <= 0:
        quantidade_por_unidade = Decimal("1")
    unidade = form.get("unidade") or "UN"
    produto_id = execute(
        """
        INSERT INTO produtos (codigo,descricao,categoria,unidade,quantidade_por_unidade_compra,unidade_base,fornecedor_id,estoque_seguranca,status)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """,
        (
            (form.get("codigo") or "Nao registrado").strip(),
            form.get("descricao"),
            form.get("categoria"),
            unidade,
            quantidade_por_unidade,
            form.get("unidade_base") or unidade,
            fornecedor_id,
            0,
            "ativo",
        ),
    )
    item_data = {"codigo_fornecedor": form.get("codigo_fornecedor") or "", "descricao": form.get("descricao") or ""}
    register_product_supplier_link(supplier["nome"], supplier.get("cnpj"), item_data, produto_id)
    flash("Produto cadastrado rapidamente. Selecione-o na entrada manual.", "success")
    return redirect(url_for("recebimentos"))


@app.get("/api/ordens/<int:ordem_id>/itens")
def api_ordem_itens(ordem_id):
    ordem = query(
        """
        SELECT o.*,f.nome fornecedor,p.codigo produto_codigo,p.descricao produto_descricao,p.unidade produto_unidade
        FROM ordens_compra o
        JOIN fornecedores f ON f.id=o.fornecedor_id
        JOIN produtos p ON p.id=o.produto_id
        WHERE o.id=%s
        """,
        (ordem_id,),
        one=True,
    )
    if not ordem:
        return jsonify({"items": []})
    return jsonify(
        {
            "fornecedor_id": ordem["fornecedor_id"],
            "items": [
                {
                    "produto_id": ordem["produto_id"],
                    "codigo": ordem["produto_codigo"],
                    "descricao": ordem["produto_descricao"],
                    "unidade": ordem["produto_unidade"],
                    "quantidade": str(ordem["quantidade"]),
                    "valor_unitario": str(ordem["preco_negociado"]),
                }
            ],
        }
    )


@app.post("/recebimentos/nfe_itens/<int:item_id>/vincular")
def vincular_item_nfe(item_id):
    ensure_nfe_schema()
    produto_id = request.form.get("produto_id")
    if not produto_id:
        flash("Selecione um codigo interno para vincular.", "danger")
        return redirect(url_for("recebimentos"))

    item = query(
        """
        SELECT i.*,n.fornecedor,n.fornecedor_cnpj,n.id nfe_id
        FROM nfe_importacao_itens i
        JOIN nfe_importacoes n ON n.id=i.nfe_importacao_id
        WHERE i.id=%s
        """,
        (item_id,),
        one=True,
    )
    product = query("SELECT id,codigo FROM produtos WHERE id=%s", (produto_id,), one=True)
    if not item or not product:
        flash("Item ou produto nao encontrado.", "danger")
        return redirect(url_for("recebimentos"))

    execute(
        """
        INSERT INTO produto_fornecedor_relacionamentos
        (fornecedor,cnpj,codigo_fornecedor,descricao_fornecedor,produto_id)
        VALUES (%s,%s,%s,%s,%s)
        ON DUPLICATE KEY UPDATE
          descricao_fornecedor=VALUES(descricao_fornecedor),
          produto_id=VALUES(produto_id),
          atualizado_em=CURRENT_TIMESTAMP
        """,
        (item["fornecedor"], item["fornecedor_cnpj"], item["codigo_fornecedor"], item["descricao"], produto_id),
    )
    execute("UPDATE nfe_importacao_itens SET produto_id=%s,status='vinculado' WHERE id=%s", (produto_id, item_id))
    refresh_receipt_status(item["nfe_id"])
    apply_receipt_effects(item["nfe_id"])
    flash("Vinculo salvo. Proximas importacoes reconhecerão este produto automaticamente.", "success")
    return redirect(url_for("recebimentos", nfe_id=item["nfe_id"]))


@app.post("/recebimentos/nfe/<int:nfe_id>/vincular_ordem")
def vincular_ordem_nfe(nfe_id):
    ensure_nfe_schema()
    ordem_id = request.form.get("ordem_id")
    if not ordem_id:
        flash("Selecione uma Ordem de Compra para vincular a NF-e.", "danger")
        return redirect(url_for("recebimentos", nfe_id=nfe_id))
    nfe = query("SELECT id,data_entrada FROM nfe_importacoes WHERE id=%s", (nfe_id,), one=True)
    ordem = query("SELECT id FROM ordens_compra WHERE id=%s", (ordem_id,), one=True)
    if not nfe or not ordem:
        flash("NF-e ou Ordem de Compra nao encontrada.", "danger")
        return redirect(url_for("recebimentos", nfe_id=nfe_id))
    execute("UPDATE nfe_importacoes SET ordem_id=%s WHERE id=%s", (ordem_id, nfe_id))
    refresh_receipt_status(nfe_id)
    execute("UPDATE recebimentos SET data_real=%s,status='entregue completo' WHERE ordem_id=%s", (nfe["data_entrada"], ordem_id))
    execute("UPDATE ordens_compra SET status_compra='recebida' WHERE id=%s", (ordem_id,))
    audit_receipt(nfe_id, "Vinculo de OC", f"Ordem {ordem_id} vinculada.", "Sistema")
    flash("NF-e vinculada a Ordem de Compra.", "success")
    return redirect(url_for("recebimentos", nfe_id=nfe_id))


@app.route("/orcamentos", methods=["GET", "POST"])
def orcamentos():
    return redirect(url_for("cadastros", modulo="orcamentos"))


@app.post("/orcamentos/<int:item_id>/excluir")
def excluir_orcamento(item_id):
    execute("DELETE FROM orcamentos WHERE id=%s", (item_id,))
    flash("Orcamento excluido.", "success")
    return redirect(url_for("cadastros", modulo="orcamentos"))


if __name__ == "__main__":
    app.run(debug=True)
