from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
import re
from xml.etree import ElementTree as ET

import mysql.connector
from flask import Flask, flash, jsonify, redirect, render_template, request, send_file, url_for
from openpyxl import Workbook, load_workbook

from config import Config


app = Flask(__name__)
app.config.from_object(Config)


DEFAULT_UNITS = (
    ("UN", "Unidade"),
    ("CX", "Caixa"),
    ("PCT", "Pacote"),
    ("KG", "Quilograma"),
    ("G", "Grama"),
    ("L", "Litro"),
    ("ML", "Mililitro"),
    ("M", "Metro"),
    ("ROLO", "Rolo"),
    ("PALLET", "Pallet"),
    ("FARDO", "Fardo"),
    ("KIT", "Kit"),
    ("PAR", "Par"),
    ("M2", "Metro Quadrado (M²)"),
    ("M3", "Metro Cúbico (M³)"),
)


def get_db():
    return mysql.connector.connect(
        host=app.config["MYSQL_HOST"],
        port=app.config["MYSQL_PORT"],
        user=app.config["MYSQL_USER"],
        password=app.config["MYSQL_PASSWORD"],
        database=app.config["MYSQL_DATABASE"],
    )


def query(sql, params=(), one=False):
    connection = get_db()
    cursor = connection.cursor(dictionary=True)
    cursor.execute(sql, params)
    result = cursor.fetchone() if one else cursor.fetchall()
    cursor.close()
    connection.close()
    return result


def execute(sql, params=()):
    connection = get_db()
    cursor = connection.cursor()
    try:
        cursor.execute(sql, params)
        connection.commit()
        return cursor.lastrowid
    except Exception:
        connection.rollback()
        raise
    finally:
        cursor.close()
        connection.close()


def decimal_value(value):
    return Decimal(str(value or 0))


def import_cell_value(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).strip()


def normalize_cnpj(value):
    digits = re.sub(r"\D", "", value or "")
    if not digits:
        return ""
    if len(digits) == 14:
        return f"{digits[:2]}.{digits[2:5]}.{digits[5:8]}/{digits[8:12]}-{digits[12:]}"
    return value.strip()


def is_valid_cnpj(value):
    digits = re.sub(r"\D", "", value or "")
    if len(digits) != 14 or digits == digits[0] * 14:
        return False

    def check_digit(base, weights):
        total = sum(int(number) * weight for number, weight in zip(base, weights))
        remainder = total % 11
        return "0" if remainder < 2 else str(11 - remainder)

    first = check_digit(digits[:12], [5, 4, 3, 2, 9, 8, 7, 6, 5, 4, 3, 2])
    second = check_digit(digits[:13], [6, 5, 4, 3, 2, 9, 8, 7, 6, 5, 4, 3, 2])
    return digits[-2:] == first + second


def ensure_supplier_schema():
    column = query(
        """
        SELECT IS_NULLABLE
        FROM information_schema.COLUMNS
        WHERE TABLE_SCHEMA=%s AND TABLE_NAME='fornecedores' AND COLUMN_NAME='cnpj'
        """,
        (app.config["MYSQL_DATABASE"],),
        one=True,
    )
    if column and column["IS_NULLABLE"] == "NO":
        execute("ALTER TABLE fornecedores MODIFY cnpj VARCHAR(18) NULL")


def ensure_nfe_schema():
    execute(
        """
        CREATE TABLE IF NOT EXISTS produto_fornecedor_relacionamentos (
          id INT AUTO_INCREMENT PRIMARY KEY,
          fornecedor VARCHAR(150) NOT NULL,
          cnpj VARCHAR(18),
          codigo_fornecedor VARCHAR(80) NOT NULL,
          descricao_fornecedor VARCHAR(255) NOT NULL,
          produto_id INT NOT NULL,
          criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
          atualizado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
          CONSTRAINT fk_rel_produto FOREIGN KEY (produto_id) REFERENCES produtos(id),
          UNIQUE KEY uk_rel_cnpj_codigo (cnpj,codigo_fornecedor),
          KEY idx_rel_fornecedor_codigo (fornecedor,codigo_fornecedor)
        )
        """
    )
    execute(
        """
        CREATE TABLE IF NOT EXISTS nfe_importacoes (
          id INT AUTO_INCREMENT PRIMARY KEY,
          ordem_id INT NULL,
          fornecedor VARCHAR(150) NOT NULL,
          fornecedor_cnpj VARCHAR(18) NOT NULL,
          numero_nf VARCHAR(30) NOT NULL,
          serie VARCHAR(10),
          chave_nfe VARCHAR(60) NOT NULL UNIQUE,
          data_emissao DATE,
          data_entrada DATE NOT NULL,
          valor_total DECIMAL(14,2) NOT NULL DEFAULT 0,
          criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
          CONSTRAINT fk_nfe_ordem FOREIGN KEY (ordem_id) REFERENCES ordens_compra(id),
          UNIQUE KEY uk_nfe_fornecedor_numero_serie (fornecedor_cnpj,numero_nf,serie)
        )
        """
    )
    if not column_exists("nfe_importacoes", "ordem_id"):
        execute("ALTER TABLE nfe_importacoes ADD COLUMN ordem_id INT NULL AFTER id")
        constraint = query(
            """
            SELECT CONSTRAINT_NAME
            FROM information_schema.TABLE_CONSTRAINTS
            WHERE TABLE_SCHEMA=%s AND TABLE_NAME='nfe_importacoes' AND CONSTRAINT_NAME='fk_nfe_ordem'
            """,
            (app.config["MYSQL_DATABASE"],),
            one=True,
        )
        if not constraint:
            execute("ALTER TABLE nfe_importacoes ADD CONSTRAINT fk_nfe_ordem FOREIGN KEY (ordem_id) REFERENCES ordens_compra(id)")
    execute(
        """
        CREATE TABLE IF NOT EXISTS nfe_importacao_itens (
          id INT AUTO_INCREMENT PRIMARY KEY,
          nfe_importacao_id INT NOT NULL,
          produto_id INT NULL,
          codigo_fornecedor VARCHAR(80),
          descricao VARCHAR(255) NOT NULL,
          ncm VARCHAR(20),
          cfop VARCHAR(10),
          cest VARCHAR(20),
          unidade VARCHAR(20),
          quantidade DECIMAL(14,4) NOT NULL DEFAULT 0,
          valor_unitario DECIMAL(14,4) NOT NULL DEFAULT 0,
          valor_total DECIMAL(14,2) NOT NULL DEFAULT 0,
          ean VARCHAR(30),
          status ENUM('vinculado','pendente') NOT NULL DEFAULT 'pendente',
          CONSTRAINT fk_nfe_item_importacao FOREIGN KEY (nfe_importacao_id) REFERENCES nfe_importacoes(id) ON DELETE CASCADE,
          CONSTRAINT fk_nfe_item_produto FOREIGN KEY (produto_id) REFERENCES produtos(id)
        )
        """
    )


def nfe_text(parent, path, namespaces=None, default=""):
    node = parent.find(path, namespaces or {})
    return (node.text or "").strip() if node is not None and node.text is not None else default


def nfe_decimal(value):
    return Decimal(str(value or "0").replace(",", "."))


def parse_nfe_xml(file_storage):
    try:
        tree = ET.parse(file_storage)
    except ET.ParseError as error:
        raise ValueError(f"XML invalido: {error}") from error

    root = tree.getroot()
    namespace = root.tag.split("}")[0].strip("{") if root.tag.startswith("{") else ""
    ns = {"n": namespace} if namespace else {}
    prefix = "n:" if namespace else ""
    inf = root.find(f".//{prefix}infNFe", ns)
    if inf is None:
        raise ValueError("O arquivo XML nao parece ser uma NF-e.")

    ide = inf.find(f"{prefix}ide", ns)
    emit = inf.find(f"{prefix}emit", ns)
    total = inf.find(f"{prefix}total/{prefix}ICMSTot", ns)
    if ide is None or emit is None:
        raise ValueError("Nao foi possivel ler cabecalho da NF-e.")

    chave = (inf.attrib.get("Id") or "").replace("NFe", "")
    numero_nf = nfe_text(ide, f"{prefix}nNF", ns)
    serie = nfe_text(ide, f"{prefix}serie", ns)
    emissao_raw = nfe_text(ide, f"{prefix}dhEmi", ns) or nfe_text(ide, f"{prefix}dEmi", ns)
    data_emissao = emissao_raw[:10] if emissao_raw else None
    fornecedor = nfe_text(emit, f"{prefix}xNome", ns)
    cnpj = normalize_cnpj(nfe_text(emit, f"{prefix}CNPJ", ns))
    valor_total = nfe_decimal(nfe_text(total, f"{prefix}vNF", ns) if total is not None else "0")

    if not chave or not numero_nf or not fornecedor or not cnpj:
        raise ValueError("NF-e sem chave, numero, fornecedor ou CNPJ.")

    items = []
    for det in inf.findall(f"{prefix}det", ns):
        prod = det.find(f"{prefix}prod", ns)
        if prod is None:
            continue
        items.append(
            {
                "codigo_fornecedor": nfe_text(prod, f"{prefix}cProd", ns),
                "descricao": nfe_text(prod, f"{prefix}xProd", ns),
                "ncm": nfe_text(prod, f"{prefix}NCM", ns),
                "cfop": nfe_text(prod, f"{prefix}CFOP", ns),
                "cest": nfe_text(prod, f"{prefix}CEST", ns),
                "unidade": nfe_text(prod, f"{prefix}uCom", ns),
                "quantidade": nfe_decimal(nfe_text(prod, f"{prefix}qCom", ns)),
                "valor_unitario": nfe_decimal(nfe_text(prod, f"{prefix}vUnCom", ns)),
                "valor_total": nfe_decimal(nfe_text(prod, f"{prefix}vProd", ns)),
                "ean": nfe_text(prod, f"{prefix}cEAN", ns),
            }
        )
    if not items:
        raise ValueError("NF-e sem produtos para importar.")

    return {
        "fornecedor": fornecedor,
        "fornecedor_cnpj": cnpj,
        "numero_nf": numero_nf,
        "serie": serie,
        "chave_nfe": chave,
        "data_emissao": data_emissao,
        "data_entrada": date.today().isoformat(),
        "valor_total": valor_total,
        "items": items,
    }


def identify_nfe_product(fornecedor, cnpj, codigo_fornecedor, descricao):
    rel = query(
        """
        SELECT p.id,p.codigo
        FROM produto_fornecedor_relacionamentos r
        JOIN produtos p ON p.id=r.produto_id
        WHERE (r.cnpj=%s OR LOWER(r.fornecedor)=LOWER(%s)) AND r.codigo_fornecedor=%s
        ORDER BY r.id DESC LIMIT 1
        """,
        (cnpj, fornecedor, codigo_fornecedor),
        one=True,
    )
    if rel:
        return rel

    product = query(
        """
        SELECT p.id,p.codigo
        FROM produtos p
        LEFT JOIN fornecedores f ON f.id=p.fornecedor_id
        WHERE p.status='ativo'
          AND (f.cnpj=%s OR LOWER(f.nome)=LOWER(%s))
          AND (p.codigo=%s OR LOWER(p.descricao)=LOWER(%s))
        ORDER BY (p.codigo=%s) DESC
        LIMIT 1
        """,
        (cnpj, fornecedor, codigo_fornecedor, descricao, codigo_fornecedor),
        one=True,
    )
    return product


def unique_product_code(base_code):
    base = re.sub(r"\s+", "-", (base_code or "").strip())[:24] or f"XML-{date.today().strftime('%Y%m%d')}"
    code = base[:30]
    suffix = 1
    while query("SELECT id FROM produtos WHERE codigo=%s", (code,), one=True):
        tail = f"-{suffix}"
        code = f"{base[:30 - len(tail)]}{tail}"
        suffix += 1
    return code


def get_or_create_nfe_supplier(fornecedor, cnpj):
    supplier = query("SELECT id FROM fornecedores WHERE cnpj=%s", (cnpj,), one=True)
    if supplier:
        return supplier["id"]
    supplier = query("SELECT id FROM fornecedores WHERE LOWER(nome)=LOWER(%s)", (fornecedor,), one=True)
    if supplier:
        return supplier["id"]
    return execute(
        "INSERT INTO fornecedores (nome,cnpj,contato,telefone,email,status) VALUES (%s,%s,%s,%s,%s,%s)",
        (fornecedor, cnpj or None, None, None, None, "ativo"),
    )


def create_product_from_nfe_item(item, fornecedor_id):
    code = unique_product_code(item["codigo_fornecedor"] or item["descricao"][:24])
    produto_id = execute(
        """
        INSERT INTO produtos
        (codigo,descricao,categoria,unidade,fornecedor_id,estoque_seguranca,status)
        VALUES (%s,%s,%s,%s,%s,%s,%s)
        """,
        (
            code,
            item["descricao"] or code,
            "Importado via NF-e",
            item["unidade"] or "UN",
            fornecedor_id,
            0,
            "ativo",
        ),
    )
    return {"id": produto_id, "codigo": code}


def column_exists(table_name, column_name):
    return query(
        """
        SELECT COLUMN_NAME
        FROM information_schema.COLUMNS
        WHERE TABLE_SCHEMA=%s AND TABLE_NAME=%s AND COLUMN_NAME=%s
        """,
        (app.config["MYSQL_DATABASE"], table_name, column_name),
        one=True,
    )


def ensure_auxiliary_registries():
    execute(
        """
        CREATE TABLE IF NOT EXISTS cadastro_categorias (
          id INT AUTO_INCREMENT PRIMARY KEY,
          nome VARCHAR(120) NOT NULL UNIQUE,
          status ENUM('ativo','inativo') NOT NULL DEFAULT 'ativo',
          criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
          atualizado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
        )
        """
    )
    execute(
        """
        CREATE TABLE IF NOT EXISTS cadastro_tipos_pagamento (
          id INT AUTO_INCREMENT PRIMARY KEY,
          nome VARCHAR(120) NOT NULL UNIQUE,
          status ENUM('ativo','inativo') NOT NULL DEFAULT 'ativo',
          criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
          atualizado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
        )
        """
    )
    for row in query(
        """
        SELECT categoria nome FROM produtos WHERE categoria IS NOT NULL AND categoria<>''
        UNION
        SELECT categoria nome FROM orcamentos WHERE categoria IS NOT NULL AND categoria<>''
        """
    ):
        execute(
            "INSERT IGNORE INTO cadastro_categorias (nome,status) VALUES (%s,'ativo')",
            (row["nome"],),
        )
    execute(
        "INSERT IGNORE INTO cadastro_categorias (nome,status) VALUES (%s,'ativo')",
        ("Importado via NF-e",),
    )
    for nome in ("PIX", "Boleto", "Cartao de Credito", "Cartao de Debito", "Dinheiro", "Transferencia"):
        execute(
            "INSERT IGNORE INTO cadastro_tipos_pagamento (nome,status) VALUES (%s,'ativo')",
            (nome,),
        )
    for row in query("SELECT DISTINCT metodo_pagamento nome FROM ordens_compra WHERE metodo_pagamento IS NOT NULL AND metodo_pagamento<>''"):
        execute(
            "INSERT IGNORE INTO cadastro_tipos_pagamento (nome,status) VALUES (%s,'ativo')",
            (row["nome"],),
        )


def ensure_product_schema():
    execute(
        """
        CREATE TABLE IF NOT EXISTS unidades_medida (
          id INT AUTO_INCREMENT PRIMARY KEY,
          codigo VARCHAR(20) NOT NULL UNIQUE,
          nome VARCHAR(120) NOT NULL,
          status ENUM('ativo','inativo') NOT NULL DEFAULT 'ativo',
          criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        """
    )
    for codigo, nome in DEFAULT_UNITS:
        execute(
            """
            INSERT INTO unidades_medida (codigo,nome,status)
            VALUES (%s,%s,'ativo')
            ON DUPLICATE KEY UPDATE nome=VALUES(nome)
            """,
            (codigo, nome),
        )
    if not column_exists("produtos", "fornecedor_id"):
        execute("ALTER TABLE produtos ADD COLUMN fornecedor_id INT NULL AFTER unidade")
    if not column_exists("produtos", "estoque_seguranca"):
        execute("ALTER TABLE produtos ADD COLUMN estoque_seguranca DECIMAL(12,2) NOT NULL DEFAULT 0 AFTER fornecedor_id")
    execute("UPDATE produtos SET unidade='UN' WHERE LOWER(unidade) IN ('un','unidade')")
    execute("UPDATE produtos SET unidade='L' WHERE unidade='GL'")
    execute("UPDATE produtos SET estoque_seguranca=COALESCE(NULLIF(estoque_seguranca,0),estoque_semanal,0)")
    execute(
        """
        UPDATE produtos p
        SET fornecedor_id = (
            SELECT o.fornecedor_id
            FROM ordens_compra o
            WHERE o.produto_id=p.id
            ORDER BY o.data_preenchimento DESC, o.id DESC
            LIMIT 1
        )
        WHERE p.fornecedor_id IS NULL
        """
    )
    execute(
        """
        UPDATE produtos p
        SET fornecedor_id = (SELECT f.id FROM fornecedores f WHERE f.status='ativo' ORDER BY f.nome LIMIT 1)
        WHERE p.fornecedor_id IS NULL
        """
    )
    constraint = query(
        """
        SELECT CONSTRAINT_NAME
        FROM information_schema.TABLE_CONSTRAINTS
        WHERE TABLE_SCHEMA=%s AND TABLE_NAME='produtos' AND CONSTRAINT_NAME='fk_produto_fornecedor'
        """,
        (app.config["MYSQL_DATABASE"],),
        one=True,
    )
    if not constraint:
        execute("ALTER TABLE produtos ADD CONSTRAINT fk_produto_fornecedor FOREIGN KEY (fornecedor_id) REFERENCES fornecedores(id)")


@app.template_filter("moeda")
def moeda(value):
    number = decimal_value(value)
    return f"R$ {number:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


@app.context_processor
def inject_globals():
    return {"hoje": date.today(), "ano": date.today().year}


@app.route("/")
def dashboard():
    indicadores = query(
        """
        SELECT
          COALESCE(SUM(CASE WHEN YEAR(data_preenchimento)=YEAR(CURDATE())
                            AND MONTH(data_preenchimento)=MONTH(CURDATE()) THEN valor_total END), 0) total_mes,
          SUM(status_compra='aguardando autorizacao') aguardando,
          SUM(data_entrega < CURDATE() AND status_compra NOT IN ('recebida','cancelada')) atrasadas
        FROM ordens_compra
        """,
        one=True,
    )
    categorias = query(
        """SELECT p.categoria, COALESCE(SUM(o.valor_total),0) total
           FROM produtos p LEFT JOIN ordens_compra o ON o.produto_id=p.id
           GROUP BY p.categoria ORDER BY total DESC LIMIT 6"""
    )
    ultimas = query(
        """SELECT o.*, f.nome fornecedor, p.descricao produto
           FROM ordens_compra o JOIN fornecedores f ON f.id=o.fornecedor_id
           JOIN produtos p ON p.id=o.produto_id ORDER BY o.id DESC LIMIT 8"""
    )
    return render_template("dashboard.html", indicadores=indicadores, categorias=categorias, ultimas=ultimas)


def registry_usage(module, name):
    if module == "categorias":
        usages = []
        product_count = query("SELECT COUNT(*) total FROM produtos WHERE categoria=%s", (name,), one=True)["total"]
        budget_count = query("SELECT COUNT(*) total FROM orcamentos WHERE categoria=%s", (name,), one=True)["total"]
        if product_count:
            usages.append(f"Produtos ({product_count})")
        if budget_count:
            usages.append(f"Orcamentos ({budget_count})")
        return usages
    if module == "pagamentos":
        count = query("SELECT COUNT(*) total FROM ordens_compra WHERE metodo_pagamento=%s", (name,), one=True)["total"]
        return [f"Ordens de compra ({count})"] if count else []
    return []


@app.route("/cadastros", methods=["GET", "POST"])
def cadastros():
    ensure_auxiliary_registries()
    module = request.args.get("modulo", "categorias")
    if module not in ("categorias", "pagamentos"):
        module = "categorias"
    configs = {
        "categorias": {
            "table": "cadastro_categorias",
            "title": "Categorias de Produto",
            "field": "Nome da categoria",
            "modal": "categoriaForm",
        },
        "pagamentos": {
            "table": "cadastro_tipos_pagamento",
            "title": "Tipos de Pagamento",
            "field": "Nome do tipo de pagamento",
            "modal": "pagamentoForm",
        },
    }
    config = configs[module]

    if request.method == "POST":
        form = request.form
        form_module = form.get("modulo", module)
        config = configs.get(form_module, config)
        item_id = form.get("id")
        nome = (form.get("nome") or "").strip()
        status = form.get("status") or "ativo"
        if not nome:
            flash("Informe o nome do cadastro.", "danger")
            return redirect(url_for("cadastros", modulo=form_module))
        duplicate = query(
            f"SELECT id FROM {config['table']} WHERE LOWER(nome)=LOWER(%s) AND (%s='' OR id<>%s)",
            (nome, item_id or "", item_id or 0),
            one=True,
        )
        if duplicate:
            flash("Ja existe um cadastro com este nome.", "danger")
            return redirect(url_for("cadastros", modulo=form_module))
        if item_id:
            old = query(f"SELECT nome FROM {config['table']} WHERE id=%s", (item_id,), one=True)
            usages = registry_usage(form_module, old["nome"]) if old and old["nome"] != nome else []
            if usages:
                flash(f"Nao e possivel renomear. Cadastro em uso em: {', '.join(usages)}.", "danger")
                return redirect(url_for("cadastros", modulo=form_module))
            execute(f"UPDATE {config['table']} SET nome=%s,status=%s WHERE id=%s", (nome, status, item_id))
        else:
            execute(f"INSERT INTO {config['table']} (nome,status) VALUES (%s,%s)", (nome, status))
        flash("Cadastro salvo com sucesso.", "success")
        return redirect(url_for("cadastros", modulo=form_module))

    search = (request.args.get("q") or "").strip()
    sort = request.args.get("sort", "nome")
    direction = request.args.get("dir", "asc")
    page = max(int(request.args.get("page", 1) or 1), 1)
    per_page = 10
    allowed_sort = {"nome", "status", "criado_em", "atualizado_em"}
    if sort not in allowed_sort:
        sort = "nome"
    direction = "desc" if direction == "desc" else "asc"
    where = "WHERE nome LIKE %s" if search else ""
    params = (f"%{search}%",) if search else ()
    total = query(f"SELECT COUNT(*) total FROM {config['table']} {where}", params, one=True)["total"]
    offset = (page - 1) * per_page
    rows = query(
        f"SELECT * FROM {config['table']} {where} ORDER BY {sort} {direction} LIMIT %s OFFSET %s",
        params + (per_page, offset),
    )
    total_pages = max((total + per_page - 1) // per_page, 1)
    counts = {
        "categorias": query("SELECT COUNT(*) total FROM cadastro_categorias", one=True)["total"],
        "pagamentos": query("SELECT COUNT(*) total FROM cadastro_tipos_pagamento", one=True)["total"],
    }
    return render_template(
        "cadastros.html",
        module=module,
        config=config,
        rows=rows,
        search=search,
        sort=sort,
        direction=direction,
        page=page,
        total_pages=total_pages,
        total=total,
        counts=counts,
    )


@app.post("/cadastros/<module>/<int:item_id>/excluir")
def excluir_cadastro(module, item_id):
    ensure_auxiliary_registries()
    configs = {
        "categorias": "cadastro_categorias",
        "pagamentos": "cadastro_tipos_pagamento",
    }
    table = configs.get(module)
    if not table:
        flash("Cadastro invalido.", "danger")
        return redirect(url_for("cadastros"))
    item = query(f"SELECT nome FROM {table} WHERE id=%s", (item_id,), one=True)
    if not item:
        flash("Cadastro nao encontrado.", "danger")
        return redirect(url_for("cadastros", modulo=module))
    usages = registry_usage(module, item["nome"])
    if usages:
        flash(f"Nao e possivel excluir. Cadastro em uso em: {', '.join(usages)}.", "danger")
        return redirect(url_for("cadastros", modulo=module))
    execute(f"DELETE FROM {table} WHERE id=%s", (item_id,))
    flash("Cadastro excluido.", "success")
    return redirect(url_for("cadastros", modulo=module))


@app.route("/fornecedores", methods=["GET", "POST"])
def fornecedores():
    ensure_supplier_schema()
    if request.method == "POST":
        form = request.form
        fornecedor_id = form.get("id")
        params = (form["nome"], form["cnpj"], form.get("contato"), form.get("telefone"), form.get("email"), form["status"])
        try:
            if fornecedor_id:
                execute("UPDATE fornecedores SET nome=%s,cnpj=%s,contato=%s,telefone=%s,email=%s,status=%s WHERE id=%s", params + (fornecedor_id,))
            else:
                execute("INSERT INTO fornecedores (nome,cnpj,contato,telefone,email,status) VALUES (%s,%s,%s,%s,%s,%s)", params)
        except mysql.connector.IntegrityError:
            flash("Ja existe um fornecedor cadastrado com este CNPJ.", "danger")
            return redirect(url_for("fornecedores"))
        flash("Fornecedor salvo com sucesso.", "success")
        return redirect(url_for("fornecedores"))
    return render_template("fornecedores.html", fornecedores=query("SELECT * FROM fornecedores ORDER BY nome"))


@app.get("/fornecedores/modelo")
def baixar_modelo_fornecedores():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Fornecedores"
    sheet.append(["Fornecedor", "CNPJ", "Contato", "Telefone", "Email"])
    sheet.append(["Fornecedor Exemplo", "00.000.000/0000-00", "João Silva", "(11) 99999-9999", "joao@exemplo.com"])
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name="modelo_fornecedores.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.post("/fornecedores/<int:item_id>/excluir")
def excluir_fornecedor(item_id):
    try:
        execute("DELETE FROM fornecedores WHERE id=%s", (item_id,))
        flash("Fornecedor excluido.", "success")
    except mysql.connector.Error:
        flash("Fornecedor possui ordens vinculadas e nao pode ser excluido.", "danger")
    return redirect(url_for("fornecedores"))


@app.post("/fornecedores/importar")
def importar_fornecedores():
    ensure_supplier_schema()
    file = request.files.get("arquivo")
    if not file or not file.filename:
        flash("Selecione um arquivo XLSX para importar.", "danger")
        return redirect(url_for("fornecedores"))
    if not file.filename.lower().endswith(".xlsx"):
        flash("Envie um arquivo no formato .xlsx.", "danger")
        return redirect(url_for("fornecedores"))

    try:
        workbook = load_workbook(file, read_only=True, data_only=True)
        sheet = workbook.active
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        expected_headers = ["Fornecedor", "CNPJ", "Contato", "Telefone", "Email"]
        headers = [import_cell_value(value) for value in (header_row or []) if import_cell_value(value)]
        if headers != expected_headers:
            workbook.close()
            flash("O arquivo deve conter exatamente as colunas: Fornecedor, CNPJ, Contato, Telefone, Email.", "danger")
            return redirect(url_for("fornecedores"))

        total = imported = ignored = errors = 0
        error_messages = []
        seen_cnpjs = set()
        seen_names = set()
        for row in sheet.iter_rows(min_row=2, values_only=True):
            values = [import_cell_value(value) for value in row[:5]]
            if not any(values):
                continue

            total += 1
            nome, cnpj_raw, contato, telefone, email = values
            cnpj = normalize_cnpj(cnpj_raw)
            row_errors = []

            if not nome:
                row_errors.append("fornecedor sem nome")
            if cnpj and not is_valid_cnpj(cnpj):
                row_errors.append("CNPJ invalido")

            name_key = nome.strip().lower()
            cnpj_key = re.sub(r"\D", "", cnpj)
            if cnpj_key:
                if cnpj_key in seen_cnpjs or query("SELECT id FROM fornecedores WHERE REPLACE(REPLACE(REPLACE(cnpj,'.',''),'/',''),'-','')=%s", (cnpj_key,), one=True):
                    row_errors.append("CNPJ duplicado")
            elif name_key and (name_key in seen_names or query("SELECT id FROM fornecedores WHERE LOWER(nome)=%s", (name_key,), one=True)):
                row_errors.append("fornecedor duplicado por nome")

            if row_errors:
                errors += len(row_errors)
                ignored += 1
                error_messages.append(f"Linha {total + 1}: {', '.join(row_errors)}")
                continue

            execute(
                "INSERT INTO fornecedores (nome,cnpj,contato,telefone,email,status) VALUES (%s,%s,%s,%s,%s,%s)",
                (nome, cnpj or None, contato or None, telefone or None, email or None, "ativo"),
            )
            imported += 1
            if cnpj_key:
                seen_cnpjs.add(cnpj_key)
            else:
                seen_names.add(name_key)

        workbook.close()
        summary = (
            f"Importacao concluida. Total de linhas lidas: {total}. "
            f"Fornecedores importados: {imported}. Fornecedores ignorados: {ignored}. "
            f"Erros encontrados: {errors}."
        )
        if error_messages:
            summary += " " + " | ".join(error_messages[:5])
            if len(error_messages) > 5:
                summary += f" | Mais {len(error_messages) - 5} erro(s)."
        flash(summary, "success" if imported else "warning")
    except Exception as error:
        flash(f"Nao foi possivel importar o arquivo: {error}", "danger")
    return redirect(url_for("fornecedores"))


@app.route("/produtos", methods=["GET", "POST"])
def produtos():
    ensure_auxiliary_registries()
    ensure_product_schema()
    if request.method == "POST":
        form = request.form
        product_id = form.get("id")
        estoque_seguranca = decimal_value(form.get("estoque_seguranca"))
        if estoque_seguranca <= 0:
            flash("Estoque de seguranca deve ser um numero positivo.", "danger")
            return redirect(url_for("produtos"))
        params = (
            form["codigo"],
            form["descricao"],
            form["categoria"],
            form["unidade"],
            form["fornecedor_id"],
            estoque_seguranca,
            form["status"],
        )
        try:
            if product_id:
                execute("UPDATE produtos SET codigo=%s,descricao=%s,categoria=%s,unidade=%s,fornecedor_id=%s,estoque_seguranca=%s,status=%s WHERE id=%s", params + (product_id,))
            else:
                execute("INSERT INTO produtos (codigo,descricao,categoria,unidade,fornecedor_id,estoque_seguranca,status) VALUES (%s,%s,%s,%s,%s,%s,%s)", params)
        except mysql.connector.IntegrityError:
            flash("Ja existe um produto cadastrado com este codigo.", "danger")
            return redirect(url_for("produtos"))
        flash("Produto salvo com sucesso.", "success")
        return redirect(url_for("produtos"))
    produtos_rows = query(
        """
        SELECT p.*, f.nome fornecedor_nome, u.nome unidade_nome
        FROM produtos p
        LEFT JOIN fornecedores f ON f.id=p.fornecedor_id
        LEFT JOIN unidades_medida u ON u.codigo=p.unidade
        ORDER BY p.descricao
        """
    )
    categorias = query("SELECT nome categoria FROM cadastro_categorias WHERE status='ativo' ORDER BY nome")
    unidades = query("SELECT codigo,nome FROM unidades_medida WHERE status='ativo' ORDER BY nome")
    fornecedores_ativos = query("SELECT id,nome FROM fornecedores WHERE status='ativo' ORDER BY nome")
    return render_template(
        "produtos.html",
        produtos=produtos_rows,
        categorias=categorias,
        unidades=unidades,
        fornecedores=fornecedores_ativos,
    )


@app.post("/produtos/<int:item_id>/excluir")
def excluir_produto(item_id):
    try:
        execute("DELETE FROM produtos WHERE id=%s", (item_id,))
        flash("Produto excluido.", "success")
    except mysql.connector.Error:
        flash("Produto possui ordens vinculadas e nao pode ser excluido.", "danger")
    return redirect(url_for("produtos"))


@app.route("/ordens")
def ordens():
    rows = query("""SELECT o.*, f.nome fornecedor, p.descricao produto FROM ordens_compra o
                    JOIN fornecedores f ON f.id=o.fornecedor_id JOIN produtos p ON p.id=o.produto_id
                    ORDER BY o.id DESC""")
    return render_template("ordens.html", ordens=rows)


@app.route("/ordens/nova", methods=["GET", "POST"])
@app.route("/ordens/<int:item_id>/editar", methods=["GET", "POST"])
def nova_ordem(item_id=None):
    ensure_auxiliary_registries()
    ordem = query("SELECT * FROM ordens_compra WHERE id=%s", (item_id,), one=True) if item_id else None
    if request.method == "POST":
        form = request.form
        quantidade = decimal_value(form["quantidade"])
        preco = decimal_value(form["preco_negociado"])
        frete = decimal_value(form.get("frete"))
        total = quantidade * preco + frete
        params = (form["data_preenchimento"], form["tipo_material"], form["fornecedor_id"], form["produto_id"], quantidade, preco, frete, total, form["data_entrega"], form["metodo_pagamento"], form.get("prazo_dias") or 0, form.get("nota_fiscal"), form["status_compra"])
        if item_id:
            execute("""UPDATE ordens_compra SET data_preenchimento=%s,tipo_material=%s,fornecedor_id=%s,produto_id=%s,
                       quantidade=%s,preco_negociado=%s,frete=%s,valor_total=%s,data_entrega=%s,metodo_pagamento=%s,
                       prazo_dias=%s,nota_fiscal=%s,status_compra=%s WHERE id=%s""", params + (item_id,))
        else:
            connection = get_db()
            cursor = connection.cursor()
            try:
                cursor.execute("SELECT COALESCE(MAX(id),0)+1 FROM ordens_compra")
                next_id = cursor.fetchone()[0]
                numero = f"OC-{date.today().year}-{next_id:05d}"
                cursor.execute("""INSERT INTO ordens_compra
                    (numero_oc,data_preenchimento,tipo_material,fornecedor_id,produto_id,quantidade,preco_negociado,frete,
                     valor_total,data_entrega,metodo_pagamento,prazo_dias,nota_fiscal,status_compra)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""", (numero,) + params)
                ordem_id = cursor.lastrowid
                cursor.execute("INSERT INTO autorizacoes (ordem_id,status) VALUES (%s,'pendente')", (ordem_id,))
                cursor.execute("INSERT INTO recebimentos (ordem_id,data_prevista,status) VALUES (%s,%s,'aguardando')", (ordem_id, form["data_entrega"]))
                connection.commit()
            except Exception:
                connection.rollback()
                raise
            finally:
                cursor.close()
                connection.close()
        flash("Ordem de compra salva com sucesso.", "success")
        return redirect(url_for("ordens"))
    return render_template(
        "nova_ordem.html",
        ordem=ordem,
        fornecedores=query("SELECT id,nome FROM fornecedores WHERE status='ativo' ORDER BY nome"),
        produtos=query("SELECT id,codigo,descricao FROM produtos WHERE status='ativo' ORDER BY descricao"),
        tipos_pagamento=query("SELECT nome FROM cadastro_tipos_pagamento WHERE status='ativo' ORDER BY nome"),
    )


@app.post("/ordens/<int:item_id>/excluir")
def excluir_ordem(item_id):
    execute("DELETE FROM ordens_compra WHERE id=%s", (item_id,))
    flash("Ordem excluida.", "success")
    return redirect(url_for("ordens"))


@app.get("/api/produtos/<int:produto_id>/historico")
def historico_produto(produto_id):
    data = query("""SELECT
        (SELECT preco_negociado FROM ordens_compra WHERE produto_id=%s ORDER BY data_preenchimento DESC,id DESC LIMIT 1) ultimo_preco,
        AVG(o.preco_negociado) preco_medio, MAX(o.data_preenchimento) ultima_data,
        (SELECT f.nome FROM ordens_compra x JOIN fornecedores f ON f.id=x.fornecedor_id WHERE x.produto_id=%s ORDER BY x.data_preenchimento DESC,x.id DESC LIMIT 1) ultimo_fornecedor,
        AVG(CASE WHEN r.data_real IS NOT NULL THEN DATEDIFF(r.data_real,o.data_preenchimento) ELSE o.prazo_dias END) prazo_medio
        FROM ordens_compra o LEFT JOIN recebimentos r ON r.ordem_id=o.id WHERE o.produto_id=%s""", (produto_id, produto_id, produto_id), one=True)
    for key, value in data.items():
        if isinstance(value, (Decimal, date, datetime)):
            data[key] = str(value)
    return jsonify(data)


@app.route("/autorizacoes", methods=["GET", "POST"])
def autorizacoes():
    if request.method == "POST":
        form = request.form
        approval_date = form.get("data_aprovacao") or None
        execute("UPDATE autorizacoes SET autorizado_por=%s,data_aprovacao=%s,status=%s,observacao=%s WHERE id=%s", (form.get("autorizado_por"), approval_date, form["status"], form.get("observacao"), form["id"]))
        status_ordem = {"aprovada": "aprovada", "reprovada": "reprovada", "pendente": "aguardando autorizacao"}[form["status"]]
        execute("UPDATE ordens_compra o JOIN autorizacoes a ON a.ordem_id=o.id SET o.status_compra=%s WHERE a.id=%s", (status_ordem, form["id"]))
        flash("Autorizacao atualizada.", "success")
        return redirect(url_for("autorizacoes"))
    rows = query("""SELECT a.*,o.numero_oc,o.valor_total FROM autorizacoes a JOIN ordens_compra o ON o.id=a.ordem_id ORDER BY a.id DESC""")
    return render_template("autorizacoes.html", autorizacoes=rows)


@app.route("/recebimentos", methods=["GET", "POST"])
def recebimentos():
    ensure_nfe_schema()
    if request.method == "POST":
        form = request.form
        real = form.get("data_real") or None
        status = form["status"]
        execute("UPDATE recebimentos SET data_real=%s,status=%s WHERE id=%s", (real, status, form["id"]))
        if status == "entregue completo":
            execute("UPDATE ordens_compra o JOIN recebimentos r ON r.ordem_id=o.id SET o.status_compra='recebida' WHERE r.id=%s", (form["id"],))
        flash("Recebimento atualizado.", "success")
        return redirect(url_for("recebimentos"))
    rows = query("""SELECT r.*,o.numero_oc,
        CASE WHEN r.data_real IS NOT NULL THEN GREATEST(DATEDIFF(r.data_real,r.data_prevista),0)
             WHEN r.data_prevista<CURDATE() AND r.status NOT IN ('entregue parcial','entregue completo') THEN DATEDIFF(CURDATE(),r.data_prevista) ELSE 0 END dias_atraso,
        CASE WHEN r.data_prevista<CURDATE() AND r.data_real IS NULL AND r.status='aguardando' THEN 'atrasado' ELSE r.status END status_exibicao
        FROM recebimentos r JOIN ordens_compra o ON o.id=r.ordem_id ORDER BY r.data_prevista""")
    nfe_rows = query(
        """
        SELECT n.*,o.numero_oc
        FROM nfe_importacoes n
        LEFT JOIN ordens_compra o ON o.id=n.ordem_id
        ORDER BY n.criado_em DESC, n.id DESC
        """
    )
    selected_nfe_id = request.args.get("nfe_id") or (nfe_rows[0]["id"] if nfe_rows else None)
    selected_nfe = None
    nfe_items = []
    produtos_ativos = query("SELECT id,codigo,descricao FROM produtos WHERE status='ativo' ORDER BY descricao")
    ordens_disponiveis = query(
        """
        SELECT o.id,o.numero_oc,o.data_entrega,f.nome fornecedor
        FROM ordens_compra o
        JOIN fornecedores f ON f.id=o.fornecedor_id
        ORDER BY o.id DESC
        """
    )
    if selected_nfe_id:
        selected_nfe = query(
            """
            SELECT n.*,o.numero_oc
            FROM nfe_importacoes n
            LEFT JOIN ordens_compra o ON o.id=n.ordem_id
            WHERE n.id=%s
            """,
            (selected_nfe_id,),
            one=True,
        )
        nfe_items = query(
            """
            SELECT i.*,p.codigo codigo_interno,p.descricao descricao_interna,n.fornecedor,n.numero_nf,n.data_entrada
            FROM nfe_importacao_itens i
            JOIN nfe_importacoes n ON n.id=i.nfe_importacao_id
            LEFT JOIN produtos p ON p.id=i.produto_id
            WHERE i.nfe_importacao_id=%s
            ORDER BY i.id
            """,
            (selected_nfe_id,),
        )
    return render_template(
        "recebimentos.html",
        recebimentos=rows,
        nfe_rows=nfe_rows,
        selected_nfe=selected_nfe,
        nfe_items=nfe_items,
        produtos_ativos=produtos_ativos,
        ordens_disponiveis=ordens_disponiveis,
    )


@app.post("/recebimentos/importar_xml")
def importar_xml_nfe():
    ensure_nfe_schema()
    file = request.files.get("xml")
    if not file or not file.filename:
        flash("Selecione um arquivo XML de NF-e.", "danger")
        return redirect(url_for("recebimentos"))
    if not file.filename.lower().endswith(".xml"):
        flash("Envie um arquivo no formato .xml.", "danger")
        return redirect(url_for("recebimentos"))

    try:
        data = parse_nfe_xml(file)
        exists = query(
            """
            SELECT id FROM nfe_importacoes
            WHERE chave_nfe=%s OR (fornecedor_cnpj=%s AND numero_nf=%s AND COALESCE(serie,'')=COALESCE(%s,''))
            """,
            (data["chave_nfe"], data["fornecedor_cnpj"], data["numero_nf"], data["serie"]),
            one=True,
        )
        if exists:
            flash("Esta Nota Fiscal ja foi importada.", "warning")
            return redirect(url_for("recebimentos", nfe_id=exists["id"]))

        fornecedor_id = get_or_create_nfe_supplier(data["fornecedor"], data["fornecedor_cnpj"])
        identified_items = []
        created_products = 0
        for item in data["items"]:
            product = identify_nfe_product(
                data["fornecedor"],
                data["fornecedor_cnpj"],
                item["codigo_fornecedor"],
                item["descricao"],
            )
            if not product:
                product = create_product_from_nfe_item(item, fornecedor_id)
                created_products += 1
            identified_items.append((item, product))

        connection = get_db()
        cursor = connection.cursor()
        try:
            cursor.execute(
                """
                INSERT INTO nfe_importacoes
                (ordem_id,fornecedor,fornecedor_cnpj,numero_nf,serie,chave_nfe,data_emissao,data_entrada,valor_total)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """,
                (
                    None,
                    data["fornecedor"],
                    data["fornecedor_cnpj"],
                    data["numero_nf"],
                    data["serie"],
                    data["chave_nfe"],
                    data["data_emissao"],
                    data["data_entrada"],
                    data["valor_total"],
                ),
            )
            nfe_id = cursor.lastrowid
            for item, product in identified_items:
                cursor.execute(
                    """
                    INSERT INTO nfe_importacao_itens
                    (nfe_importacao_id,produto_id,codigo_fornecedor,descricao,ncm,cfop,cest,unidade,quantidade,valor_unitario,valor_total,ean,status)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    """,
                    (
                        nfe_id,
                        product["id"],
                        item["codigo_fornecedor"],
                        item["descricao"],
                        item["ncm"],
                        item["cfop"],
                        item["cest"],
                        item["unidade"],
                        item["quantidade"],
                        item["valor_unitario"],
                        item["valor_total"],
                        item["ean"],
                        "vinculado",
                    ),
                )
                cursor.execute(
                    """
                    INSERT INTO produto_fornecedor_relacionamentos
                    (fornecedor,cnpj,codigo_fornecedor,descricao_fornecedor,produto_id)
                    VALUES (%s,%s,%s,%s,%s)
                    ON DUPLICATE KEY UPDATE
                      descricao_fornecedor=VALUES(descricao_fornecedor),
                      produto_id=VALUES(produto_id),
                      atualizado_em=CURRENT_TIMESTAMP
                    """,
                    (
                        data["fornecedor"],
                        data["fornecedor_cnpj"],
                        item["codigo_fornecedor"],
                        item["descricao"],
                        product["id"],
                    ),
                )
            connection.commit()
        except Exception:
            connection.rollback()
            raise
        finally:
            cursor.close()
            connection.close()

        extra = f" {created_products} produto(s) cadastrado(s) automaticamente." if created_products else ""
        flash(f"XML de NF-e importado com sucesso.{extra} Agora vincule a Ordem de Compra.", "success")
        return redirect(url_for("recebimentos", nfe_id=nfe_id))
    except ValueError as error:
        flash(str(error), "danger")
    except mysql.connector.IntegrityError:
        flash("Esta Nota Fiscal ja foi importada.", "warning")
    except Exception as error:
        flash(f"Nao foi possivel importar o XML: {error}", "danger")
    return redirect(url_for("recebimentos"))


@app.post("/recebimentos/nfe_itens/<int:item_id>/vincular")
def vincular_item_nfe(item_id):
    ensure_nfe_schema()
    produto_id = request.form.get("produto_id")
    if not produto_id:
        flash("Selecione um codigo interno para vincular.", "danger")
        return redirect(url_for("recebimentos"))

    item = query(
        """
        SELECT i.*,n.fornecedor,n.fornecedor_cnpj,n.id nfe_id
        FROM nfe_importacao_itens i
        JOIN nfe_importacoes n ON n.id=i.nfe_importacao_id
        WHERE i.id=%s
        """,
        (item_id,),
        one=True,
    )
    product = query("SELECT id,codigo FROM produtos WHERE id=%s", (produto_id,), one=True)
    if not item or not product:
        flash("Item ou produto nao encontrado.", "danger")
        return redirect(url_for("recebimentos"))

    execute(
        """
        INSERT INTO produto_fornecedor_relacionamentos
        (fornecedor,cnpj,codigo_fornecedor,descricao_fornecedor,produto_id)
        VALUES (%s,%s,%s,%s,%s)
        ON DUPLICATE KEY UPDATE
          descricao_fornecedor=VALUES(descricao_fornecedor),
          produto_id=VALUES(produto_id),
          atualizado_em=CURRENT_TIMESTAMP
        """,
        (item["fornecedor"], item["fornecedor_cnpj"], item["codigo_fornecedor"], item["descricao"], produto_id),
    )
    execute("UPDATE nfe_importacao_itens SET produto_id=%s,status='vinculado' WHERE id=%s", (produto_id, item_id))
    flash("Vinculo salvo. Proximas importacoes reconhecerão este produto automaticamente.", "success")
    return redirect(url_for("recebimentos", nfe_id=item["nfe_id"]))


@app.post("/recebimentos/nfe/<int:nfe_id>/vincular_ordem")
def vincular_ordem_nfe(nfe_id):
    ensure_nfe_schema()
    ordem_id = request.form.get("ordem_id")
    if not ordem_id:
        flash("Selecione uma Ordem de Compra para vincular a NF-e.", "danger")
        return redirect(url_for("recebimentos", nfe_id=nfe_id))
    nfe = query("SELECT id,data_entrada FROM nfe_importacoes WHERE id=%s", (nfe_id,), one=True)
    ordem = query("SELECT id FROM ordens_compra WHERE id=%s", (ordem_id,), one=True)
    if not nfe or not ordem:
        flash("NF-e ou Ordem de Compra nao encontrada.", "danger")
        return redirect(url_for("recebimentos", nfe_id=nfe_id))
    execute("UPDATE nfe_importacoes SET ordem_id=%s WHERE id=%s", (ordem_id, nfe_id))
    execute("UPDATE recebimentos SET data_real=%s,status='entregue completo' WHERE ordem_id=%s", (nfe["data_entrada"], ordem_id))
    execute("UPDATE ordens_compra SET status_compra='recebida' WHERE id=%s", (ordem_id,))
    flash("NF-e vinculada a Ordem de Compra.", "success")
    return redirect(url_for("recebimentos", nfe_id=nfe_id))


@app.route("/orcamentos", methods=["GET", "POST"])
def orcamentos():
    if request.method == "POST":
        form = request.form
        budget_id = form.get("id")
        params = (form["categoria"], form["ano"], form["mes"], form["orcamento_previsto"])
        if budget_id:
            execute("UPDATE orcamentos SET categoria=%s,ano=%s,mes=%s,orcamento_previsto=%s WHERE id=%s", params + (budget_id,))
        else:
            execute("INSERT INTO orcamentos (categoria,ano,mes,orcamento_previsto) VALUES (%s,%s,%s,%s)", params)
        flash("Orcamento salvo.", "success")
        return redirect(url_for("orcamentos"))
    rows = query("""SELECT b.*,COALESCE(SUM(o.valor_total),0) total_comprado,
        b.orcamento_previsto-COALESCE(SUM(o.valor_total),0) diferenca
        FROM orcamentos b LEFT JOIN produtos p ON p.categoria=b.categoria
        LEFT JOIN ordens_compra o ON o.produto_id=p.id AND YEAR(o.data_preenchimento)=b.ano AND MONTH(o.data_preenchimento)=b.mes
        GROUP BY b.id ORDER BY b.ano DESC,b.mes DESC,b.categoria""")
    return render_template("orcamentos.html", orcamentos=rows)


@app.post("/orcamentos/<int:item_id>/excluir")
def excluir_orcamento(item_id):
    execute("DELETE FROM orcamentos WHERE id=%s", (item_id,))
    flash("Orcamento excluido.", "success")
    return redirect(url_for("orcamentos"))


if __name__ == "__main__":
    app.run(debug=True)
