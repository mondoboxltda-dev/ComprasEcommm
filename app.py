import os
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
import re
import unicodedata
from xml.etree import ElementTree as ET

import mysql.connector
try:
    import psycopg
    from psycopg.rows import dict_row
except ImportError:
    psycopg = None
    dict_row = None

DB_ERROR_TYPES = (mysql.connector.Error,)
DB_INTEGRITY_ERROR_TYPES = (mysql.connector.IntegrityError,)
if psycopg is not None:
    DB_ERROR_TYPES = DB_ERROR_TYPES + (psycopg.Error,)
    DB_INTEGRITY_ERROR_TYPES = DB_INTEGRITY_ERROR_TYPES + (psycopg.IntegrityError,)

from flask import Flask, flash, g, jsonify, redirect, render_template, request, send_file, session, url_for
from openpyxl import Workbook, load_workbook
from werkzeug.security import check_password_hash, generate_password_hash

from config import Config


app = Flask(__name__)
app.config.from_object(Config)


AUTH_SCHEMA_READY = False
POSTGRES_POOL = None

PERMISSION_MODULES = (
    ("dashboard", "Dashboard"),
    ("cadastros", "Cadastros"),
    ("fornecedores", "Fornecedores"),
    ("produtos", "Produtos"),
    ("historico_compras", "HistÃ³rico de Compras"),
    ("ordens", "Ordens de compra"),
    ("autorizacoes", "AutorizaÃ§Ãµes"),
    ("recebimentos", "Recebimentos"),
    ("orcamentos", "OrÃ§amentos"),
    ("usuarios", "UsuÃ¡rios e permissÃµes"),
)

ENDPOINT_MODULES = {
    "dashboard": "dashboard",
    "cadastros": "cadastros",
    "excluir_cadastro": "cadastros",
    "salvar_orcamento_mensal": "orcamentos",
    "excluir_orcamento_mensal": "orcamentos",
    "baixar_modelo_orcamentos": "orcamentos",
    "importar_orcamentos_mensais": "orcamentos",
    "fornecedores": "fornecedores",
    "baixar_modelo_fornecedores": "fornecedores",
    "excluir_fornecedor": "fornecedores",
    "importar_fornecedores": "fornecedores",
    "produtos": "produtos",
    "baixar_modelo_produtos": "produtos",
    "importar_produtos_xlsx": "produtos",
    "excluir_produto": "produtos",
    "ordens": "ordens",
    "nova_ordem": "ordens",
    "excluir_ordem": "ordens",
    "produtos_por_fornecedor": "ordens",
    "consumo_categoria_produto": "ordens",
    "historico_produto": "produtos",
    "historico_compras": "historico_compras",
    "exportar_historico_compras": "historico_compras",
    "autorizacoes": "autorizacoes",
    "recebimentos": "recebimentos",
    "excluir_recebimento": "recebimentos",
    "importar_xml_nfe": "recebimentos",
    "salvar_recebimento_manual": "recebimentos",
    "cadastrar_produto_rapido_recebimento": "recebimentos",
    "api_ordem_itens": "recebimentos",
    "vincular_item_nfe": "recebimentos",
    "vincular_ordem_nfe": "recebimentos",
    "orcamentos": "orcamentos",
    "excluir_orcamento": "orcamentos",
    "usuarios": "usuarios",
}


DEFAULT_UNITS = (
    ("UN", "Unidade"),
    ("CX", "Caixa"),
    ("PCT", "Pacote"),
    ("KG", "Quilograma"),
    ("G", "Grama"),
    ("L", "Litro"),
    ("ML", "Mililitro"),
    ("M", "Metro"),
    ("ROLO", "Rolo"),
    ("PALLET", "Pallet"),
    ("FARDO", "Fardo"),
    ("KIT", "Kit"),
    ("PAR", "Par"),
    ("M2", "Metro Quadrado (MÂ²)"),
    ("M3", "Metro CÃºbico (MÂ³)"),
)


def using_postgres():
    return bool(app.config.get("DATABASE_URL"))


def _normalize_sql_for_postgres(sql):
    """Compatibilidade leve para consultas MySQL comuns durante a migração."""
    if not using_postgres():
        return sql
    converted = sql
    converted = re.sub(r"\bCURDATE\(\)", "CURRENT_DATE", converted, flags=re.IGNORECASE)
    converted = re.sub(r"\bNOW\(\)", "CURRENT_TIMESTAMP", converted, flags=re.IGNORECASE)
    converted = re.sub(r"YEAR\(([^)]+)\)", r"EXTRACT(YEAR FROM \1)", converted, flags=re.IGNORECASE)
    converted = re.sub(r"MONTH\(([^)]+)\)", r"EXTRACT(MONTH FROM \1)", converted, flags=re.IGNORECASE)
    converted = re.sub(r"DATEDIFF\(([^,]+),\s*([^)]+)\)", r"(\1::date - \2::date)", converted, flags=re.IGNORECASE)
    return converted


def get_db():
    if using_postgres():
        if psycopg is None:
            raise RuntimeError("psycopg nao esta instalado. Rode: pip install -r requirements.txt")
        return psycopg.connect(app.config["DATABASE_URL"], row_factory=dict_row)
    return mysql.connector.connect(
        host=app.config["MYSQL_HOST"],
        port=app.config["MYSQL_PORT"],
        user=app.config["MYSQL_USER"],
        password=app.config["MYSQL_PASSWORD"],
        database=app.config["MYSQL_DATABASE"],
    )


def get_request_db():
    """Reutiliza uma conexão por requisição e evita handshakes repetidos."""
    connection = g.get("db_connection")
    if connection is not None:
        return connection
    if using_postgres():
        if psycopg is None:
            raise RuntimeError("psycopg nao esta instalado. Rode: pip install -r requirements.txt")
        connection = psycopg.connect(
            app.config["DATABASE_URL"],
            row_factory=dict_row,
            prepare_threshold=None,
            connect_timeout=app.config["DB_CONNECT_TIMEOUT"],
        )
    else:
        connection = get_db()
    g.db_connection = connection
    return connection


@app.teardown_appcontext
def close_request_db(_exception=None):
    connection = g.pop("db_connection", None)
    if connection is None:
        return
    try:
        connection.rollback()
    except Exception:
        pass
    connection.close()


def query(sql, params=(), one=False):
    connection = get_request_db()
    sql = _normalize_sql_for_postgres(sql)
    if using_postgres():
        cursor = connection.cursor()
    else:
        cursor = connection.cursor(dictionary=True)
    cursor.execute(sql, params)
    result = cursor.fetchone() if one else cursor.fetchall()
    cursor.close()
    return result


def execute(sql, params=()):
    connection = get_request_db()
    sql = _normalize_sql_for_postgres(sql)
    cursor = connection.cursor()
    try:
        cursor.execute(sql, params)
        inserted_id = None
        if using_postgres() and cursor.description:
            row = cursor.fetchone()
            if row:
                inserted_id = row.get("id") if isinstance(row, dict) else row[0]
        elif using_postgres() and sql.lstrip().upper().startswith("INSERT"):
            try:
                cursor.execute("SELECT LASTVAL() AS id")
                row = cursor.fetchone()
                inserted_id = row.get("id") if isinstance(row, dict) else row[0]
            except Exception:
                inserted_id = None
        else:
            inserted_id = getattr(cursor, "lastrowid", None)
        connection.commit()
        return inserted_id
    except Exception:
        connection.rollback()
        raise
    finally:
        cursor.close()


def ensure_auth_schema():
    global AUTH_SCHEMA_READY
    if AUTH_SCHEMA_READY:
        return
    if using_postgres():
        execute(
            """
            CREATE TABLE IF NOT EXISTS usuarios (
              id SERIAL PRIMARY KEY,
              nome VARCHAR(160) NOT NULL,
              login VARCHAR(80) NOT NULL UNIQUE,
              senha_hash VARCHAR(255) NOT NULL,
              tipo_acesso VARCHAR(30) NOT NULL DEFAULT 'comum',
              status VARCHAR(20) NOT NULL DEFAULT 'ativo',
              ultimo_login TIMESTAMP NULL,
              criado_em TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
              atualizado_em TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
            )
            """
        )
        execute(
            """
            CREATE TABLE IF NOT EXISTS usuario_permissoes (
              id SERIAL PRIMARY KEY,
              usuario_id INT NOT NULL REFERENCES usuarios(id) ON DELETE CASCADE,
              modulo VARCHAR(60) NOT NULL,
              pode_visualizar SMALLINT NOT NULL DEFAULT 1,
              UNIQUE (usuario_id, modulo)
            )
            """
        )
    else:
        execute(
            """
            CREATE TABLE IF NOT EXISTS usuarios (
              id INT AUTO_INCREMENT PRIMARY KEY,
              nome VARCHAR(160) NOT NULL,
              login VARCHAR(80) NOT NULL UNIQUE,
              senha_hash VARCHAR(255) NOT NULL,
              tipo_acesso ENUM('administrador','comum') NOT NULL DEFAULT 'comum',
              status ENUM('ativo','inativo') NOT NULL DEFAULT 'ativo',
              ultimo_login DATETIME NULL,
              criado_em DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
              atualizado_em DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
            """
        )
        if not column_exists("usuarios", "ultimo_login"):
            execute("ALTER TABLE usuarios ADD COLUMN ultimo_login DATETIME NULL AFTER status")
        execute(
            """
            CREATE TABLE IF NOT EXISTS usuario_permissoes (
              id INT AUTO_INCREMENT PRIMARY KEY,
              usuario_id INT NOT NULL,
              modulo VARCHAR(60) NOT NULL,
              pode_visualizar TINYINT(1) NOT NULL DEFAULT 1,
              UNIQUE KEY uk_usuario_modulo (usuario_id,modulo),
              CONSTRAINT fk_usuario_permissao_usuario FOREIGN KEY (usuario_id) REFERENCES usuarios(id) ON DELETE CASCADE
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
            """
        )
    total = query("SELECT COUNT(*) total FROM usuarios", one=True)["total"]
    if not total:
        admin_id = execute(
            """
            INSERT INTO usuarios (nome,login,senha_hash,tipo_acesso,status)
            VALUES (%s,%s,%s,%s,%s)
            """,
            (os.getenv("ADMIN_INITIAL_NAME", "Administrador"), os.getenv("ADMIN_INITIAL_LOGIN", "admin"), generate_password_hash(os.getenv("ADMIN_INITIAL_PASSWORD", "troque-esta-senha")), "administrador", "ativo"),
        )
        for module, _label in PERMISSION_MODULES:
            execute(
                "INSERT INTO usuario_permissoes (usuario_id,modulo,pode_visualizar) VALUES (%s,%s,1)",
                (admin_id, module),
            )
    AUTH_SCHEMA_READY = True

def get_user_permissions(user_id):
    rows = query(
        "SELECT modulo FROM usuario_permissoes WHERE usuario_id=%s AND pode_visualizar=1",
        (user_id,),
    )
    return {row["modulo"] for row in rows}


def endpoint_module(endpoint):
    if endpoint == "cadastros" and request.args.get("modulo") == "orcamentos":
        return "orcamentos"
    return ENDPOINT_MODULES.get(endpoint)


def can_access(module):
    user = getattr(g, "current_user", None)
    if not module or not user:
        return False
    if user["tipo_acesso"] == "administrador":
        return True
    return module in getattr(g, "user_permissions", set())


def first_allowed_url():
    for module, _label in PERMISSION_MODULES:
        if can_access(module):
            if module == "orcamentos":
                return url_for("cadastros", modulo="orcamentos")
            if module == "usuarios":
                return url_for("usuarios")
            if module == "historico_compras":
                return url_for("historico_compras")
            if module == "ordens":
                return url_for("ordens")
            if module == "recebimentos":
                return url_for("recebimentos")
            if module == "autorizacoes":
                return url_for("autorizacoes")
            return url_for(module)
    return url_for("logout")


@app.before_request
def require_login():
    if request.endpoint in ("login", "static"):
        return None
    ensure_auth_schema()
    user_id = session.get("user_id")
    if not user_id:
        return redirect(url_for("login", next=request.full_path if request.query_string else request.path))
    user = query(
        "SELECT id,nome,login,tipo_acesso,status FROM usuarios WHERE id=%s",
        (user_id,),
        one=True,
    )
    if not user or user["status"] != "ativo":
        session.clear()
        flash("Acesso encerrado. Faca login novamente.", "warning")
        return redirect(url_for("login"))
    g.current_user = user
    g.user_permissions = {module for module, _label in PERMISSION_MODULES} if user["tipo_acesso"] == "administrador" else get_user_permissions(user["id"])
    module = endpoint_module(request.endpoint)
    if module and not can_access(module):
        if request.path.startswith("/api/"):
            return jsonify({"error": "Acesso nao autorizado."}), 403
        flash("Voce nao tem permissao para acessar esta tela.", "warning")
        return redirect(first_allowed_url())
    return None


def decimal_value(value):
    return Decimal(str(value or 0))


def money_value(value):
    if value is None or value == "":
        return Decimal("0")
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))
    text = str(value).strip().replace("R$", "").replace(" ", "")
    if "," in text:
        text = text.replace(".", "").replace(",", ".")
    return Decimal(text)


def parse_optional_date_br(value):
    if not value:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValueError("Data deve estar no formato DD/MM/AAAA.")


def normalize_iso_date(value, field_name="Data", default_today=False):
    if value is None or str(value).strip() == "":
        if default_today:
            return date.today().isoformat()
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    try:
        return datetime.strptime(text, "%Y-%m-%d").date().isoformat()
    except ValueError as error:
        raise ValueError(f"{field_name} deve estar no formato AAAA-MM-DD.") from error


def infer_purchase_conversion(product):
    code = (product.get("codigo") or "").upper().replace(" ", "")
    description = (product.get("descricao") or "").upper()
    unit = (product.get("unidade") or "").upper()
    if unit not in ("PALLET", "CX", "CAIXA", "ROLO"):
        return None
    match = re.search(r"(\d+(?:[,.]\d+)?)\s*(?:CHAPA|CHAPAS|PECA|PECAS|P[EÃ‡]A|P[EÃ‡]AS)", description)
    if match:
        factor = money_value(match.group(1))
        base_unit = "CHAPAS" if "CHAPA" in match.group(0) else "PECAS"
        return factor, base_unit
    if "MDF4MM" in code:
        return Decimal("126"), "CHAPAS"
    if "MDF2MM" in code:
        return Decimal("168"), "CHAPAS"
    return None


def normalize_purchase_unit(unit):
    value = (unit or "").strip().upper()
    aliases = {
        "PT": "PALLET",
        "PL": "PALLET",
        "PAL": "PALLET",
        "CH": "CHAPAS",
        "CHAPA": "CHAPAS",
        "PC": "PECAS",
        "PECA": "PECAS",
        "PEÃ‡A": "PECAS",
        "UNID": "UN",
        "UNIDADE": "UN",
    }
    return aliases.get(value, value or "UN")


def conversion_for_receipt_item(product, item_unit):
    product_unit = normalize_purchase_unit(product.get("unidade"))
    purchase_unit = normalize_purchase_unit(item_unit or product.get("unidade"))
    product_factor = decimal_value(product.get("quantidade_por_unidade_compra") or 1)
    if product_factor <= 0:
        product_factor = Decimal("1")
    if purchase_unit == product_unit:
        return purchase_unit, product_factor
    if purchase_unit in ("PALLET", "CX", "CAIXA", "ROLO"):
        return purchase_unit, product_factor
    return purchase_unit, Decimal("1")


def import_cell_value(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).strip()


def normalize_header(value):
    text = unicodedata.normalize("NFKD", import_cell_value(value))
    text = "".join(char for char in text if not unicodedata.combining(char))
    return re.sub(r"\s+", " ", text).strip().lower()


def normalize_cnpj(value):
    digits = re.sub(r"\D", "", value or "")
    if not digits:
        return ""
    if len(digits) == 14:
        return f"{digits[:2]}.{digits[2:5]}.{digits[5:8]}/{digits[8:12]}-{digits[12:]}"
    return value.strip()


def is_valid_cnpj(value):
    digits = re.sub(r"\D", "", value or "")
    if len(digits) != 14 or digits == digits[0] * 14:
        return False

    def check_digit(base, weights):
        total = sum(int(number) * weight for number, weight in zip(base, weights))
        remainder = total % 11
        return "0" if remainder < 2 else str(11 - remainder)

    first = check_digit(digits[:12], [5, 4, 3, 2, 9, 8, 7, 6, 5, 4, 3, 2])
    second = check_digit(digits[:13], [6, 5, 4, 3, 2, 9, 8, 7, 6, 5, 4, 3, 2])
    return digits[-2:] == first + second


def ensure_supplier_schema():
    if using_postgres():
        return
    column = query(
        """
        SELECT IS_NULLABLE
        FROM information_schema.COLUMNS
        WHERE TABLE_SCHEMA=%s AND TABLE_NAME='fornecedores' AND COLUMN_NAME='cnpj'
        """,
        (app.config["MYSQL_DATABASE"],),
        one=True,
    )
    if column and column["IS_NULLABLE"] == "NO":
        execute("ALTER TABLE fornecedores MODIFY cnpj VARCHAR(18) NULL")


def ensure_nfe_schema():
    if using_postgres():
        return
    execute(
        """
        CREATE TABLE IF NOT EXISTS produto_fornecedor_relacionamentos (
          id INT AUTO_INCREMENT PRIMARY KEY,
          fornecedor VARCHAR(150) NOT NULL,
          cnpj VARCHAR(18),
          codigo_fornecedor VARCHAR(80) NOT NULL,
          descricao_fornecedor VARCHAR(255) NOT NULL,
          produto_id INT NOT NULL,
          criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
          atualizado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
          CONSTRAINT fk_rel_produto FOREIGN KEY (produto_id) REFERENCES produtos(id),
          UNIQUE KEY uk_rel_cnpj_codigo (cnpj,codigo_fornecedor),
          KEY idx_rel_fornecedor_codigo (fornecedor,codigo_fornecedor)
        )
        """
    )
    execute(
        """
        CREATE TABLE IF NOT EXISTS nfe_importacoes (
          id INT AUTO_INCREMENT PRIMARY KEY,
          ordem_id INT NULL,
          fornecedor VARCHAR(150) NOT NULL,
          fornecedor_cnpj VARCHAR(18) NOT NULL,
          numero_nf VARCHAR(30) NOT NULL,
          serie VARCHAR(10),
          chave_nfe VARCHAR(60) NOT NULL UNIQUE,
          data_emissao DATE,
          data_entrada DATE NOT NULL,
          valor_total DECIMAL(14,2) NOT NULL DEFAULT 0,
          criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
          CONSTRAINT fk_nfe_ordem FOREIGN KEY (ordem_id) REFERENCES ordens_compra(id),
          UNIQUE KEY uk_nfe_fornecedor_numero_serie (fornecedor_cnpj,numero_nf,serie)
        )
        """
    )
    if not column_exists("nfe_importacoes", "ordem_id"):
        execute("ALTER TABLE nfe_importacoes ADD COLUMN ordem_id INT NULL AFTER id")
        constraint = query(
            """
            SELECT CONSTRAINT_NAME
            FROM information_schema.TABLE_CONSTRAINTS
            WHERE TABLE_SCHEMA=%s AND TABLE_NAME='nfe_importacoes' AND CONSTRAINT_NAME='fk_nfe_ordem'
            """,
            (app.config["MYSQL_DATABASE"],),
            one=True,
        )
        if not constraint:
            execute("ALTER TABLE nfe_importacoes ADD CONSTRAINT fk_nfe_ordem FOREIGN KEY (ordem_id) REFERENCES ordens_compra(id)")
    nfe_columns = (
        ("origem", "ALTER TABLE nfe_importacoes ADD COLUMN origem ENUM('XML','Manual') NOT NULL DEFAULT 'XML' AFTER ordem_id"),
        ("tipo_documento", "ALTER TABLE nfe_importacoes ADD COLUMN tipo_documento VARCHAR(40) NOT NULL DEFAULT 'Nota Fiscal' AFTER numero_nf"),
        ("condicao_pagamento", "ALTER TABLE nfe_importacoes ADD COLUMN condicao_pagamento VARCHAR(120) NULL AFTER valor_total"),
        ("observacoes", "ALTER TABLE nfe_importacoes ADD COLUMN observacoes TEXT NULL AFTER condicao_pagamento"),
        ("data_entrega", "ALTER TABLE nfe_importacoes ADD COLUMN data_entrega DATE NULL AFTER data_entrada"),
        ("tipo_entrega", "ALTER TABLE nfe_importacoes ADD COLUMN tipo_entrega ENUM('completa','parcial') NULL AFTER data_entrega"),
        ("ordem_parcial_id", "ALTER TABLE nfe_importacoes ADD COLUMN ordem_parcial_id INT NULL AFTER tipo_entrega"),
        ("usuario_responsavel", "ALTER TABLE nfe_importacoes ADD COLUMN usuario_responsavel VARCHAR(120) NOT NULL DEFAULT 'Sistema' AFTER observacoes"),
        ("status_recebimento", "ALTER TABLE nfe_importacoes ADD COLUMN status_recebimento VARCHAR(60) NOT NULL DEFAULT 'Pendente' AFTER usuario_responsavel"),
        ("confirmado_em", "ALTER TABLE nfe_importacoes ADD COLUMN confirmado_em DATETIME NULL AFTER status_recebimento"),
    )
    for column, statement in nfe_columns:
        if not column_exists("nfe_importacoes", column):
            execute(statement)
    execute(
        """
        CREATE TABLE IF NOT EXISTS nfe_importacao_itens (
          id INT AUTO_INCREMENT PRIMARY KEY,
          nfe_importacao_id INT NOT NULL,
          produto_id INT NULL,
          codigo_fornecedor VARCHAR(80),
          descricao VARCHAR(255) NOT NULL,
          ncm VARCHAR(20),
          cfop VARCHAR(10),
          cest VARCHAR(20),
          unidade VARCHAR(20),
          quantidade DECIMAL(14,4) NOT NULL DEFAULT 0,
          valor_unitario DECIMAL(14,4) NOT NULL DEFAULT 0,
          valor_total DECIMAL(14,2) NOT NULL DEFAULT 0,
          ean VARCHAR(30),
          status ENUM('vinculado','pendente') NOT NULL DEFAULT 'pendente',
          CONSTRAINT fk_nfe_item_importacao FOREIGN KEY (nfe_importacao_id) REFERENCES nfe_importacoes(id) ON DELETE CASCADE,
          CONSTRAINT fk_nfe_item_produto FOREIGN KEY (produto_id) REFERENCES produtos(id)
        )
        """
    )
    if not column_exists("nfe_importacao_itens", "desconto"):
        execute("ALTER TABLE nfe_importacao_itens ADD COLUMN desconto DECIMAL(14,2) NOT NULL DEFAULT 0 AFTER valor_unitario")
    execute(
        """
        CREATE TABLE IF NOT EXISTS movimentacoes_estoque (
          id INT AUTO_INCREMENT PRIMARY KEY,
          produto_id INT NOT NULL,
          recebimento_id INT NOT NULL,
          recebimento_item_id INT NOT NULL,
          data_movimentacao DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
          tipo VARCHAR(30) NOT NULL,
          quantidade DECIMAL(14,4) NOT NULL,
          valor_unitario DECIMAL(14,4) NOT NULL DEFAULT 0,
          valor_total DECIMAL(14,2) NOT NULL DEFAULT 0,
          origem VARCHAR(20) NOT NULL,
          usuario VARCHAR(120) NOT NULL DEFAULT 'Sistema',
          UNIQUE KEY uk_mov_item (recebimento_item_id,tipo),
          CONSTRAINT fk_mov_produto FOREIGN KEY (produto_id) REFERENCES produtos(id),
          CONSTRAINT fk_mov_recebimento FOREIGN KEY (recebimento_id) REFERENCES nfe_importacoes(id),
          CONSTRAINT fk_mov_item FOREIGN KEY (recebimento_item_id) REFERENCES nfe_importacao_itens(id) ON DELETE CASCADE
        )
        """
    )
    execute(
        """
        CREATE TABLE IF NOT EXISTS historico_custos (
          id INT AUTO_INCREMENT PRIMARY KEY,
          produto_id INT NOT NULL,
          fornecedor VARCHAR(150) NOT NULL,
          documento VARCHAR(60),
          data_entrada DATE NOT NULL,
          quantidade DECIMAL(14,4) NOT NULL,
          valor_unitario DECIMAL(14,4) NOT NULL,
          valor_total DECIMAL(14,2) NOT NULL,
          origem VARCHAR(20) NOT NULL,
          recebimento_id INT NOT NULL,
          recebimento_item_id INT NOT NULL,
          criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
          UNIQUE KEY uk_custo_item (recebimento_item_id,origem),
          CONSTRAINT fk_custo_produto FOREIGN KEY (produto_id) REFERENCES produtos(id),
          CONSTRAINT fk_custo_recebimento FOREIGN KEY (recebimento_id) REFERENCES nfe_importacoes(id),
          CONSTRAINT fk_custo_item FOREIGN KEY (recebimento_item_id) REFERENCES nfe_importacao_itens(id) ON DELETE CASCADE
        )
        """
    )
    if table_exists("historico_custos"):
        if not column_exists("historico_custos", "unidade_compra"):
            execute("ALTER TABLE historico_custos ADD COLUMN unidade_compra VARCHAR(30) NULL AFTER quantidade")
        if not column_exists("historico_custos", "quantidade_por_unidade_compra"):
            execute("ALTER TABLE historico_custos ADD COLUMN quantidade_por_unidade_compra DECIMAL(14,4) NOT NULL DEFAULT 1 AFTER unidade_compra")
        if not column_exists("historico_custos", "unidade_base"):
            execute("ALTER TABLE historico_custos ADD COLUMN unidade_base VARCHAR(30) NULL AFTER quantidade_por_unidade_compra")
        if not column_exists("historico_custos", "quantidade_total_base"):
            execute("ALTER TABLE historico_custos ADD COLUMN quantidade_total_base DECIMAL(14,4) NOT NULL DEFAULT 0 AFTER unidade_base")
        if not column_exists("historico_custos", "valor_unitario_base"):
            execute("ALTER TABLE historico_custos ADD COLUMN valor_unitario_base DECIMAL(14,4) NOT NULL DEFAULT 0 AFTER valor_unitario")
    execute(
        """
        CREATE TABLE IF NOT EXISTS recebimento_auditoria (
          id INT AUTO_INCREMENT PRIMARY KEY,
          recebimento_id INT NOT NULL,
          usuario VARCHAR(120) NOT NULL DEFAULT 'Sistema',
          acao VARCHAR(80) NOT NULL,
          origem VARCHAR(20) NOT NULL,
          detalhes TEXT,
          criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
          CONSTRAINT fk_audit_recebimento FOREIGN KEY (recebimento_id) REFERENCES nfe_importacoes(id) ON DELETE CASCADE
        )
        """
    )


def nfe_text(parent, path, namespaces=None, default=""):
    node = parent.find(path, namespaces or {})
    return (node.text or "").strip() if node is not None and node.text is not None else default


def nfe_decimal(value):
    return Decimal(str(value or "0").replace(",", "."))


def parse_nfe_xml(file_storage):
    try:
        tree = ET.parse(file_storage)
    except ET.ParseError as error:
        raise ValueError(f"XML invalido: {error}") from error

    root = tree.getroot()
    namespace = root.tag.split("}")[0].strip("{") if root.tag.startswith("{") else ""
    ns = {"n": namespace} if namespace else {}
    prefix = "n:" if namespace else ""
    inf = root.find(f".//{prefix}infNFe", ns)
    if inf is None:
        raise ValueError("O arquivo XML nao parece ser uma NF-e.")

    ide = inf.find(f"{prefix}ide", ns)
    emit = inf.find(f"{prefix}emit", ns)
    total = inf.find(f"{prefix}total/{prefix}ICMSTot", ns)
    if ide is None or emit is None:
        raise ValueError("Nao foi possivel ler cabecalho da NF-e.")

    chave = (inf.attrib.get("Id") or "").replace("NFe", "")
    numero_nf = nfe_text(ide, f"{prefix}nNF", ns)
    serie = nfe_text(ide, f"{prefix}serie", ns)
    emissao_raw = nfe_text(ide, f"{prefix}dhEmi", ns) or nfe_text(ide, f"{prefix}dEmi", ns)
    data_emissao = emissao_raw[:10] if emissao_raw else None
    fornecedor = nfe_text(emit, f"{prefix}xNome", ns)
    cnpj = normalize_cnpj(nfe_text(emit, f"{prefix}CNPJ", ns))
    telefone = nfe_text(emit, f"{prefix}enderEmit/{prefix}fone", ns)
    valor_total = nfe_decimal(nfe_text(total, f"{prefix}vNF", ns) if total is not None else "0")

    if not chave or not numero_nf or not fornecedor or not cnpj:
        raise ValueError("NF-e sem chave, numero, fornecedor ou CNPJ.")

    items = []
    for det in inf.findall(f"{prefix}det", ns):
        prod = det.find(f"{prefix}prod", ns)
        if prod is None:
            continue
        items.append(
            {
                "codigo_fornecedor": nfe_text(prod, f"{prefix}cProd", ns),
                "descricao": nfe_text(prod, f"{prefix}xProd", ns),
                "ncm": nfe_text(prod, f"{prefix}NCM", ns),
                "cfop": nfe_text(prod, f"{prefix}CFOP", ns),
                "cest": nfe_text(prod, f"{prefix}CEST", ns),
                "unidade": nfe_text(prod, f"{prefix}uCom", ns),
                "quantidade": nfe_decimal(nfe_text(prod, f"{prefix}qCom", ns)),
                "valor_unitario": nfe_decimal(nfe_text(prod, f"{prefix}vUnCom", ns)),
                "valor_total": nfe_decimal(nfe_text(prod, f"{prefix}vProd", ns)),
                "ean": nfe_text(prod, f"{prefix}cEAN", ns),
            }
        )
    if not items:
        raise ValueError("NF-e sem produtos para importar.")

    return {
        "fornecedor": fornecedor,
        "fornecedor_cnpj": cnpj,
        "fornecedor_telefone": telefone,
        "numero_nf": numero_nf,
        "serie": serie,
        "chave_nfe": chave,
        "data_emissao": data_emissao,
        "valor_total": valor_total,
        "items": items,
    }


def audit_receipt(recebimento_id, acao, detalhes="", usuario="Sistema", origem=None):
    receipt = query("SELECT origem FROM nfe_importacoes WHERE id=%s", (recebimento_id,), one=True)
    execute(
        """
        INSERT INTO recebimento_auditoria (recebimento_id,usuario,acao,origem,detalhes)
        VALUES (%s,%s,%s,%s,%s)
        """,
        (recebimento_id, usuario or "Sistema", acao, origem or (receipt["origem"] if receipt else "Manual"), detalhes),
    )


def apply_receipt_effects(recebimento_id):
    ensure_product_schema()
    ensure_nfe_schema()
    receipt = query("SELECT * FROM nfe_importacoes WHERE id=%s", (recebimento_id,), one=True)
    if not receipt:
        return {"processed": 0}
    items = query(
        """
        SELECT i.*
        FROM nfe_importacao_itens i
        WHERE i.nfe_importacao_id=%s AND i.produto_id IS NOT NULL
        ORDER BY i.id
        """,
        (recebimento_id,),
    )
    processed = 0
    for item in items:
        exists = query(
            "SELECT id FROM movimentacoes_estoque WHERE recebimento_item_id=%s AND tipo='entrada'",
            (item["id"],),
            one=True,
        )
        if exists:
            continue
        product = query(
            """
            SELECT estoque_atual,custo_atual,unidade,quantidade_por_unidade_compra,unidade_base
            FROM produtos
            WHERE id=%s
            """,
            (item["produto_id"],),
            one=True,
        )
        if not product:
            continue
        qty = decimal_value(item["quantidade"])
        unit_cost = decimal_value(item["valor_unitario"])
        total_value = decimal_value(item["valor_total"])
        purchase_unit, purchase_factor = conversion_for_receipt_item(product, item.get("unidade"))
        base_qty = qty * purchase_factor
        base_unit_cost = total_value / base_qty if base_qty > 0 else unit_cost
        current_qty = decimal_value(product["estoque_atual"])
        current_cost = decimal_value(product["custo_atual"])
        new_qty = current_qty + qty
        new_cost = ((current_qty * current_cost) + total_value) / new_qty if new_qty > 0 else unit_cost
        execute(
            "UPDATE produtos SET estoque_atual=%s,custo_atual=%s WHERE id=%s",
            (new_qty, new_cost, item["produto_id"]),
        )
        execute(
            """
            INSERT INTO movimentacoes_estoque
            (produto_id,recebimento_id,recebimento_item_id,tipo,quantidade,valor_unitario,valor_total,origem,usuario)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """,
            (
                item["produto_id"],
                recebimento_id,
                item["id"],
                "entrada",
                qty,
                unit_cost,
                total_value,
                receipt["origem"],
                receipt.get("usuario_responsavel") or "Sistema",
            ),
        )
        execute(
            """
            INSERT INTO historico_custos
            (produto_id,fornecedor,documento,data_entrada,quantidade,unidade_compra,quantidade_por_unidade_compra,
             unidade_base,quantidade_total_base,valor_unitario,valor_unitario_base,valor_total,origem,recebimento_id,recebimento_item_id)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """,
            (
                item["produto_id"],
                receipt["fornecedor"],
                receipt["numero_nf"],
                receipt["data_entrada"],
                qty,
                purchase_unit,
                purchase_factor,
                product.get("unidade_base") or product.get("unidade") or item.get("unidade") or "UN",
                base_qty,
                unit_cost,
                base_unit_cost,
                total_value,
                receipt["origem"],
                recebimento_id,
                item["id"],
            ),
        )
        processed += 1
    audit_receipt(recebimento_id, "Processamento de estoque/custo", f"{processed} item(ns) processado(s).", receipt.get("usuario_responsavel") or "Sistema", receipt["origem"])
    return {"processed": processed}


def calculate_receipt_status(recebimento_id):
    receipt = query(
        "SELECT id,origem,ordem_id,status_recebimento FROM nfe_importacoes WHERE id=%s",
        (recebimento_id,),
        one=True,
    )
    if not receipt:
        return None
    pending = query(
        """
        SELECT COUNT(1) total
        FROM nfe_importacao_itens
        WHERE nfe_importacao_id=%s AND status='pendente'
        """,
        (recebimento_id,),
        one=True,
    )["total"]
    if pending:
        return "Produto Pendente de Cadastro"
    if receipt["status_recebimento"] == "Recebido Parcialmente":
        return "Recebido Parcialmente"
    if receipt["origem"] == "XML" and not receipt["ordem_id"]:
        return "Aguardando Ordem de Compra"
    return "Recebido"


def refresh_receipt_status(recebimento_id):
    status = calculate_receipt_status(recebimento_id)
    if status:
        execute("UPDATE nfe_importacoes SET status_recebimento=%s WHERE id=%s", (status, recebimento_id))
    return status


def refresh_receipt_statuses():
    for receipt in query("SELECT id FROM nfe_importacoes"):
        refresh_receipt_status(receipt["id"])


def recalculate_products_after_receipt_delete(cursor, product_ids):
    for product_id in product_ids:
        cursor.execute(
            """
            SELECT
              COALESCE(SUM(quantidade),0) estoque,
              COALESCE(SUM(valor_total),0) valor_total,
              COALESCE(SUM(quantidade),0) quantidade_total
            FROM historico_custos
            WHERE produto_id=%s
            """,
            (product_id,),
        )
        totals = cursor.fetchone()
        stock = decimal_value(totals["estoque"] if totals else 0)
        total_value = decimal_value(totals["valor_total"] if totals else 0)
        total_qty = decimal_value(totals["quantidade_total"] if totals else 0)
        current_cost = total_value / total_qty if total_qty > 0 else Decimal("0")
        cursor.execute(
            "UPDATE produtos SET estoque_atual=%s,custo_atual=%s WHERE id=%s",
            (stock, current_cost, product_id),
        )


def update_order_receipt_agenda(ordem_id, data_entrada, received_qty=None):
    if not ordem_id:
        return
    ordem = query(
        """
        SELECT o.id,COALESCE(SUM(i.quantidade),o.quantidade) quantidade
        FROM ordens_compra o
        LEFT JOIN ordem_compra_itens i ON i.ordem_id=o.id
        WHERE o.id=%s
        GROUP BY o.id
        """,
        (ordem_id,),
        one=True,
    )
    if not ordem:
        raise ValueError("Pedido de compra informado nao existe.")
    if received_qty is not None and decimal_value(received_qty) < decimal_value(ordem["quantidade"]):
        execute("UPDATE ordens_compra SET status_compra='em compra' WHERE id=%s", (ordem_id,))
        execute("UPDATE recebimentos SET data_real=%s,status='entregue parcial' WHERE ordem_id=%s", (data_entrada, ordem_id))
        return
    execute("UPDATE ordens_compra SET status_compra='recebida' WHERE id=%s", (ordem_id,))
    execute("UPDATE recebimentos SET data_real=%s,status='entregue completo' WHERE ordem_id=%s", (data_entrada, ordem_id))


def delivered_quantities_from_form(form):
    item_ids = form.getlist("delivered_item_id[]")
    quantities = form.getlist("delivered_quantidade[]")
    delivered = {}
    for index, item_id in enumerate(item_ids):
        if not item_id:
            continue
        delivered[int(item_id)] = decimal_value(quantities[index] if index < len(quantities) else 0)
    return delivered


def order_delivery_items(ordem_id):
    return query(
        """
        SELECT i.*,p.descricao produto,p.codigo,
               COALESCE(i.unidade_compra,p.unidade,'UN') unidade
        FROM ordem_compra_itens i
        JOIN produtos p ON p.id=i.produto_id
        WHERE i.ordem_id=%s
        ORDER BY i.id
        """,
        (ordem_id,),
    )


def next_partial_order_number(original_number):
    base = f"{original_number}-PARCIAL"
    existing = query(
        "SELECT numero_oc FROM ordens_compra WHERE numero_oc=%s OR numero_oc LIKE %s ORDER BY id DESC",
        (base, f"{base}-%"),
    )
    if not existing:
        return base
    return f"{base}-{len(existing) + 1}"


def create_partial_purchase_order(original_order, remaining_items, new_delivery_date):
    first = remaining_items[0]
    numero = next_partial_order_number(original_order["numero_oc"])
    total = sum(item["valor_total_item"] for item in remaining_items)
    total_freight = sum(item["frete"] for item in remaining_items)
    ordem_id = execute(
        """
        INSERT INTO ordens_compra
        (numero_oc,data_preenchimento,tipo_material,fornecedor_id,produto_id,quantidade,preco_negociado,frete,
         valor_total,data_entrega,metodo_pagamento,parcelas,prazos_parcelas,prazo_dias,nota_fiscal,status_compra,ordem_original_id)
        VALUES (%s,CURDATE(),%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """,
        (
            numero,
            first.get("categoria") or original_order.get("tipo_material"),
            original_order["fornecedor_id"],
            first["produto_id"],
            first["quantidade"],
            first["preco_negociado"],
            total_freight,
            total,
            new_delivery_date,
            original_order["metodo_pagamento"],
            original_order.get("parcelas") or 1,
            original_order.get("prazos_parcelas"),
            original_order.get("prazo_dias") or 0,
            original_order.get("nota_fiscal"),
            "parcial",
            original_order["id"],
        ),
    )
    for item in remaining_items:
        execute(
            """
            INSERT INTO ordem_compra_itens
            (ordem_id,produto_id,quantidade,preco_negociado,frete,valor_total_item,categoria,unidade_compra,quantidade_por_unidade,unidade_base)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """,
            (
                ordem_id,
                item["produto_id"],
                item["quantidade"],
                item["preco_negociado"],
                item["frete"],
                item["valor_total_item"],
                item.get("categoria"),
                item.get("unidade_compra"),
                item.get("quantidade_por_unidade") or 1,
                item.get("unidade_base"),
            ),
        )
    execute("INSERT INTO recebimentos (ordem_id,data_prevista,status) VALUES (%s,%s,'aguardando')", (ordem_id, new_delivery_date))
    return {"id": ordem_id, "numero_oc": numero}


def process_order_delivery(ordem_id, receipt_id, tipo_entrega, data_entrega, delivered_quantities=None, data_entrega_restante=None, documento=None):
    if not ordem_id:
        return None
    ensure_order_schema()
    if tipo_entrega not in ("completa", "parcial"):
        raise ValueError("Informe se a entrega e completa ou parcial.")
    original = query("SELECT * FROM ordens_compra WHERE id=%s", (ordem_id,), one=True)
    if not original:
        raise ValueError("Pedido de compra informado nao existe.")
    items = order_delivery_items(ordem_id)
    if not items:
        raise ValueError("Pedido de compra sem produtos vinculados.")
    if tipo_entrega == "completa":
        execute("UPDATE ordens_compra SET status_compra='entregue completo' WHERE id=%s", (ordem_id,))
        execute("UPDATE recebimentos SET data_real=%s,status='entregue completo' WHERE ordem_id=%s", (data_entrega, ordem_id))
        audit_receipt(receipt_id, "Entrega completa", f"Ordem {original['numero_oc']} entregue completamente. Documento {documento or '-'}.", "Sistema")
        return {"tipo": "completa", "partial_order": None}

    if not data_entrega_restante:
        raise ValueError("Informe a nova data de entrega do restante da ordem.")
    delivered_quantities = delivered_quantities or {}
    remaining_items = []
    audit_lines = []
    has_remaining = False
    for item in items:
        original_qty = decimal_value(item["quantidade"])
        delivered_qty = delivered_quantities.get(int(item["id"]), Decimal("0"))
        if delivered_qty < 0:
            raise ValueError("Quantidade entregue nao pode ser negativa.")
        if delivered_qty > original_qty:
            raise ValueError(f"Quantidade entregue maior que a quantidade original do item {item['produto']}.")
        remaining_qty = original_qty - delivered_qty
        has_remaining = has_remaining or remaining_qty > 0
        audit_lines.append(f"{item['produto']}: entregue {delivered_qty}, restante {remaining_qty}")
        if remaining_qty > 0:
            ratio = remaining_qty / original_qty if original_qty > 0 else Decimal("0")
            remaining_freight = decimal_value(item["frete"]) * ratio
            remaining_items.append(
                {
                    **item,
                    "quantidade": remaining_qty,
                    "frete": remaining_freight,
                    "valor_total_item": (remaining_qty * decimal_value(item["preco_negociado"])) + remaining_freight,
                }
            )
    if not has_remaining:
        raise ValueError("Para entrega parcial, pelo menos um item precisa ter quantidade restante.")
    partial_order = create_partial_purchase_order(original, remaining_items, data_entrega_restante)
    execute("UPDATE ordens_compra SET status_compra='entrega parcial recebida' WHERE id=%s", (ordem_id,))
    execute("UPDATE recebimentos SET data_real=%s,status='entregue parcial' WHERE ordem_id=%s", (data_entrega, ordem_id))
    audit_receipt(
        receipt_id,
        "Entrega parcial",
        f"Ordem original {original['numero_oc']}. Documento {documento or '-'}. {' | '.join(audit_lines)}. Nova ordem parcial {partial_order['numero_oc']} com entrega em {data_entrega_restante}.",
        "Sistema",
    )
    return {"tipo": "parcial", "partial_order": partial_order}


def serialize_row(row):
    serialized = {}
    for key, value in row.items():
        if isinstance(value, Decimal):
            serialized[key] = float(value)
        elif isinstance(value, (date, datetime)):
            serialized[key] = value.isoformat()
        else:
            serialized[key] = value
    return serialized


def purchase_history_where(filters):
    where = []
    params = []
    if filters.get("codigo"):
        where.append("p.codigo LIKE %s")
        params.append(f"%{filters['codigo']}%")
    if filters.get("produto"):
        where.append("p.descricao LIKE %s")
        params.append(f"%{filters['produto']}%")
    if filters.get("fornecedor"):
        where.append("h.fornecedor LIKE %s")
        params.append(f"%{filters['fornecedor']}%")
    if filters.get("categoria"):
        where.append("p.categoria=%s")
        params.append(filters["categoria"])
    if filters.get("nf"):
        where.append("h.documento LIKE %s")
        params.append(f"%{filters['nf']}%")
    if filters.get("data_inicio"):
        where.append("h.data_entrada >= %s")
        params.append(filters["data_inicio"])
    if filters.get("data_fim"):
        where.append("h.data_entrada <= %s")
        params.append(filters["data_fim"])
    if filters.get("unidade"):
        where.append("p.unidade=%s")
        params.append(filters["unidade"])
    if filters.get("preco_min"):
        where.append("h.valor_unitario >= %s")
        params.append(money_value(filters["preco_min"]))
    if filters.get("preco_max"):
        where.append("h.valor_unitario <= %s")
        params.append(money_value(filters["preco_max"]))
    if filters.get("origem"):
        where.append("h.origem=%s")
        params.append(filters["origem"])
    if filters.get("status"):
        where.append("n.status_recebimento=%s")
        params.append(filters["status"])
    return ("WHERE " + " AND ".join(where)) if where else "", params


def request_purchase_history_filters():
    return {
        "codigo": request.args.get("codigo") or "",
        "produto": request.args.get("produto") or "",
        "fornecedor": request.args.get("fornecedor") or "",
        "categoria": request.args.get("categoria") or "",
        "nf": request.args.get("nf") or "",
        "data_inicio": request.args.get("data_inicio") or "",
        "data_fim": request.args.get("data_fim") or "",
        "unidade": request.args.get("unidade") or "",
        "preco_min": request.args.get("preco_min") or "",
        "preco_max": request.args.get("preco_max") or "",
        "origem": request.args.get("origem") or "",
        "status": request.args.get("status") or "",
    }


def identify_nfe_product(fornecedor, cnpj, codigo_fornecedor, descricao):
    rel = query(
        """
        SELECT p.id,p.codigo
        FROM produto_fornecedor_relacionamentos r
        JOIN produtos p ON p.id=r.produto_id
        WHERE (r.cnpj=%s OR LOWER(r.fornecedor)=LOWER(%s)) AND r.codigo_fornecedor=%s
        ORDER BY r.id DESC LIMIT 1
        """,
        (cnpj, fornecedor, codigo_fornecedor),
        one=True,
    )
    if rel:
        return rel

    product = query(
        """
        SELECT p.id,p.codigo
        FROM produtos p
        LEFT JOIN fornecedores f ON f.id=p.fornecedor_id
        WHERE p.status <> 'inativo'
          AND (f.cnpj=%s OR LOWER(f.nome)=LOWER(%s))
          AND (p.codigo=%s OR LOWER(p.descricao)=LOWER(%s))
        ORDER BY (p.codigo=%s) DESC
        LIMIT 1
        """,
        (cnpj, fornecedor, codigo_fornecedor, descricao, codigo_fornecedor),
        one=True,
    )
    if product:
        return product
    return None


def get_or_create_nfe_supplier(fornecedor, cnpj, telefone=""):
    supplier = query("SELECT id,nome FROM fornecedores WHERE cnpj=%s", (cnpj,), one=True) if cnpj else None
    if supplier:
        return {"id": supplier["id"], "nome": supplier["nome"], "created": False}
    supplier = query("SELECT id,nome FROM fornecedores WHERE LOWER(nome)=LOWER(%s)", (fornecedor,), one=True)
    if supplier:
        return {"id": supplier["id"], "nome": supplier["nome"], "created": False}
    supplier_id = execute(
        "INSERT INTO fornecedores (nome,cnpj,contato,telefone,email,status) VALUES (%s,%s,%s,%s,%s,%s)",
        (
            fornecedor or "NÃ£o informado",
            cnpj or None,
            "NÃ£o informado",
            telefone or "NÃ£o informado",
            "NÃ£o informado",
            "ativo",
        ),
    )
    return {"id": supplier_id, "nome": fornecedor or "NÃ£o informado", "created": True}


def create_pending_product_from_nfe_item(item, fornecedor_id):
    code = "NÃ£o registrado"
    descricao = item["descricao"] or "Produto importado da NF-e"
    produto_id = execute(
        """
        INSERT INTO produtos
        (codigo,descricao,categoria,unidade,quantidade_por_unidade_compra,unidade_base,fornecedor_id,estoque_seguranca,status)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """,
        (
            code,
            descricao,
            "Importado via NF-e",
            item["unidade"] or "UN",
            1,
            item["unidade"] or "UN",
            fornecedor_id,
            0,
            "pendente revisao",
        ),
    )
    return {"id": produto_id, "codigo": code, "descricao": descricao, "created": True}


def register_product_supplier_link(fornecedor, cnpj, item, produto_id):
    if not item.get("codigo_fornecedor"):
        return
    if using_postgres():
        execute(
            """
            INSERT INTO produto_fornecedor_relacionamentos
            (fornecedor,cnpj,codigo_fornecedor,descricao_fornecedor,produto_id)
            VALUES (%s,%s,%s,%s,%s)
            ON CONFLICT (cnpj,codigo_fornecedor) DO UPDATE SET
              descricao_fornecedor=EXCLUDED.descricao_fornecedor,
              produto_id=EXCLUDED.produto_id,
              atualizado_em=CURRENT_TIMESTAMP
            """,
            (
                fornecedor,
                cnpj,
                item["codigo_fornecedor"],
                item["descricao"],
                produto_id,
            ),
        )
        return
    execute(
        """
        INSERT INTO produto_fornecedor_relacionamentos
        (fornecedor,cnpj,codigo_fornecedor,descricao_fornecedor,produto_id)
        VALUES (%s,%s,%s,%s,%s)
        ON DUPLICATE KEY UPDATE
          descricao_fornecedor=VALUES(descricao_fornecedor),
          produto_id=VALUES(produto_id),
          atualizado_em=CURRENT_TIMESTAMP
        """,
        (
            fornecedor,
            cnpj,
            item["codigo_fornecedor"],
            item["descricao"],
            produto_id,
        ),
    )

def column_exists(table_name, column_name):
    if using_postgres():
        return query(
            """
            SELECT column_name
            FROM information_schema.columns
            WHERE table_schema='public' AND table_name=%s AND column_name=%s
            """,
            (table_name, column_name),
            one=True,
        )
    return query(
        """
        SELECT COLUMN_NAME
        FROM information_schema.COLUMNS
        WHERE TABLE_SCHEMA=%s AND TABLE_NAME=%s AND COLUMN_NAME=%s
        """,
        (app.config["MYSQL_DATABASE"], table_name, column_name),
        one=True,
    )


def table_exists(table_name):
    if using_postgres():
        return query(
            """
            SELECT table_name
            FROM information_schema.tables
            WHERE table_schema='public' AND table_name=%s
            """,
            (table_name,),
            one=True,
        )
    return query(
        """
        SELECT TABLE_NAME
        FROM information_schema.TABLES
        WHERE TABLE_SCHEMA=%s AND TABLE_NAME=%s
        """,
        (app.config["MYSQL_DATABASE"], table_name),
        one=True,
    )

def ensure_auxiliary_registries():
    if using_postgres():
        return
    execute(
        """
        CREATE TABLE IF NOT EXISTS cadastro_categorias (
          id INT AUTO_INCREMENT PRIMARY KEY,
          nome VARCHAR(120) NOT NULL UNIQUE,
          status ENUM('ativo','inativo') NOT NULL DEFAULT 'ativo',
          criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
          atualizado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
        )
        """
    )
    execute(
        """
        CREATE TABLE IF NOT EXISTS cadastro_tipos_pagamento (
          id INT AUTO_INCREMENT PRIMARY KEY,
          nome VARCHAR(120) NOT NULL UNIQUE,
          status ENUM('ativo','inativo') NOT NULL DEFAULT 'ativo',
          criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
          atualizado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
        )
        """
    )
    for row in query(
        """
        SELECT categoria nome FROM produtos WHERE categoria IS NOT NULL AND categoria<>''
        UNION
        SELECT categoria nome FROM orcamentos WHERE categoria IS NOT NULL AND categoria<>''
        """
    ):
        execute(
            "INSERT IGNORE INTO cadastro_categorias (nome,status) VALUES (%s,'ativo')",
            (row["nome"],),
        )
    execute(
        "INSERT IGNORE INTO cadastro_categorias (nome,status) VALUES (%s,'ativo')",
        ("Importado via NF-e",),
    )
    for nome in ("PIX", "Boleto", "Cartao de Credito", "Cartao de Debito", "Dinheiro", "Transferencia"):
        execute(
            "INSERT IGNORE INTO cadastro_tipos_pagamento (nome,status) VALUES (%s,'ativo')",
            (nome,),
        )
    for row in query("SELECT DISTINCT metodo_pagamento nome FROM ordens_compra WHERE metodo_pagamento IS NOT NULL AND metodo_pagamento<>''"):
        execute(
            "INSERT IGNORE INTO cadastro_tipos_pagamento (nome,status) VALUES (%s,'ativo')",
            (row["nome"],),
        )

def ensure_product_schema():
    if using_postgres():
        return
    execute(
        """
        CREATE TABLE IF NOT EXISTS unidades_medida (
          id INT AUTO_INCREMENT PRIMARY KEY,
          codigo VARCHAR(20) NOT NULL UNIQUE,
          nome VARCHAR(120) NOT NULL,
          status ENUM('ativo','inativo') NOT NULL DEFAULT 'ativo',
          criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        """
    )
    for codigo, nome in DEFAULT_UNITS:
        execute(
            """
            INSERT INTO unidades_medida (codigo,nome,status)
            VALUES (%s,%s,'ativo')
            ON DUPLICATE KEY UPDATE nome=VALUES(nome)
            """,
            (codigo, nome),
        )
    if not column_exists("produtos", "fornecedor_id"):
        execute("ALTER TABLE produtos ADD COLUMN fornecedor_id INT NULL AFTER unidade")
    if not column_exists("produtos", "estoque_seguranca"):
        execute("ALTER TABLE produtos ADD COLUMN estoque_seguranca DECIMAL(12,2) NOT NULL DEFAULT 0 AFTER fornecedor_id")
    if not column_exists("produtos", "unidade_estoque_seguranca"):
        execute("ALTER TABLE produtos ADD COLUMN unidade_estoque_seguranca VARCHAR(30) NULL AFTER estoque_seguranca")
    if not column_exists("produtos", "quantidade_por_unidade_compra"):
        execute("ALTER TABLE produtos ADD COLUMN quantidade_por_unidade_compra DECIMAL(14,4) NOT NULL DEFAULT 1 AFTER unidade")
    if not column_exists("produtos", "unidade_base"):
        execute("ALTER TABLE produtos ADD COLUMN unidade_base VARCHAR(30) NOT NULL DEFAULT 'UN' AFTER quantidade_por_unidade_compra")
    if not column_exists("produtos", "estoque_atual"):
        execute("ALTER TABLE produtos ADD COLUMN estoque_atual DECIMAL(14,4) NOT NULL DEFAULT 0 AFTER estoque_seguranca")
    if not column_exists("produtos", "custo_atual"):
        execute("ALTER TABLE produtos ADD COLUMN custo_atual DECIMAL(14,4) NOT NULL DEFAULT 0 AFTER estoque_atual")
    if not column_exists("historico_custos", "unidade_compra"):
        execute("ALTER TABLE historico_custos ADD COLUMN unidade_compra VARCHAR(30) NULL AFTER quantidade")
    if not column_exists("historico_custos", "quantidade_por_unidade_compra"):
        execute("ALTER TABLE historico_custos ADD COLUMN quantidade_por_unidade_compra DECIMAL(14,4) NOT NULL DEFAULT 1 AFTER unidade_compra")
    if not column_exists("historico_custos", "unidade_base"):
        execute("ALTER TABLE historico_custos ADD COLUMN unidade_base VARCHAR(30) NULL AFTER quantidade_por_unidade_compra")
    if not column_exists("historico_custos", "quantidade_total_base"):
        execute("ALTER TABLE historico_custos ADD COLUMN quantidade_total_base DECIMAL(14,4) NOT NULL DEFAULT 0 AFTER unidade_base")
    if not column_exists("historico_custos", "valor_unitario_base"):
        execute("ALTER TABLE historico_custos ADD COLUMN valor_unitario_base DECIMAL(14,4) NOT NULL DEFAULT 0 AFTER valor_unitario")
    execute("ALTER TABLE produtos MODIFY codigo VARCHAR(30) NOT NULL")
    unique_code_indexes = query(
        """
        SELECT INDEX_NAME
        FROM information_schema.STATISTICS
        WHERE TABLE_SCHEMA=%s
          AND TABLE_NAME='produtos'
          AND COLUMN_NAME='codigo'
          AND NON_UNIQUE=0
          AND INDEX_NAME <> 'PRIMARY'
        """,
        (app.config["MYSQL_DATABASE"],),
    )
    for index in unique_code_indexes:
        index_name = index["INDEX_NAME"].replace("`", "``")
        execute(f"ALTER TABLE produtos DROP INDEX `{index_name}`")
    execute(
        """
        ALTER TABLE produtos
        MODIFY status ENUM('ativo','inativo','pendente revisao') NOT NULL DEFAULT 'ativo'
        """
    )
    execute("UPDATE produtos SET unidade='UN' WHERE LOWER(unidade) IN ('un','unidade')")
    execute("UPDATE produtos SET unidade='L' WHERE unidade='GL'")
    execute("UPDATE produtos SET quantidade_por_unidade_compra=1 WHERE quantidade_por_unidade_compra IS NULL OR quantidade_por_unidade_compra<=0")
    execute("UPDATE produtos SET unidade_base=unidade WHERE unidade_base IS NULL OR unidade_base=''")
    for product in query(
        """
        SELECT id,codigo,descricao,unidade,quantidade_por_unidade_compra,unidade_base
        FROM produtos
        WHERE (
            UPPER(unidade) IN ('PALLET','CX','CAIXA','ROLO')
            AND (
              COALESCE(quantidade_por_unidade_compra,1)<=1
              OR UPPER(COALESCE(unidade_base,'')) IN ('','UN')
            )
          )
          OR (
            UPPER(codigo) LIKE 'MDF%%'
            AND UPPER(COALESCE(unidade_base,'')) IN ('','UN')
          )
        """
    ):
        inferred = infer_purchase_conversion(product)
        if inferred:
            execute(
                "UPDATE produtos SET quantidade_por_unidade_compra=%s,unidade_base=%s WHERE id=%s",
                (inferred[0], inferred[1], product["id"]),
            )
        elif (product.get("codigo") or "").upper().replace(" ", "").startswith("MDF"):
            execute("UPDATE produtos SET unidade_base=%s WHERE id=%s", ("CHAPAS", product["id"]))
    if table_exists("historico_custos"):
        execute(
            """
            UPDATE historico_custos h
            JOIN produtos p ON p.id=h.produto_id
            LEFT JOIN nfe_importacao_itens i ON i.id=h.recebimento_item_id
            SET h.unidade_compra=CASE UPPER(COALESCE(NULLIF(i.unidade,''),p.unidade))
                    WHEN 'PT' THEN 'PALLET'
                    WHEN 'PL' THEN 'PALLET'
                    WHEN 'PAL' THEN 'PALLET'
                    WHEN 'CH' THEN 'CHAPAS'
                    WHEN 'CHAPA' THEN 'CHAPAS'
                    WHEN 'PC' THEN 'PECAS'
                    WHEN 'PECA' THEN 'PECAS'
                    WHEN 'PEÃ‡A' THEN 'PECAS'
                    WHEN 'UNID' THEN 'UN'
                    WHEN 'UNIDADE' THEN 'UN'
                    ELSE UPPER(COALESCE(NULLIF(i.unidade,''),p.unidade,'UN'))
                END,
                h.quantidade_por_unidade_compra=CASE
                    WHEN (
                        CASE UPPER(COALESCE(NULLIF(i.unidade,''),p.unidade))
                            WHEN 'PT' THEN 'PALLET'
                            WHEN 'PL' THEN 'PALLET'
                            WHEN 'PAL' THEN 'PALLET'
                            WHEN 'CH' THEN 'CHAPAS'
                            WHEN 'CHAPA' THEN 'CHAPAS'
                            WHEN 'PC' THEN 'PECAS'
                            WHEN 'PECA' THEN 'PECAS'
                            WHEN 'PEÃ‡A' THEN 'PECAS'
                            WHEN 'UNID' THEN 'UN'
                            WHEN 'UNIDADE' THEN 'UN'
                            ELSE UPPER(COALESCE(NULLIF(i.unidade,''),p.unidade,'UN'))
                        END
                    ) = (
                        CASE UPPER(COALESCE(p.unidade,'UN'))
                            WHEN 'PT' THEN 'PALLET'
                            WHEN 'PL' THEN 'PALLET'
                            WHEN 'PAL' THEN 'PALLET'
                            WHEN 'CH' THEN 'CHAPAS'
                            WHEN 'CHAPA' THEN 'CHAPAS'
                            WHEN 'PC' THEN 'PECAS'
                            WHEN 'PECA' THEN 'PECAS'
                            WHEN 'PEÃ‡A' THEN 'PECAS'
                            WHEN 'UNID' THEN 'UN'
                            WHEN 'UNIDADE' THEN 'UN'
                            ELSE UPPER(COALESCE(p.unidade,'UN'))
                        END
                    )
                    OR (
                        CASE UPPER(COALESCE(NULLIF(i.unidade,''),p.unidade))
                            WHEN 'PT' THEN 'PALLET'
                            WHEN 'PL' THEN 'PALLET'
                            WHEN 'PAL' THEN 'PALLET'
                            ELSE UPPER(COALESCE(NULLIF(i.unidade,''),p.unidade,'UN'))
                        END
                    ) IN ('PALLET','CX','CAIXA','ROLO')
                    THEN COALESCE(NULLIF(p.quantidade_por_unidade_compra,0),1)
                    ELSE 1
                END,
                h.unidade_base=COALESCE(NULLIF(p.unidade_base,''),p.unidade),
                h.quantidade_total_base=h.quantidade * h.quantidade_por_unidade_compra,
                h.valor_unitario_base=h.valor_total / NULLIF(h.quantidade * h.quantidade_por_unidade_compra,0)
            WHERE h.produto_id=p.id
            """
        )
        execute(
            """
            UPDATE historico_custos
            SET quantidade_total_base=quantidade * COALESCE(NULLIF(quantidade_por_unidade_compra,0),1),
                valor_unitario_base=valor_total / NULLIF(quantidade * COALESCE(NULLIF(quantidade_por_unidade_compra,0),1),0)
            """
        )
    execute("UPDATE produtos SET estoque_seguranca=COALESCE(NULLIF(estoque_seguranca,0),estoque_semanal,0)")
    execute(
        """
        UPDATE produtos p
        SET fornecedor_id = (
            SELECT o.fornecedor_id
            FROM ordens_compra o
            WHERE o.produto_id=p.id
            ORDER BY o.data_preenchimento DESC, o.id DESC
            LIMIT 1
        )
        WHERE p.fornecedor_id IS NULL
        """
    )
    execute(
        """
        UPDATE produtos p
        SET fornecedor_id = (SELECT f.id FROM fornecedores f WHERE f.status='ativo' ORDER BY f.nome LIMIT 1)
        WHERE p.fornecedor_id IS NULL
        """
    )
    constraint = query(
        """
        SELECT CONSTRAINT_NAME
        FROM information_schema.TABLE_CONSTRAINTS
        WHERE TABLE_SCHEMA=%s AND TABLE_NAME='produtos' AND CONSTRAINT_NAME='fk_produto_fornecedor'
        """,
        (app.config["MYSQL_DATABASE"],),
        one=True,
    )
    if not constraint:
        execute("ALTER TABLE produtos ADD CONSTRAINT fk_produto_fornecedor FOREIGN KEY (fornecedor_id) REFERENCES fornecedores(id)")


def ensure_order_schema():
    if using_postgres():
        return
    if not column_exists("ordens_compra", "parcelas"):
        execute("ALTER TABLE ordens_compra ADD COLUMN parcelas TINYINT NOT NULL DEFAULT 1 AFTER metodo_pagamento")
    if not column_exists("ordens_compra", "prazos_parcelas"):
        execute("ALTER TABLE ordens_compra ADD COLUMN prazos_parcelas VARCHAR(255) NULL AFTER parcelas")
    if not column_exists("ordens_compra", "ordem_original_id"):
        execute("ALTER TABLE ordens_compra ADD COLUMN ordem_original_id INT NULL AFTER status_compra")
    execute(
        """
        CREATE TABLE IF NOT EXISTS ordem_compra_itens (
          id INT AUTO_INCREMENT PRIMARY KEY,
          ordem_id INT NOT NULL,
          produto_id INT NOT NULL,
          quantidade DECIMAL(14,4) NOT NULL,
          preco_negociado DECIMAL(14,4) NOT NULL DEFAULT 0,
          frete DECIMAL(14,2) NOT NULL DEFAULT 0,
          valor_total_item DECIMAL(14,2) NOT NULL DEFAULT 0,
          categoria VARCHAR(120),
          unidade_compra VARCHAR(30),
          quantidade_por_unidade DECIMAL(14,4) NOT NULL DEFAULT 1,
          unidade_base VARCHAR(30),
          criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
          CONSTRAINT fk_ordem_item_ordem FOREIGN KEY (ordem_id) REFERENCES ordens_compra(id) ON DELETE CASCADE,
          CONSTRAINT fk_ordem_item_produto FOREIGN KEY (produto_id) REFERENCES produtos(id)
        )
        """
    )
    execute(
        """
        INSERT INTO ordem_compra_itens
        (ordem_id,produto_id,quantidade,preco_negociado,frete,valor_total_item,categoria,unidade_compra,quantidade_por_unidade,unidade_base)
        SELECT o.id,o.produto_id,o.quantidade,o.preco_negociado,o.frete,o.valor_total,p.categoria,p.unidade,
               COALESCE(NULLIF(p.quantidade_por_unidade_compra,0),1),COALESCE(p.unidade_base,p.unidade)
        FROM ordens_compra o
        JOIN produtos p ON p.id=o.produto_id
        LEFT JOIN ordem_compra_itens i ON i.ordem_id=o.id
        WHERE i.id IS NULL
        """
    )


@app.template_filter("moeda")
def moeda(value):
    number = decimal_value(value)
    return f"R$ {number:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


@app.context_processor
def inject_globals():
    return {
        "hoje": date.today(),
        "ano": date.today().year,
        "current_user": getattr(g, "current_user", None),
        "permission_modules": PERMISSION_MODULES,
        "can_access": can_access,
    }


@app.route("/login", methods=["GET", "POST"])
def login():
    ensure_auth_schema()
    if request.method == "POST":
        login_name = (request.form.get("login") or "").strip()
        password = request.form.get("senha") or ""
        user = query(
            "SELECT id,nome,login,senha_hash,tipo_acesso,status FROM usuarios WHERE login=%s",
            (login_name,),
            one=True,
        )
        if user and user["status"] == "ativo" and check_password_hash(user["senha_hash"], password):
            execute("UPDATE usuarios SET ultimo_login=NOW() WHERE id=%s", (user["id"],))
            session.clear()
            session["user_id"] = user["id"]
            session["user_name"] = user["nome"]
            next_url = request.form.get("next") or url_for("dashboard")
            if not next_url.startswith("/"):
                next_url = url_for("dashboard")
            return redirect(next_url)
        flash("Usuario ou senha invalidos.", "danger")
    return render_template("login.html", next_url=request.args.get("next") or request.form.get("next") or "")


@app.get("/logout")
def logout():
    session.clear()
    flash("Sessao encerrada.", "success")
    return redirect(url_for("login"))


@app.route("/usuarios", methods=["GET", "POST"])
def usuarios():
    if getattr(g, "current_user", {}).get("tipo_acesso") != "administrador":
        flash("Apenas administradores podem gerenciar usuarios.", "warning")
        return redirect(first_allowed_url())
    if request.method == "POST":
        form = request.form
        action = form.get("action") or "save"
        user_id = form.get("id")
        if action == "reset_password":
            senha = form.get("senha") or ""
            if not user_id or not senha or len(senha) < 6:
                flash("Informe uma nova senha com no minimo 6 caracteres.", "danger")
                return redirect(url_for("usuarios"))
            execute("UPDATE usuarios SET senha_hash=%s WHERE id=%s", (generate_password_hash(senha), user_id))
            flash("Senha atualizada com sucesso.", "success")
            return redirect(url_for("usuarios"))
        if action == "toggle_admin":
            if not user_id:
                return redirect(url_for("usuarios"))
            user = query("SELECT id,tipo_acesso FROM usuarios WHERE id=%s", (user_id,), one=True)
            if not user:
                flash("Usuario nao encontrado.", "warning")
                return redirect(url_for("usuarios"))
            if user["id"] == g.current_user["id"]:
                flash("Voce nao pode alterar o proprio perfil de administrador.", "danger")
                return redirect(url_for("usuarios"))
            new_type = "comum" if user["tipo_acesso"] == "administrador" else "administrador"
            execute("UPDATE usuarios SET tipo_acesso=%s WHERE id=%s", (new_type, user_id))
            execute("DELETE FROM usuario_permissoes WHERE usuario_id=%s", (user_id,))
            modules_to_save = {module for module, _label in PERMISSION_MODULES}
            if new_type == "comum":
                modules_to_save.discard("usuarios")
            for module in modules_to_save:
                execute("INSERT INTO usuario_permissoes (usuario_id,modulo,pode_visualizar) VALUES (%s,%s,1)", (user_id, module))
            flash("Perfil de acesso atualizado.", "success")
            return redirect(url_for("usuarios"))
        if action == "delete":
            if not user_id:
                return redirect(url_for("usuarios"))
            if int(user_id) == g.current_user["id"]:
                flash("Voce nao pode excluir o proprio usuario logado.", "danger")
                return redirect(url_for("usuarios"))
            execute("DELETE FROM usuarios WHERE id=%s", (user_id,))
            flash("Usuario removido com sucesso.", "success")
            return redirect(url_for("usuarios"))

        nome = (form.get("nome") or "").strip()
        login_name = (form.get("login") or "").strip()
        senha = form.get("senha") or ""
        if not nome and login_name:
            nome = login_name.split("@")[0]
        tipo_acesso = form.get("tipo_acesso") if form.get("tipo_acesso") in ("administrador", "comum") else "comum"
        status = form.get("status") if form.get("status") in ("ativo", "inativo") else "ativo"
        selected_modules = set(form.getlist("modulos"))
        if not nome or not login_name:
            flash("Nome e usuario/login sao obrigatorios.", "danger")
            return redirect(url_for("usuarios"))
        if tipo_acesso == "administrador":
            selected_modules = {module for module, _label in PERMISSION_MODULES}
        if user_id and str(user_id).isdigit():
            existing = query("SELECT id FROM usuarios WHERE login=%s AND id<>%s", (login_name, user_id), one=True)
            if existing:
                flash("Ja existe outro usuario com este login.", "danger")
                return redirect(url_for("usuarios"))
            if int(user_id) == g.current_user["id"] and status != "ativo":
                flash("Voce nao pode inativar o proprio usuario logado.", "danger")
                return redirect(url_for("usuarios"))
            if senha:
                execute(
                    "UPDATE usuarios SET nome=%s,login=%s,senha_hash=%s,tipo_acesso=%s,status=%s WHERE id=%s",
                    (nome, login_name, generate_password_hash(senha), tipo_acesso, status, user_id),
                )
            else:
                execute(
                    "UPDATE usuarios SET nome=%s,login=%s,tipo_acesso=%s,status=%s WHERE id=%s",
                    (nome, login_name, tipo_acesso, status, user_id),
                )
            execute("DELETE FROM usuario_permissoes WHERE usuario_id=%s", (user_id,))
            target_user_id = int(user_id)
            flash("Usuario atualizado com sucesso.", "success")
        else:
            if not senha or len(senha) < 6:
                flash("Senha obrigatoria para novo usuario.", "danger")
                return redirect(url_for("usuarios"))
            existing = query("SELECT id FROM usuarios WHERE login=%s", (login_name,), one=True)
            if existing:
                flash("Ja existe um usuario com este login.", "danger")
                return redirect(url_for("usuarios"))
            target_user_id = execute(
                """
                INSERT INTO usuarios (nome,login,senha_hash,tipo_acesso,status)
                VALUES (%s,%s,%s,%s,%s)
                """,
                (nome, login_name, generate_password_hash(senha), tipo_acesso, status),
            )
            flash("Usuario cadastrado com sucesso.", "success")
        for module in selected_modules:
            if module in {item[0] for item in PERMISSION_MODULES}:
                execute(
                    "INSERT INTO usuario_permissoes (usuario_id,modulo,pode_visualizar) VALUES (%s,%s,1)",
                    (target_user_id, module),
                )
        return redirect(url_for("usuarios"))

    users = query("SELECT id,nome,login,tipo_acesso,status,ultimo_login,criado_em,atualizado_em FROM usuarios ORDER BY nome")
    permissions = query("SELECT usuario_id,modulo FROM usuario_permissoes WHERE pode_visualizar=1")
    permissions_by_user = {}
    for row in permissions:
        permissions_by_user.setdefault(row["usuario_id"], set()).add(row["modulo"])
    return render_template(
        "usuarios.html",
        usuarios=users,
        permissions_by_user=permissions_by_user,
        modules=PERMISSION_MODULES,
    )


@app.route("/")
def dashboard():
    if using_postgres():
        indicadores = query(
            """
            SELECT
              COALESCE(SUM(CASE WHEN EXTRACT(YEAR FROM data_preenchimento)=EXTRACT(YEAR FROM CURRENT_DATE)
                                  AND EXTRACT(MONTH FROM data_preenchimento)=EXTRACT(MONTH FROM CURRENT_DATE)
                                THEN valor_total ELSE 0 END), 0) total_mes,
              COALESCE(SUM(CASE WHEN status_compra='aguardando autorizacao' THEN 1 ELSE 0 END), 0) aguardando,
              COALESCE(SUM(CASE WHEN data_entrega < CURRENT_DATE AND status_compra NOT IN ('recebida','cancelada') THEN 1 ELSE 0 END), 0) atrasadas
            FROM ordens_compra
            """,
            one=True,
        )
        ultimas_sql = """
            SELECT o.*, f.nome fornecedor,
                   COALESCE(STRING_AGG(p.descricao, ' | ' ORDER BY i.id), '-') produto
            FROM ordens_compra o
            JOIN fornecedores f ON f.id=o.fornecedor_id
            LEFT JOIN ordem_compra_itens i ON i.ordem_id=o.id
            LEFT JOIN produtos p ON p.id=i.produto_id
            GROUP BY o.id, f.nome
            ORDER BY o.id DESC LIMIT 8
        """
    else:
        indicadores = query(
            """
            SELECT
              COALESCE(SUM(CASE WHEN YEAR(data_preenchimento)=YEAR(CURDATE())
                                AND MONTH(data_preenchimento)=MONTH(CURDATE()) THEN valor_total END), 0) total_mes,
              SUM(status_compra='aguardando autorizacao') aguardando,
              SUM(data_entrega < CURDATE() AND status_compra NOT IN ('recebida','cancelada')) atrasadas
            FROM ordens_compra
            """,
            one=True,
        )
        ultimas_sql = """
            SELECT o.*, f.nome fornecedor,
                   COALESCE(GROUP_CONCAT(p.descricao ORDER BY i.id SEPARATOR ' | '), '-') produto
            FROM ordens_compra o
            JOIN fornecedores f ON f.id=o.fornecedor_id
            LEFT JOIN ordem_compra_itens i ON i.ordem_id=o.id
            LEFT JOIN produtos p ON p.id=i.produto_id
            GROUP BY o.id
            ORDER BY o.id DESC LIMIT 8
        """
    categorias = query(
        """SELECT p.categoria, COALESCE(SUM(i.valor_total_item),0) total
           FROM produtos p
           LEFT JOIN ordem_compra_itens i ON i.produto_id=p.id
           GROUP BY p.categoria ORDER BY total DESC LIMIT 6"""
    )
    ultimas = query(ultimas_sql)
    return render_template("dashboard.html", indicadores=indicadores, categorias=categorias, ultimas=ultimas)


def registry_usage(module, name):
    if module == "categorias":
        usages = []
        product_count = query("SELECT COUNT(*) total FROM produtos WHERE categoria=%s", (name,), one=True)["total"]
        ensure_monthly_budget_schema()
        budget_count = query(
            """
            SELECT COUNT(*) total
            FROM orcamentos o
            JOIN cadastro_categorias c ON c.id=o.categoria_id
            WHERE c.nome=%s
            """,
            (name,),
            one=True,
        )["total"]
        if product_count:
            usages.append(f"Produtos ({product_count})")
        if budget_count:
            usages.append(f"Orcamentos ({budget_count})")
        return usages
    if module == "pagamentos":
        count = query("SELECT COUNT(*) total FROM ordens_compra WHERE metodo_pagamento=%s", (name,), one=True)["total"]
        return [f"Ordens de compra ({count})"] if count else []
    return []


def ensure_monthly_budget_schema():
    if using_postgres():
        return
    execute(
        """
        CREATE TABLE IF NOT EXISTS orcamentos (
          id INT AUTO_INCREMENT PRIMARY KEY,
          mes TINYINT NOT NULL,
          ano SMALLINT NOT NULL,
          categoria_id INT NULL,
          categoria VARCHAR(80) NULL,
          valor_orcamento DECIMAL(14,2) NULL,
          orcamento_previsto DECIMAL(14,2) NULL,
          data_importacao TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
          usuario_importacao VARCHAR(120) NULL
        )
        """
    )
    if not column_exists("orcamentos", "categoria_id"):
        execute("ALTER TABLE orcamentos ADD COLUMN categoria_id INT NULL AFTER ano")
    if not column_exists("orcamentos", "valor_orcamento"):
        execute("ALTER TABLE orcamentos ADD COLUMN valor_orcamento DECIMAL(14,2) NULL AFTER categoria_id")
    if not column_exists("orcamentos", "data_importacao"):
        execute("ALTER TABLE orcamentos ADD COLUMN data_importacao TIMESTAMP DEFAULT CURRENT_TIMESTAMP AFTER valor_orcamento")
    if not column_exists("orcamentos", "usuario_importacao"):
        execute("ALTER TABLE orcamentos ADD COLUMN usuario_importacao VARCHAR(120) NULL AFTER data_importacao")
    if column_exists("orcamentos", "categoria"):
        execute("ALTER TABLE orcamentos MODIFY categoria VARCHAR(80) NULL")
    execute(
        """
        UPDATE orcamentos o
        JOIN cadastro_categorias c ON LOWER(c.nome)=LOWER(o.categoria)
        SET o.categoria_id=c.id
        WHERE o.categoria_id IS NULL AND o.categoria IS NOT NULL
        """
    )
    execute("UPDATE orcamentos SET valor_orcamento=orcamento_previsto WHERE valor_orcamento IS NULL AND orcamento_previsto IS NOT NULL")
    execute("UPDATE orcamentos SET orcamento_previsto=valor_orcamento WHERE orcamento_previsto IS NULL AND valor_orcamento IS NOT NULL")
    unique_indexes = query(
        """
        SELECT INDEX_NAME
        FROM information_schema.STATISTICS
        WHERE TABLE_SCHEMA=%s AND TABLE_NAME='orcamentos' AND NON_UNIQUE=0 AND INDEX_NAME<>'PRIMARY'
        GROUP BY INDEX_NAME
        """,
        (app.config["MYSQL_DATABASE"],),
    )
    for index in unique_indexes:
        index_name = index["INDEX_NAME"].replace("`", "``")
        execute(f"ALTER TABLE orcamentos DROP INDEX `{index_name}`")
    index_exists = query(
        """
        SELECT INDEX_NAME
        FROM information_schema.STATISTICS
        WHERE TABLE_SCHEMA=%s AND TABLE_NAME='orcamentos' AND INDEX_NAME='uk_orcamento_categoria_competencia'
        """,
        (app.config["MYSQL_DATABASE"],),
        one=True,
    )
    if not index_exists:
        execute("ALTER TABLE orcamentos ADD UNIQUE KEY uk_orcamento_categoria_competencia (categoria_id,ano,mes)")


def budget_spent_sql():
    if using_postgres():
        return """
            SELECT
              c.id categoria_id,
              EXTRACT(YEAR FROM n.data_entrada) ano,
              EXTRACT(MONTH FROM n.data_entrada) mes,
              COALESCE(SUM(i.valor_total),0) total_gasto
            FROM nfe_importacao_itens i
            JOIN nfe_importacoes n ON n.id=i.nfe_importacao_id
            JOIN produtos p ON p.id=i.produto_id
            JOIN cadastro_categorias c ON LOWER(c.nome)=LOWER(p.categoria)
            GROUP BY c.id,EXTRACT(YEAR FROM n.data_entrada),EXTRACT(MONTH FROM n.data_entrada)
        """
    return """
        SELECT
          c.id categoria_id,
          YEAR(n.data_entrada) ano,
          MONTH(n.data_entrada) mes,
          COALESCE(SUM(i.valor_total),0) total_gasto
        FROM nfe_importacao_itens i
        JOIN nfe_importacoes n ON n.id=i.nfe_importacao_id
        JOIN produtos p ON p.id=i.produto_id
        JOIN cadastro_categorias c ON LOWER(c.nome)=LOWER(p.categoria)
        GROUP BY c.id,YEAR(n.data_entrada),MONTH(n.data_entrada)
    """


def budget_status(valor_orcamento, valor_gasto):
    budget = decimal_value(valor_orcamento)
    spent = decimal_value(valor_gasto)
    if budget == 0 and spent == 0:
        return "SEM ORCAMENTO"
    if (budget == 0 and spent > 0) or spent > budget:
        return "ORCAMENTO ESTOURADO"
    return "DENTRO DO ORCAMENTO"


def build_budget_context():
    ensure_auxiliary_registries()
    ensure_product_schema()
    ensure_nfe_schema()
    ensure_order_schema()
    ensure_monthly_budget_schema()

    mes = request.args.get("mes") or str(date.today().month)
    ano_filter = request.args.get("ano") or str(date.today().year)
    categoria_id = request.args.get("categoria_id") or ""
    month = int(mes)
    year = int(ano_filter)
    params = [year, month, year, month]
    filters = ["(b.id IS NOT NULL OR COALESCE(g.total_gasto,0) > 0)"]
    if categoria_id:
        filters.append("c.id=%s")
        params.append(int(categoria_id))
    where = " AND ".join(filters)
    rows = query(
        f"""
        SELECT
          b.id,
          %s mes,
          %s ano,
          c.id categoria_id,
          COALESCE(b.valor_orcamento,0) valor_orcamento,
          b.data_importacao,
          b.usuario_importacao,
          c.nome categoria,
          COALESCE(g.total_gasto,0) valor_gasto,
          (COALESCE(b.valor_orcamento,0)-COALESCE(g.total_gasto,0)) saldo,
          CASE
            WHEN COALESCE(b.valor_orcamento,0) > 0 THEN (COALESCE(g.total_gasto,0)/b.valor_orcamento)*100
            WHEN COALESCE(g.total_gasto,0) > 0 THEN 100
            ELSE 0
          END percentual
        FROM cadastro_categorias c
        LEFT JOIN orcamentos b ON b.categoria_id=c.id AND b.ano=%s AND b.mes=%s
        LEFT JOIN ({budget_spent_sql()}) g
          ON g.categoria_id=c.id AND g.ano=%s AND g.mes=%s
        WHERE {where}
        ORDER BY c.nome
        """,
        (month, year, *tuple(params)),
    )
    for row in rows:
        row["status"] = budget_status(row["valor_orcamento"], row["valor_gasto"])
    total_budget = sum(decimal_value(row["valor_orcamento"]) for row in rows)
    total_spent = sum(decimal_value(row["valor_gasto"]) for row in rows)
    summary = {
        "orcamento": total_budget,
        "gasto": total_spent,
        "saldo": total_budget - total_spent,
        "percentual": (total_spent / total_budget * 100) if total_budget else (Decimal("100") if total_spent > 0 else Decimal("0")),
        "status": budget_status(total_budget, total_spent),
    }
    return {
        "budget_rows": rows,
        "budget_summary": summary,
        "budget_filters": {"mes": int(mes), "ano": int(ano_filter), "categoria_id": categoria_id},
        "budget_categories": query("SELECT id,nome FROM cadastro_categorias WHERE status='ativo' ORDER BY nome"),
        "budget_years": query("SELECT DISTINCT ano FROM orcamentos ORDER BY ano DESC"),
    }


def registry_counts():
    return query(
        """
        SELECT
          (SELECT COUNT(*) FROM cadastro_categorias) categorias,
          (SELECT COUNT(*) FROM cadastro_tipos_pagamento) pagamentos,
          (SELECT COUNT(*) FROM orcamentos) orcamentos
        """,
        one=True,
    )


@app.route("/cadastros", methods=["GET", "POST"])
def cadastros():
    ensure_auxiliary_registries()
    ensure_monthly_budget_schema()
    module = request.args.get("modulo", "categorias")
    if module not in ("categorias", "pagamentos", "orcamentos"):
        module = "categorias"
    if module == "orcamentos":
        context = build_budget_context()
        counts = registry_counts()
        return render_template("cadastros.html", module=module, counts=counts, **context)
    configs = {
        "categorias": {
            "table": "cadastro_categorias",
            "title": "Categorias de Produto",
            "field": "Nome da categoria",
            "modal": "categoriaForm",
        },
        "pagamentos": {
            "table": "cadastro_tipos_pagamento",
            "title": "Tipos de Pagamento",
            "field": "Nome do tipo de pagamento",
            "modal": "pagamentoForm",
        },
    }
    config = configs[module]

    if request.method == "POST":
        form = request.form
        form_module = form.get("modulo", module)
        config = configs.get(form_module, config)
        item_id = form.get("id")
        nome = (form.get("nome") or "").strip()
        status = form.get("status") or "ativo"
        if not nome:
            flash("Informe o nome do cadastro.", "danger")
            return redirect(url_for("cadastros", modulo=form_module))
        duplicate = query(
            f"SELECT id FROM {config['table']} WHERE LOWER(nome)=LOWER(%s) AND (%s='' OR id<>%s)",
            (nome, item_id or "", item_id or 0),
            one=True,
        )
        if duplicate:
            flash("Ja existe um cadastro com este nome.", "danger")
            return redirect(url_for("cadastros", modulo=form_module))
        if item_id:
            old = query(f"SELECT nome FROM {config['table']} WHERE id=%s", (item_id,), one=True)
            usages = registry_usage(form_module, old["nome"]) if old and old["nome"] != nome else []
            if usages:
                flash(f"Nao e possivel renomear. Cadastro em uso em: {', '.join(usages)}.", "danger")
                return redirect(url_for("cadastros", modulo=form_module))
            execute(f"UPDATE {config['table']} SET nome=%s,status=%s WHERE id=%s", (nome, status, item_id))
        else:
            execute(f"INSERT INTO {config['table']} (nome,status) VALUES (%s,%s)", (nome, status))
        flash("Cadastro salvo com sucesso.", "success")
        return redirect(url_for("cadastros", modulo=form_module))

    search = (request.args.get("q") or "").strip()
    sort = request.args.get("sort", "nome")
    direction = request.args.get("dir", "asc")
    page = max(int(request.args.get("page", 1) or 1), 1)
    per_page = 10
    allowed_sort = {"nome", "status", "criado_em", "atualizado_em"}
    if sort not in allowed_sort:
        sort = "nome"
    direction = "desc" if direction == "desc" else "asc"
    where = "WHERE nome LIKE %s" if search else ""
    params = (f"%{search}%",) if search else ()
    total = query(f"SELECT COUNT(*) total FROM {config['table']} {where}", params, one=True)["total"]
    offset = (page - 1) * per_page
    rows = query(
        f"SELECT * FROM {config['table']} {where} ORDER BY {sort} {direction} LIMIT %s OFFSET %s",
        params + (per_page, offset),
    )
    total_pages = max((total + per_page - 1) // per_page, 1)
    counts = registry_counts()
    return render_template(
        "cadastros.html",
        module=module,
        config=config,
        rows=rows,
        search=search,
        sort=sort,
        direction=direction,
        page=page,
        total_pages=total_pages,
        total=total,
        counts=counts,
    )


@app.post("/cadastros/<module>/<int:item_id>/excluir")
def excluir_cadastro(module, item_id):
    ensure_auxiliary_registries()
    configs = {
        "categorias": "cadastro_categorias",
        "pagamentos": "cadastro_tipos_pagamento",
    }
    table = configs.get(module)
    if not table:
        flash("Cadastro invalido.", "danger")
        return redirect(url_for("cadastros"))
    item = query(f"SELECT nome FROM {table} WHERE id=%s", (item_id,), one=True)
    if not item:
        flash("Cadastro nao encontrado.", "danger")
        return redirect(url_for("cadastros", modulo=module))
    usages = registry_usage(module, item["nome"])
    if usages:
        flash(f"Nao e possivel excluir. Cadastro em uso em: {', '.join(usages)}.", "danger")
        return redirect(url_for("cadastros", modulo=module))
    execute(f"DELETE FROM {table} WHERE id=%s", (item_id,))
    flash("Cadastro excluido.", "success")
    return redirect(url_for("cadastros", modulo=module))


@app.post("/cadastros/orcamentos/salvar")
def salvar_orcamento_mensal():
    ensure_monthly_budget_schema()
    form = request.form
    budget_id = form.get("id")
    mes = int(form.get("mes") or 0)
    ano_budget = int(form.get("ano") or 0)
    categoria_id = int(form.get("categoria_id") or 0)
    valor = money_value(form.get("valor_orcamento"))
    if mes < 1 or mes > 12 or ano_budget <= 0 or not categoria_id or valor < 0:
        flash("Informe mes, ano, categoria e valor validos.", "danger")
        return redirect(url_for("cadastros", modulo="orcamentos"))
    categoria = query("SELECT id FROM cadastro_categorias WHERE id=%s", (categoria_id,), one=True)
    if not categoria:
        flash("Categoria nao encontrada.", "danger")
        return redirect(url_for("cadastros", modulo="orcamentos"))
    duplicate = query(
        """
        SELECT id FROM orcamentos
        WHERE categoria_id=%s AND ano=%s AND mes=%s AND (%s='' OR id<>%s)
        """,
        (categoria_id, ano_budget, mes, budget_id or "", budget_id or 0),
        one=True,
    )
    if duplicate:
        execute(
            "UPDATE orcamentos SET valor_orcamento=%s,orcamento_previsto=%s,data_importacao=CURRENT_TIMESTAMP WHERE id=%s",
            (valor, valor, duplicate["id"]),
        )
        flash("Orcamento existente atualizado.", "success")
    elif budget_id:
        execute(
            """
            UPDATE orcamentos
            SET mes=%s,ano=%s,categoria_id=%s,valor_orcamento=%s,orcamento_previsto=%s,data_importacao=CURRENT_TIMESTAMP
            WHERE id=%s
            """,
            (mes, ano_budget, categoria_id, valor, valor, budget_id),
        )
        flash("Orcamento atualizado.", "success")
    else:
        execute(
            """
            INSERT INTO orcamentos (mes,ano,categoria_id,valor_orcamento,orcamento_previsto,usuario_importacao)
            VALUES (%s,%s,%s,%s,%s,%s)
            """,
            (mes, ano_budget, categoria_id, valor, valor, "Sistema"),
        )
        flash("Orcamento cadastrado.", "success")
    return redirect(url_for("cadastros", modulo="orcamentos", mes=mes, ano=ano_budget))


@app.post("/cadastros/orcamentos/<int:item_id>/excluir")
def excluir_orcamento_mensal(item_id):
    ensure_monthly_budget_schema()
    row = query("SELECT mes,ano FROM orcamentos WHERE id=%s", (item_id,), one=True)
    execute("DELETE FROM orcamentos WHERE id=%s", (item_id,))
    flash("Orcamento excluido.", "success")
    if row:
        return redirect(url_for("cadastros", modulo="orcamentos", mes=row["mes"], ano=row["ano"]))
    return redirect(url_for("cadastros", modulo="orcamentos"))


@app.get("/cadastros/orcamentos/modelo")
def baixar_modelo_orcamentos():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Orcamentos Mensais"
    sheet.append(["MÃªs", "Ano", "Categoria", "Valor de Compra"])
    sheet.append([date.today().month, date.today().year, "Categoria Exemplo", 50000])
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name="modelo_orcamentos_mensais.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.post("/cadastros/orcamentos/importar")
def importar_orcamentos_mensais():
    ensure_auxiliary_registries()
    ensure_monthly_budget_schema()
    file = request.files.get("arquivo")
    if not file or not file.filename:
        flash("Selecione um arquivo XLSX.", "danger")
        return redirect(url_for("cadastros", modulo="orcamentos"))
    if not file.filename.lower().endswith(".xlsx"):
        flash("O arquivo deve estar no formato .xlsx.", "danger")
        return redirect(url_for("cadastros", modulo="orcamentos"))

    categories = {
        row["nome"].strip().lower(): row["id"]
        for row in query("SELECT id,nome FROM cadastro_categorias WHERE status='ativo'")
    }
    total = imported = updated = errors = 0
    error_messages = []
    try:
        workbook = load_workbook(file, data_only=True)
        sheet = workbook.active
        header = [normalize_header(cell.value) for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        required = ["mes", "ano", "categoria", "valor de compra"]
        if header != required:
            flash("O XLSX deve conter exatamente as colunas: Mes, Ano, Categoria, Valor de Compra.", "danger")
            workbook.close()
            return redirect(url_for("cadastros", modulo="orcamentos"))
        for line_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if all(value is None or str(value).strip() == "" for value in row):
                continue
            total += 1
            row_errors = []
            mes_raw, ano_raw, categoria_raw, valor_raw = row[:4]
            try:
                mes = int(mes_raw)
            except Exception:
                mes = 0
                row_errors.append("mes invalido")
            try:
                ano_budget = int(ano_raw)
            except Exception:
                ano_budget = 0
                row_errors.append("ano invalido")
            categoria_nome = import_cell_value(categoria_raw)
            categoria_id = categories.get(categoria_nome.lower())
            if not categoria_nome:
                row_errors.append("categoria obrigatoria")
            elif not categoria_id:
                row_errors.append("categoria nao cadastrada")
            try:
                valor = money_value(valor_raw)
            except Exception:
                valor = Decimal("-1")
                row_errors.append("valor invalido")
            if mes < 1 or mes > 12:
                row_errors.append("mes fora de 1 a 12")
            if ano_budget <= 0:
                row_errors.append("ano obrigatorio")
            if valor < 0:
                row_errors.append("valor negativo")
            if row_errors:
                errors += 1
                error_messages.append(f"Linha {line_number}: {', '.join(row_errors)}")
                continue
            existing = query(
                "SELECT id FROM orcamentos WHERE categoria_id=%s AND ano=%s AND mes=%s",
                (categoria_id, ano_budget, mes),
                one=True,
            )
            if existing:
                execute(
                    "UPDATE orcamentos SET valor_orcamento=%s,orcamento_previsto=%s,data_importacao=CURRENT_TIMESTAMP,usuario_importacao=%s WHERE id=%s",
                    (valor, valor, "Importacao XLSX", existing["id"]),
                )
                updated += 1
            else:
                execute(
                    """
                    INSERT INTO orcamentos (mes,ano,categoria_id,valor_orcamento,orcamento_previsto,usuario_importacao)
                    VALUES (%s,%s,%s,%s,%s,%s)
                    """,
                    (mes, ano_budget, categoria_id, valor, valor, "Importacao XLSX"),
                )
                imported += 1
        workbook.close()
        summary = (
            f"Importacao concluida. Linhas lidas: {total}. "
            f"Novos orcamentos: {imported}. Atualizados: {updated}. Erros: {errors}."
        )
        if error_messages:
            summary += " " + " | ".join(error_messages[:6])
            if len(error_messages) > 6:
                summary += f" | Mais {len(error_messages) - 6} erro(s)."
        flash(summary, "success" if imported or updated else "warning")
    except Exception as error:
        flash(f"Nao foi possivel importar o orcamento: {error}", "danger")
    return redirect(url_for("cadastros", modulo="orcamentos"))


@app.route("/fornecedores", methods=["GET", "POST"])
def fornecedores():
    ensure_supplier_schema()
    if request.method == "POST":
        form = request.form
        fornecedor_id = form.get("id")
        params = (form["nome"], form["cnpj"], form.get("contato"), form.get("telefone"), form.get("email"), form["status"])
        try:
            if fornecedor_id:
                execute("UPDATE fornecedores SET nome=%s,cnpj=%s,contato=%s,telefone=%s,email=%s,status=%s WHERE id=%s", params + (fornecedor_id,))
            else:
                execute("INSERT INTO fornecedores (nome,cnpj,contato,telefone,email,status) VALUES (%s,%s,%s,%s,%s,%s)", params)
        except DB_INTEGRITY_ERROR_TYPES:
            flash("Ja existe um fornecedor cadastrado com este CNPJ.", "danger")
            return redirect(url_for("fornecedores"))
        flash("Fornecedor salvo com sucesso.", "success")
        return redirect(url_for("fornecedores"))
    return render_template("fornecedores.html", fornecedores=query("SELECT * FROM fornecedores ORDER BY nome"))


@app.get("/fornecedores/modelo")
def baixar_modelo_fornecedores():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Fornecedores"
    sheet.append(["Fornecedor", "CNPJ", "Contato", "Telefone", "Email"])
    sheet.append(["Fornecedor Exemplo", "00.000.000/0000-00", "JoÃ£o Silva", "(11) 99999-9999", "joao@exemplo.com"])
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name="modelo_fornecedores.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.post("/fornecedores/<int:item_id>/excluir")
def excluir_fornecedor(item_id):
    try:
        execute("DELETE FROM fornecedores WHERE id=%s", (item_id,))
        flash("Fornecedor excluido.", "success")
    except DB_ERROR_TYPES:
        flash("Fornecedor possui ordens vinculadas e nao pode ser excluido.", "danger")
    return redirect(url_for("fornecedores"))


@app.post("/fornecedores/importar")
def importar_fornecedores():
    ensure_supplier_schema()
    file = request.files.get("arquivo")
    if not file or not file.filename:
        flash("Selecione um arquivo XLSX para importar.", "danger")
        return redirect(url_for("fornecedores"))
    if not file.filename.lower().endswith(".xlsx"):
        flash("Envie um arquivo no formato .xlsx.", "danger")
        return redirect(url_for("fornecedores"))

    try:
        workbook = load_workbook(file, read_only=True, data_only=True)
        sheet = workbook.active
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        expected_headers = ["Fornecedor", "CNPJ", "Contato", "Telefone", "Email"]
        headers = [import_cell_value(value) for value in (header_row or []) if import_cell_value(value)]
        if headers != expected_headers:
            workbook.close()
            flash("O arquivo deve conter exatamente as colunas: Fornecedor, CNPJ, Contato, Telefone, Email.", "danger")
            return redirect(url_for("fornecedores"))

        total = imported = ignored = errors = 0
        error_messages = []
        seen_cnpjs = set()
        seen_names = set()
        for row in sheet.iter_rows(min_row=2, values_only=True):
            values = [import_cell_value(value) for value in row[:5]]
            if not any(values):
                continue

            total += 1
            nome, cnpj_raw, contato, telefone, email = values
            cnpj = normalize_cnpj(cnpj_raw)
            row_errors = []

            if not nome:
                row_errors.append("fornecedor sem nome")
            if cnpj and not is_valid_cnpj(cnpj):
                row_errors.append("CNPJ invalido")

            name_key = nome.strip().lower()
            cnpj_key = re.sub(r"\D", "", cnpj)
            if cnpj_key:
                if cnpj_key in seen_cnpjs or query("SELECT id FROM fornecedores WHERE REPLACE(REPLACE(REPLACE(cnpj,'.',''),'/',''),'-','')=%s", (cnpj_key,), one=True):
                    row_errors.append("CNPJ duplicado")
            elif name_key and (name_key in seen_names or query("SELECT id FROM fornecedores WHERE LOWER(nome)=%s", (name_key,), one=True)):
                row_errors.append("fornecedor duplicado por nome")

            if row_errors:
                errors += len(row_errors)
                ignored += 1
                error_messages.append(f"Linha {total + 1}: {', '.join(row_errors)}")
                continue

            execute(
                "INSERT INTO fornecedores (nome,cnpj,contato,telefone,email,status) VALUES (%s,%s,%s,%s,%s,%s)",
                (nome, cnpj or None, contato or None, telefone or None, email or None, "ativo"),
            )
            imported += 1
            if cnpj_key:
                seen_cnpjs.add(cnpj_key)
            else:
                seen_names.add(name_key)

        workbook.close()
        summary = (
            f"Importacao concluida. Total de linhas lidas: {total}. "
            f"Fornecedores importados: {imported}. Fornecedores ignorados: {ignored}. "
            f"Erros encontrados: {errors}."
        )
        if error_messages:
            summary += " " + " | ".join(error_messages[:5])
            if len(error_messages) > 5:
                summary += f" | Mais {len(error_messages) - 5} erro(s)."
        flash(summary, "success" if imported else "warning")
    except Exception as error:
        flash(f"Nao foi possivel importar o arquivo: {error}", "danger")
    return redirect(url_for("fornecedores"))


PRODUCT_IMPORT_HEADERS = [
    "Codigo",
    "Nome do Produto",
    "Categoria",
    "Unidade de Compra",
    "Quantidade por Unidade",
    "Tipo da Quantidade por Unidade",
    "Fornecedor",
    "Estoque de Seguranca",
    "Unidade do Estoque de Seguranca",
    "Ultimo Preco",
    "Ultimo Preco Base",
    "Ultima Compra",
    "Ultimo Fornecedor",
    "Preco Medio",
    "Preco Medio Base",
    "Status",
]

PRODUCT_IMPORT_KEYS = {
    "codigo": "Codigo",
    "nome do produto": "Nome do Produto",
    "categoria": "Categoria",
    "unidade de compra": "Unidade de Compra",
    "quantidade por unidade": "Quantidade por Unidade",
    "tipo da quantidade por unidade": "Tipo da Quantidade por Unidade",
    "fornecedor": "Fornecedor",
    "estoque de seguranca": "Estoque de Seguranca",
    "unidade do estoque de seguranca": "Unidade do Estoque de Seguranca",
    "ultimo preco": "Ultimo Preco",
    "ultimo preco base": "Ultimo Preco Base",
    "ultima compra": "Ultima Compra",
    "ultimo fornecedor": "Ultimo Fornecedor",
    "preco medio": "Preco Medio",
    "preco medio base": "Preco Medio Base",
    "status": "Status",
}

PRODUCT_IMPORT_REQUIRED = (
    "Codigo",
    "Nome do Produto",
    "Categoria",
    "Unidade de Compra",
    "Quantidade por Unidade",
    "Fornecedor",
    "Status",
)


def product_import_cell(row_data, name):
    return row_data.get(name, "")


def get_or_create_product_import_supplier(name, auto_create):
    supplier_name = import_cell_value(name)
    if not supplier_name:
        return None, False
    supplier = query("SELECT id,nome FROM fornecedores WHERE LOWER(nome)=LOWER(%s)", (supplier_name,), one=True)
    if supplier:
        return supplier, False
    if not auto_create:
        return None, False
    supplier_id = execute(
        "INSERT INTO fornecedores (nome,cnpj,contato,telefone,email,status) VALUES (%s,%s,%s,%s,%s,%s)",
        (supplier_name, None, "Nao informado", "Nao informado", "Nao informado", "ativo"),
    )
    return {"id": supplier_id, "nome": supplier_name}, True


def register_imported_product_history(product_id, row_data, supplier_name, unidade, quantidade_por_unidade, unidade_base):
    purchase_date = parse_optional_date_br(product_import_cell(row_data, "Ultima Compra"))
    last_supplier = import_cell_value(product_import_cell(row_data, "Ultimo Fornecedor")) or supplier_name
    unit_price = money_value(product_import_cell(row_data, "Ultimo Preco") or product_import_cell(row_data, "Preco Medio"))
    base_price = money_value(product_import_cell(row_data, "Ultimo Preco Base") or product_import_cell(row_data, "Preco Medio Base"))
    if not purchase_date or (unit_price <= 0 and base_price <= 0):
        return False
    if unit_price <= 0 and base_price > 0:
        unit_price = base_price * quantidade_por_unidade
    if base_price <= 0:
        base_price = unit_price / quantidade_por_unidade if quantidade_por_unidade > 0 else unit_price
    quantity = Decimal("1")
    base_qty = quantidade_por_unidade
    total_value = unit_price * quantity
    receipt_id = execute(
        """
        INSERT INTO nfe_importacoes
        (origem,fornecedor,fornecedor_cnpj,numero_nf,tipo_documento,serie,chave_nfe,data_emissao,data_entrada,valor_total,usuario_responsavel,status_recebimento,confirmado_em)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,NOW())
        """,
        (
            "Manual",
            last_supplier,
            "Nao informado",
            f"IMP-PROD-{product_id}",
            "Sem Documento",
            "XLSX",
            f"PROD-XLSX-{product_id}-{datetime.now().strftime('%Y%m%d%H%M%S%f')}",
            purchase_date,
            purchase_date,
            total_value,
            "Importacao XLSX",
            "Recebido",
        ),
    )
    item_id = execute(
        """
        INSERT INTO nfe_importacao_itens
        (nfe_importacao_id,produto_id,codigo_fornecedor,descricao,unidade,quantidade,valor_unitario,valor_total,status)
        SELECT %s,id,codigo,descricao,%s,%s,%s,%s,'vinculado'
        FROM produtos WHERE id=%s
        """,
        (receipt_id, unidade, quantity, unit_price, total_value, product_id),
    )
    execute(
        """
        INSERT INTO historico_custos
        (produto_id,fornecedor,documento,data_entrada,quantidade,unidade_compra,quantidade_por_unidade_compra,
         unidade_base,quantidade_total_base,valor_unitario,valor_unitario_base,valor_total,origem,recebimento_id,recebimento_item_id)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """,
        (
            product_id,
            last_supplier,
            f"IMP-PROD-{product_id}",
            purchase_date,
            quantity,
            unidade,
            quantidade_por_unidade,
            unidade_base,
            base_qty,
            unit_price,
            base_price,
            total_value,
            "Importacao XLSX",
            receipt_id,
            item_id,
        ),
    )
    return True


@app.route("/produtos", methods=["GET", "POST"])
def produtos():
    ensure_auxiliary_registries()
    ensure_product_schema()
    if request.method == "POST":
        form = request.form
        product_id = form.get("id")
        estoque_seguranca = decimal_value(form.get("estoque_seguranca"))
        if estoque_seguranca <= 0:
            flash("Estoque de seguranca deve ser um numero positivo.", "danger")
            return redirect(url_for("produtos"))
        quantidade_por_unidade = decimal_value(form.get("quantidade_por_unidade_compra") or 1)
        if quantidade_por_unidade <= 0:
            flash("Quantidade por unidade de compra deve ser um numero positivo.", "danger")
            return redirect(url_for("produtos"))
        codigo = form["codigo"].strip()
        params = (
            codigo,
            form["descricao"],
            form["categoria"],
            form["unidade"],
            quantidade_por_unidade,
            form.get("unidade_base") or form["unidade"],
            form["fornecedor_id"],
            estoque_seguranca,
            form["unidade"],
            form["status"],
        )
        try:
            if product_id:
                execute("UPDATE produtos SET codigo=%s,descricao=%s,categoria=%s,unidade=%s,quantidade_por_unidade_compra=%s,unidade_base=%s,fornecedor_id=%s,estoque_seguranca=%s,unidade_estoque_seguranca=%s,status=%s WHERE id=%s", params + (product_id,))
            else:
                execute("INSERT INTO produtos (codigo,descricao,categoria,unidade,quantidade_por_unidade_compra,unidade_base,fornecedor_id,estoque_seguranca,unidade_estoque_seguranca,status) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)", params)
        except DB_INTEGRITY_ERROR_TYPES:
            flash("Nao foi possivel salvar o produto. Verifique os campos vinculados.", "danger")
            return redirect(url_for("produtos"))
        flash("Produto salvo com sucesso.", "success")
        return redirect(url_for("produtos"))
    product_filters = {
        "codigo": request.args.get("codigo") or "",
        "produto": request.args.get("produto") or "",
        "categoria": request.args.get("categoria") or "",
    }
    where = []
    params = []
    if product_filters["codigo"]:
        where.append("p.codigo LIKE %s")
        params.append(f"%{product_filters['codigo']}%")
    if product_filters["produto"]:
        where.append("p.descricao LIKE %s")
        params.append(f"%{product_filters['produto']}%")
    if product_filters["categoria"]:
        where.append("p.categoria=%s")
        params.append(product_filters["categoria"])
    where_sql = "WHERE " + " AND ".join(where) if where else ""
    produtos_rows = query(
        f"""
        SELECT p.*, f.nome fornecedor_nome, u.nome unidade_nome,
          (SELECT h.valor_unitario FROM historico_custos h WHERE h.produto_id=p.id ORDER BY h.data_entrada DESC,h.id DESC LIMIT 1) ultimo_preco,
          (SELECT h.valor_unitario_base FROM historico_custos h WHERE h.produto_id=p.id ORDER BY h.data_entrada DESC,h.id DESC LIMIT 1) ultimo_preco_base,
          (SELECT h.data_entrada FROM historico_custos h WHERE h.produto_id=p.id ORDER BY h.data_entrada DESC,h.id DESC LIMIT 1) ultima_compra,
          (SELECT h.fornecedor FROM historico_custos h WHERE h.produto_id=p.id ORDER BY h.data_entrada DESC,h.id DESC LIMIT 1) ultimo_fornecedor,
          (SELECT AVG(h.valor_unitario) FROM historico_custos h WHERE h.produto_id=p.id) preco_medio_historico,
          (SELECT AVG(h.valor_unitario_base) FROM historico_custos h WHERE h.produto_id=p.id) preco_medio_base_historico
        FROM produtos p
        LEFT JOIN fornecedores f ON f.id=p.fornecedor_id
        LEFT JOIN unidades_medida u ON u.codigo=p.unidade
        {where_sql}
        ORDER BY p.descricao
        """,
        tuple(params),
    )
    categorias = query("SELECT nome categoria FROM cadastro_categorias WHERE status='ativo' ORDER BY nome")
    unidades = query("SELECT codigo,nome FROM unidades_medida WHERE status='ativo' ORDER BY nome")
    fornecedores_ativos = query("SELECT id,nome FROM fornecedores WHERE status='ativo' ORDER BY nome")
    return render_template(
        "produtos.html",
        produtos=produtos_rows,
        categorias=categorias,
        unidades=unidades,
        fornecedores=fornecedores_ativos,
        filtros=product_filters,
    )


@app.get("/produtos/modelo")
def baixar_modelo_produtos():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Produtos"
    sheet.append(PRODUCT_IMPORT_HEADERS)
    sheet.append([
        "MDF4MM",
        "MDF CRU 04MM 2,75 X 1,85",
        "2.1.2 Materia-Prima MDF",
        "PALLET",
        126,
        "CHAPAS",
        "Fornecedor Exemplo",
        10,
        "PALLET",
        7994.72,
        63.45,
        "26/06/2026",
        "Fornecedor Exemplo",
        7994.72,
        63.45,
        "Ativo",
    ])
    for column_cells in sheet.columns:
        letter = column_cells[0].column_letter
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[letter].width = min(max(max_length + 2, 14), 34)
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name="modelo_produtos.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.post("/produtos/importar")
def importar_produtos_xlsx():
    ensure_auxiliary_registries()
    ensure_product_schema()
    ensure_nfe_schema()
    file = request.files.get("arquivo")
    auto_create_suppliers = request.form.get("auto_create_suppliers") == "1"
    allow_existing_codes = request.form.get("allow_existing_codes") == "1"
    if not file or not file.filename:
        flash("Selecione um arquivo XLSX.", "danger")
        return redirect(url_for("produtos"))
    if not file.filename.lower().endswith(".xlsx"):
        flash("O arquivo deve estar no formato .xlsx.", "danger")
        return redirect(url_for("produtos"))
    try:
        workbook = load_workbook(file, data_only=True)
        sheet = workbook.active
        if sheet.max_row < 2:
            flash("A planilha esta vazia ou nao possui linhas de produtos.", "danger")
            return redirect(url_for("produtos"))
        raw_headers = [cell.value for cell in sheet[1]]
        headers = {}
        for index, header in enumerate(raw_headers, start=1):
            key = PRODUCT_IMPORT_KEYS.get(normalize_header(header))
            if key:
                headers[key] = index
        missing_headers = [header for header in PRODUCT_IMPORT_HEADERS if header not in headers]
        if missing_headers:
            flash("Cabecalho invalido. Colunas ausentes ou com nome incorreto: " + ", ".join(missing_headers), "danger")
            return redirect(url_for("produtos"))

        categories = {
            normalize_header(row["nome"]): row["nome"]
            for row in query("SELECT nome FROM cadastro_categorias WHERE status='ativo'")
        }
        units = {}
        for row in query("SELECT codigo,nome FROM unidades_medida WHERE status='ativo'"):
            units[normalize_header(row["codigo"])] = row["codigo"]
            units[normalize_header(row["nome"])] = row["codigo"]
        existing_codes = {
            import_cell_value(row["codigo"]).lower()
            for row in query("SELECT DISTINCT codigo FROM produtos WHERE codigo IS NOT NULL AND codigo<>''")
        }

        total = imported = error_count = history_count = created_suppliers = skipped_exact = 0
        errors = []
        duplicates = []
        seen_exact = set()

        for row_index in range(2, sheet.max_row + 1):
            values = [sheet.cell(row_index, col).value for col in range(1, len(raw_headers) + 1)]
            if all(import_cell_value(value) == "" for value in values):
                continue
            total += 1
            row_data = {
                name: sheet.cell(row_index, headers[name]).value
                for name in PRODUCT_IMPORT_HEADERS
            }
            row_errors = []
            for required in PRODUCT_IMPORT_REQUIRED:
                if import_cell_value(product_import_cell(row_data, required)) == "":
                    row_errors.append(f"{required} obrigatorio")

            codigo = import_cell_value(product_import_cell(row_data, "Codigo"))
            descricao = import_cell_value(product_import_cell(row_data, "Nome do Produto"))
            categoria_input = import_cell_value(product_import_cell(row_data, "Categoria"))
            unidade_input = import_cell_value(product_import_cell(row_data, "Unidade de Compra"))
            unidade_estoque_input = import_cell_value(product_import_cell(row_data, "Unidade do Estoque de Seguranca"))
            unidade_base = import_cell_value(product_import_cell(row_data, "Tipo da Quantidade por Unidade")) or unidade_input
            fornecedor_input = import_cell_value(product_import_cell(row_data, "Fornecedor"))
            status_input = normalize_header(product_import_cell(row_data, "Status"))

            categoria = categories.get(normalize_header(categoria_input))
            if categoria_input and not categoria:
                row_errors.append("Categoria nao cadastrada")
            unidade = units.get(normalize_header(unidade_input))
            if unidade_input and not unidade:
                row_errors.append("Unidade de Compra nao cadastrada")
            unidade_estoque = units.get(normalize_header(unidade_estoque_input)) if unidade_estoque_input else unidade
            if unidade_estoque_input and not unidade_estoque:
                row_errors.append("Unidade do Estoque de Seguranca nao cadastrada")
            status_map = {"ativo": "ativo", "inativo": "inativo"}
            status = status_map.get(status_input)
            if status_input and not status:
                row_errors.append("Status deve ser Ativo ou Inativo")

            try:
                quantidade_por_unidade = money_value(product_import_cell(row_data, "Quantidade por Unidade"))
                if quantidade_por_unidade <= 0:
                    row_errors.append("Quantidade por Unidade deve ser positiva")
            except Exception:
                quantidade_por_unidade = Decimal("0")
                row_errors.append("Quantidade por Unidade deve ser numerica")

            try:
                estoque_seguranca = money_value(product_import_cell(row_data, "Estoque de Seguranca"))
                if estoque_seguranca < 0:
                    row_errors.append("Estoque de Seguranca nao pode ser negativo")
            except Exception:
                estoque_seguranca = Decimal("0")
                row_errors.append("Estoque de Seguranca deve ser numerico")

            for price_field in ("Ultimo Preco", "Ultimo Preco Base", "Preco Medio", "Preco Medio Base"):
                if import_cell_value(product_import_cell(row_data, price_field)):
                    try:
                        if money_value(product_import_cell(row_data, price_field)) < 0:
                            row_errors.append(f"{price_field} nao pode ser negativo")
                    except Exception:
                        row_errors.append(f"{price_field} deve ser numerico")
            if import_cell_value(product_import_cell(row_data, "Ultima Compra")):
                try:
                    parse_optional_date_br(product_import_cell(row_data, "Ultima Compra"))
                except ValueError as error:
                    row_errors.append(str(error))

            duplicate_code = codigo.lower() in existing_codes
            if duplicate_code:
                duplicates.append(f"Linha {row_index}: codigo {codigo}")
                if not allow_existing_codes:
                    row_errors.append("Codigo ja cadastrado. Marque a opcao para importar mesmo assim.")

            supplier = query("SELECT id,nome FROM fornecedores WHERE LOWER(nome)=LOWER(%s)", (fornecedor_input,), one=True) if fornecedor_input else None
            should_create_supplier = bool(fornecedor_input and not supplier and auto_create_suppliers)
            if fornecedor_input and not supplier and not auto_create_suppliers:
                row_errors.append("Fornecedor nao cadastrado")

            exact_key = (codigo.lower(), descricao.lower(), str(supplier["id"]) if supplier else fornecedor_input.lower())
            if exact_key in seen_exact:
                row_errors.append("Produto duplicado dentro da planilha")
            existing_exact = query(
                """
                SELECT id FROM produtos
                WHERE LOWER(codigo)=LOWER(%s) AND LOWER(descricao)=LOWER(%s) AND fornecedor_id=%s
                LIMIT 1
                """,
                (codigo, descricao, supplier["id"] if supplier else 0),
                one=True,
            ) if supplier and codigo and descricao else None
            if existing_exact:
                skipped_exact += 1
                row_errors.append("Produto ja existe com mesmo codigo, nome e fornecedor")

            if row_errors:
                error_count += 1
                errors.append(f"Linha {row_index}: " + "; ".join(row_errors))
                continue

            supplier_created = False
            if should_create_supplier:
                supplier, supplier_created = get_or_create_product_import_supplier(fornecedor_input, True)

            product_id = execute(
                """
                INSERT INTO produtos
                (codigo,descricao,categoria,unidade,quantidade_por_unidade_compra,unidade_base,fornecedor_id,estoque_seguranca,unidade_estoque_seguranca,status)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """,
                (
                    codigo,
                    descricao,
                    categoria,
                    unidade,
                    quantidade_por_unidade,
                    unidade_base,
                    supplier["id"],
                    estoque_seguranca,
                    unidade_estoque,
                    status,
                ),
            )
            seen_exact.add(exact_key)
            existing_codes.add(codigo.lower())
            imported += 1
            if supplier_created:
                created_suppliers += 1
            if register_imported_product_history(product_id, row_data, supplier["nome"], unidade, quantidade_por_unidade, unidade_base):
                history_count += 1

        workbook.close()
        if total == 0:
            flash("A planilha nao possui linhas validas para leitura.", "danger")
            return redirect(url_for("produtos"))
        summary = (
            f"Importacao de produtos concluida. Total de linhas lidas: {total}. "
            f"Produtos importados: {imported}. Linhas com erro: {error_count}. "
            f"Historicos de custo criados: {history_count}. Fornecedores criados: {created_suppliers}."
        )
        if skipped_exact:
            summary += f" Produtos repetidos ignorados: {skipped_exact}."
        if duplicates:
            summary += " Codigos ja cadastrados identificados: " + " | ".join(duplicates[:5]) + "."
            if len(duplicates) > 5:
                summary += f" Mais {len(duplicates) - 5} codigo(s) duplicado(s)."
        if errors:
            summary += " Erros: " + " | ".join(errors[:8])
            if len(errors) > 8:
                summary += f" | Mais {len(errors) - 8} erro(s)."
        flash(summary, "success" if imported else "warning")
    except Exception as error:
        flash(f"Nao foi possivel importar produtos: {error}", "danger")
    return redirect(url_for("produtos"))


@app.post("/produtos/<int:item_id>/excluir")
def excluir_produto(item_id):
    try:
        execute("DELETE FROM produtos WHERE id=%s", (item_id,))
        flash("Produto excluido.", "success")
    except DB_ERROR_TYPES:
        flash("Produto possui ordens vinculadas e nao pode ser excluido.", "danger")
    return redirect(url_for("produtos"))


@app.route("/ordens")
def ordens():
    ensure_order_schema()
    if using_postgres():
        rows = query("""SELECT o.*, f.nome fornecedor, oo.numero_oc ordem_original_numero,
                        COUNT(i.id) total_itens,
                        STRING_AGG((p.codigo || ' - ' || p.descricao), ' | ' ORDER BY i.id) produtos_resumo,
                        SUM(i.quantidade) quantidade_itens,
                        SUM(i.quantidade * COALESCE(NULLIF(i.quantidade_por_unidade,0),1)) quantidade_base_total,
                        MIN(COALESCE(i.unidade_compra,p.unidade)) unidade,
                        MIN(COALESCE(i.unidade_base,p.unidade_base,p.unidade)) unidade_base
                        FROM ordens_compra o
                        JOIN fornecedores f ON f.id=o.fornecedor_id
                        LEFT JOIN ordens_compra oo ON oo.id=o.ordem_original_id
                        LEFT JOIN ordem_compra_itens i ON i.ordem_id=o.id
                        LEFT JOIN produtos p ON p.id=i.produto_id
                        GROUP BY o.id, f.nome, oo.numero_oc
                        ORDER BY o.id DESC""")
    else:
        rows = query("""SELECT o.*, f.nome fornecedor, oo.numero_oc ordem_original_numero,
                        COUNT(i.id) total_itens,
                        GROUP_CONCAT(CONCAT(p.codigo,' - ',p.descricao) ORDER BY i.id SEPARATOR ' | ') produtos_resumo,
                        SUM(i.quantidade) quantidade_itens,
                        SUM(i.quantidade * COALESCE(NULLIF(i.quantidade_por_unidade,0),1)) quantidade_base_total,
                        MIN(COALESCE(i.unidade_compra,p.unidade)) unidade,
                        MIN(COALESCE(i.unidade_base,p.unidade_base,p.unidade)) unidade_base
                        FROM ordens_compra o
                        JOIN fornecedores f ON f.id=o.fornecedor_id
                        LEFT JOIN ordens_compra oo ON oo.id=o.ordem_original_id
                        LEFT JOIN ordem_compra_itens i ON i.ordem_id=o.id
                        LEFT JOIN produtos p ON p.id=i.produto_id
                        GROUP BY o.id
                        ORDER BY o.id DESC""")
    return render_template("ordens.html", ordens=rows)

@app.route("/ordens/nova", methods=["GET", "POST"])
@app.route("/ordens/<int:item_id>/editar", methods=["GET", "POST"])
def nova_ordem(item_id=None):
    ensure_auxiliary_registries()
    ensure_order_schema()
    ordem = query("SELECT * FROM ordens_compra WHERE id=%s", (item_id,), one=True) if item_id else None
    ordem_itens = query(
        """
        SELECT i.*,p.codigo,p.descricao,p.fornecedor_id,p.unidade,p.unidade_base produto_unidade_base,
               p.quantidade_por_unidade_compra
        FROM ordem_compra_itens i
        JOIN produtos p ON p.id=i.produto_id
        WHERE i.ordem_id=%s
        ORDER BY i.id
        """,
        (item_id,),
    ) if item_id else []
    ordem_itens_json = [
        {
            "produto_id": row["produto_id"],
            "quantidade": str(row["quantidade"]),
            "preco_negociado": str(row["preco_negociado"]),
            "frete": str(row["frete"]),
        }
        for row in ordem_itens
    ]
    if request.method == "POST":
        form = request.form
        produto_ids = form.getlist("produto_id[]")
        quantidades = form.getlist("quantidade[]")
        precos = form.getlist("preco_negociado[]")
        fretes = form.getlist("frete_item[]")
        freight_mode = form.get("freight_mode") or "individual"
        frete_geral = money_value(form.get("frete_geral")) if freight_mode == "geral" else Decimal("0")
        parcelas = int(form.get("parcelas") or 0)
        prazos = [int(value) for value in form.getlist("prazos_parcelas[]") if str(value).strip() != ""]
        items = []
        subtotal = Decimal("0")
        for index, produto_id in enumerate(produto_ids):
            if not produto_id:
                continue
            quantidade = decimal_value(quantidades[index] if index < len(quantidades) else 0)
            preco = money_value(precos[index] if index < len(precos) else 0)
            frete_item = money_value(fretes[index] if index < len(fretes) else 0)
            if quantidade <= 0 or preco < 0 or frete_item < 0:
                flash("Revise os itens: produto, quantidade, preco e frete devem ser validos.", "danger")
                return redirect(request.url)
            product = query(
                "SELECT id,codigo,descricao,categoria,unidade,quantidade_por_unidade_compra,unidade_base FROM produtos WHERE id=%s AND fornecedor_id=%s",
                (produto_id, form["fornecedor_id"]),
                one=True,
            )
            if not product:
                flash("Todos os produtos devem pertencer ao fornecedor selecionado.", "danger")
                return redirect(request.url)
            value_without_freight = quantidade * preco
            subtotal += value_without_freight
            items.append({"product": product, "quantidade": quantidade, "preco": preco, "frete": frete_item, "base": value_without_freight})
        if not items or parcelas <= 0 or len(prazos) != parcelas:
            flash("Informe ao menos um produto, parcelas e prazo de cada parcela.", "danger")
            return redirect(request.url)
        if freight_mode == "geral" and frete_geral > 0:
            for item in items:
                item["frete"] = (item["base"] / subtotal * frete_geral) if subtotal > 0 else frete_geral / len(items)
        for item in items:
            item["total"] = item["base"] + item["frete"]
        total = sum(item["total"] for item in items)
        first = items[0]
        status_compra = form.get("status_compra") if item_id else "aguardando autorizacao"
        if status_compra in ("aprovada", "recebida"):
            status_compra = ordem["status_compra"] if ordem else "aguardando autorizacao"
        params = (
            form["data_preenchimento"],
            first["product"]["categoria"],
            form["fornecedor_id"],
            first["product"]["id"],
            first["quantidade"],
            first["preco"],
            sum(item["frete"] for item in items),
            total,
            form["data_entrega"],
            form["metodo_pagamento"],
            parcelas,
            ",".join(str(value) for value in prazos),
            prazos[0],
            form.get("nota_fiscal") or None,
            status_compra,
        )
        connection = get_db()
        cursor = connection.cursor()
        if item_id:
            try:
                cursor.execute("""UPDATE ordens_compra SET data_preenchimento=%s,tipo_material=%s,fornecedor_id=%s,produto_id=%s,
                           quantidade=%s,preco_negociado=%s,frete=%s,valor_total=%s,data_entrega=%s,metodo_pagamento=%s,
                           parcelas=%s,prazos_parcelas=%s,prazo_dias=%s,nota_fiscal=%s,status_compra=%s WHERE id=%s""", params + (item_id,))
                cursor.execute("DELETE FROM ordem_compra_itens WHERE ordem_id=%s", (item_id,))
                ordem_id = item_id
                for item in items:
                    product = item["product"]
                    cursor.execute(
                        """
                        INSERT INTO ordem_compra_itens
                        (ordem_id,produto_id,quantidade,preco_negociado,frete,valor_total_item,categoria,unidade_compra,quantidade_por_unidade,unidade_base)
                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                        """,
                        (ordem_id, product["id"], item["quantidade"], item["preco"], item["frete"], item["total"], product["categoria"], product["unidade"], product["quantidade_por_unidade_compra"] or 1, product["unidade_base"] or product["unidade"]),
                    )
                cursor.execute("UPDATE recebimentos SET data_prevista=%s WHERE ordem_id=%s", (form["data_entrega"], ordem_id))
                connection.commit()
            except Exception:
                connection.rollback()
                raise
            finally:
                cursor.close()
                connection.close()
        else:
            try:
                next_id_row = query("SELECT COALESCE(MAX(id),0)+1 next_id FROM ordens_compra", one=True)
                next_id = int(next_id_row["next_id"] or 1)
                numero = f"OC-{date.today().year}-{next_id:05d}"
                insert_sql = """INSERT INTO ordens_compra
                    (numero_oc,data_preenchimento,tipo_material,fornecedor_id,produto_id,quantidade,preco_negociado,frete,
                     valor_total,data_entrega,metodo_pagamento,parcelas,prazos_parcelas,prazo_dias,nota_fiscal,status_compra)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"""
                if using_postgres():
                    insert_sql += " RETURNING id"
                cursor.execute(insert_sql, (numero,) + params)
                if using_postgres():
                    ordem_id = cursor.fetchone()["id"]
                else:
                    ordem_id = cursor.lastrowid
                for item in items:
                    product = item["product"]
                    cursor.execute(
                        """
                        INSERT INTO ordem_compra_itens
                        (ordem_id,produto_id,quantidade,preco_negociado,frete,valor_total_item,categoria,unidade_compra,quantidade_por_unidade,unidade_base)
                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                        """,
                        (ordem_id, product["id"], item["quantidade"], item["preco"], item["frete"], item["total"], product["categoria"], product["unidade"], product["quantidade_por_unidade_compra"] or 1, product["unidade_base"] or product["unidade"]),
                    )
                cursor.execute("INSERT INTO autorizacoes (ordem_id,status) VALUES (%s,'pendente')", (ordem_id,))
                cursor.execute("INSERT INTO recebimentos (ordem_id,data_prevista,status) VALUES (%s,%s,'aguardando')", (ordem_id, form["data_entrega"]))
                connection.commit()
            except Exception:
                connection.rollback()
                raise
            finally:
                cursor.close()
                connection.close()
        flash("Ordem de compra salva com sucesso.", "success")
        return redirect(url_for("ordens"))
    return render_template(
        "nova_ordem.html",
        ordem=ordem,
        ordem_itens=ordem_itens_json,
        fornecedores=query("SELECT id,nome FROM fornecedores WHERE status='ativo' ORDER BY nome"),
        produtos=query("SELECT id,codigo,descricao,fornecedor_id,unidade,quantidade_por_unidade_compra,unidade_base,categoria FROM produtos WHERE status='ativo' ORDER BY descricao"),
        tipos_pagamento=query("SELECT nome FROM cadastro_tipos_pagamento WHERE status='ativo' ORDER BY nome"),
    )


@app.post("/ordens/<int:item_id>/excluir")
def excluir_ordem(item_id):
    execute("DELETE FROM ordens_compra WHERE id=%s", (item_id,))
    flash("Ordem excluida.", "success")
    return redirect(url_for("ordens"))


@app.get("/api/fornecedores/<int:fornecedor_id>/produtos")
def produtos_por_fornecedor(fornecedor_id):
    rows = query(
        """
        SELECT id,codigo,descricao,unidade,categoria,quantidade_por_unidade_compra,unidade_base
        FROM produtos
        WHERE fornecedor_id=%s AND status='ativo'
        ORDER BY descricao
        """,
        (fornecedor_id,),
    )
    return jsonify([serialize_row(row) for row in rows])


@app.get("/api/produtos/<int:produto_id>/consumo_categoria")
def consumo_categoria_produto(produto_id):
    product = query("SELECT id,categoria,unidade FROM produtos WHERE id=%s", (produto_id,), one=True)
    if not product:
        return jsonify({"error": "Produto nao encontrado"}), 404
    ref_date = date.today()
    if request.args.get("data_ref"):
        try:
            ref_date = datetime.strptime(request.args.get("data_ref"), "%Y-%m-%d").date()
        except ValueError:
            ref_date = date.today()
    year = ref_date.year
    month = ref_date.month
    budget = query(
        """
        SELECT o.valor_orcamento
        FROM orcamentos o
        JOIN cadastro_categorias c ON c.id=o.categoria_id
        WHERE c.nome=%s AND o.ano=%s AND o.mes=%s
        LIMIT 1
        """,
        (product["categoria"], year, month),
        one=True,
    )
    received = query(
        """
        SELECT COALESCE(SUM(h.valor_total),0) total
        FROM historico_custos h
        JOIN produtos p ON p.id=h.produto_id
        WHERE p.categoria=%s
          AND YEAR(h.data_entrada)=%s
          AND MONTH(h.data_entrada)=%s
        """,
        (product["categoria"], year, month),
        one=True,
    )
    committed_orders = query(
        """
        SELECT COALESCE(SUM(i.valor_total_item),0) total
        FROM ordem_compra_itens i
        JOIN ordens_compra o ON o.id=i.ordem_id
        JOIN produtos p ON p.id=i.produto_id
        WHERE p.categoria=%s
          AND YEAR(o.data_preenchimento)=%s
          AND MONTH(o.data_preenchimento)=%s
          AND o.status_compra IN ('aguardando autorizacao','aprovada','em compra')
        """,
        (product["categoria"], year, month),
        one=True,
    )
    received_quantities = query(
        """
        SELECT COALESCE(NULLIF(TRIM(p.unidade),''),'UN') unidade, COALESCE(SUM(h.quantidade),0) quantidade
        FROM historico_custos h
        JOIN produtos p ON p.id=h.produto_id
        WHERE p.categoria=%s
          AND YEAR(h.data_entrada)=%s
          AND MONTH(h.data_entrada)=%s
        GROUP BY COALESCE(NULLIF(TRIM(p.unidade),''),'UN')
        """,
        (product["categoria"], year, month),
    )
    order_quantities = query(
        """
        SELECT COALESCE(NULLIF(TRIM(i.unidade_compra),''),NULLIF(TRIM(p.unidade),''),'UN') unidade,
               COALESCE(SUM(i.quantidade),0) quantidade
        FROM ordem_compra_itens i
        JOIN ordens_compra o ON o.id=i.ordem_id
        JOIN produtos p ON p.id=i.produto_id
        WHERE p.categoria=%s
          AND YEAR(o.data_preenchimento)=%s
          AND MONTH(o.data_preenchimento)=%s
          AND o.status_compra IN ('aguardando autorizacao','aprovada','em compra')
        GROUP BY COALESCE(NULLIF(TRIM(i.unidade_compra),''),NULLIF(TRIM(p.unidade),''),'UN')
        """,
        (product["categoria"], year, month),
    )
    quantities = {}
    for row in received_quantities + order_quantities:
        unit = row["unidade"] or "UN"
        quantities[unit] = quantities.get(unit, Decimal("0")) + decimal_value(row["quantidade"])
    received_total = decimal_value(received["total"] if received else 0)
    committed_total = decimal_value(committed_orders["total"] if committed_orders else 0)
    return jsonify(
        {
            "categoria": product["categoria"],
            "referencia": f"{month:02d}/{year}",
            "unidade": product["unidade"] or "UN",
            "tem_orcamento": bool(budget),
            "orcamento": float(decimal_value(budget["valor_orcamento"]) if budget else Decimal("0")),
            "consumido": float(received_total + committed_total),
            "consumido_recebimentos": float(received_total),
            "consumido_ordens": float(committed_total),
            "quantidades": [
                {"unidade": unit, "quantidade": float(quantity)}
                for unit, quantity in sorted(quantities.items())
            ],
        }
    )


@app.get("/api/produtos/<int:produto_id>/historico")
def historico_produto(produto_id):
    product = query("SELECT * FROM produtos WHERE id=%s", (produto_id,), one=True)
    if not product:
        return jsonify({"error": "Produto nao encontrado"}), 404
    fornecedor_id = request.args.get("fornecedor_id")
    supplier = query("SELECT nome FROM fornecedores WHERE id=%s", (fornecedor_id,), one=True) if fornecedor_id else None
    supplier_name = supplier["nome"] if supplier else None
    supplier_filter = "AND h.fornecedor=%s" if supplier_name else ""
    supplier_params = (supplier_name,) if supplier_name else ()
    summary = query(
        f"""
        SELECT
          COUNT(*) total_compras,
          COALESCE(SUM(h.quantidade),0) quantidade_total,
          COALESCE(SUM(h.quantidade_total_base),0) quantidade_total_base,
          COALESCE(SUM(h.valor_total),0) valor_total_comprado,
          AVG(h.valor_unitario) preco_medio,
          AVG(h.valor_unitario_base) preco_medio_base,
          MIN(h.valor_unitario) menor_preco,
          MAX(h.valor_unitario) maior_preco,
          CASE WHEN SUM(h.quantidade) > 0 THEN SUM(h.valor_total)/SUM(h.quantidade) ELSE 0 END custo_medio_ponderado,
          CASE WHEN SUM(h.quantidade_total_base) > 0 THEN SUM(h.valor_total)/SUM(h.quantidade_total_base) ELSE 0 END custo_medio_base_ponderado
        FROM historico_custos h
        WHERE h.produto_id=%s {supplier_filter}
        """,
        (produto_id,) + supplier_params,
        one=True,
    )
    latest = query(
        f"""
        SELECT h.*
        FROM historico_custos h
        WHERE h.produto_id=%s {supplier_filter}
        ORDER BY h.data_entrada DESC,h.id DESC
        LIMIT 1
        """,
        (produto_id,) + supplier_params,
        one=True,
    )
    purchases = query(
        """
        SELECT h.*,n.status_recebimento
        FROM historico_custos h
        LEFT JOIN nfe_importacoes n ON n.id=h.recebimento_id
        WHERE h.produto_id=%s
        ORDER BY h.data_entrada DESC,h.id DESC
        LIMIT 8
        """,
        (produto_id,),
    )
    chart = query(
        """
        SELECT data_entrada,valor_unitario,valor_unitario_base,fornecedor,documento
        FROM historico_custos
        WHERE produto_id=%s
        ORDER BY data_entrada,id
        LIMIT 30
        """,
        (produto_id,),
    )
    best_supplier = query(
        """
        SELECT fornecedor,AVG(valor_unitario) preco_medio,AVG(valor_unitario_base) preco_medio_base,COUNT(*) total_compras
        FROM historico_custos
        WHERE produto_id=%s
        GROUP BY fornecedor
        ORDER BY preco_medio ASC,total_compras DESC
        LIMIT 1
        """,
        (produto_id,),
        one=True,
    )
    general = query(
        """
        SELECT AVG(valor_unitario) preco_medio_geral,AVG(valor_unitario_base) preco_medio_base_geral
        FROM historico_custos
        WHERE produto_id=%s
        """,
        (produto_id,),
        one=True,
    )
    avg_price = decimal_value(summary["preco_medio"]) if summary and summary["preco_medio"] is not None else Decimal("0")
    last_price = decimal_value(latest["valor_unitario"]) if latest else Decimal("0")
    alert_percent = ((last_price - avg_price) / avg_price * 100) if avg_price and last_price > avg_price else Decimal("0")
    return jsonify(
        {
            "produto": serialize_row(product),
            "resumo": serialize_row(summary),
            "ultima_compra": serialize_row(latest) if latest else None,
            "ultimas_compras": [serialize_row(row) for row in purchases],
            "grafico": [serialize_row(row) for row in chart],
            "melhor_fornecedor": serialize_row(best_supplier) if best_supplier else None,
            "preco_medio_geral": float(decimal_value(general["preco_medio_geral"]) if general and general["preco_medio_geral"] is not None else Decimal("0")),
            "preco_medio_base_geral": float(decimal_value(general["preco_medio_base_geral"]) if general and general["preco_medio_base_geral"] is not None else Decimal("0")),
            "alerta_percentual": float(alert_percent),
        }
    )


@app.route("/historico-compras")
def historico_compras():
    ensure_nfe_schema()
    filters = request_purchase_history_filters()
    where_sql, params = purchase_history_where(filters)
    rows = query(
        f"""
        SELECT
          h.id,h.data_entrada,h.fornecedor,h.documento,h.quantidade,h.unidade_compra,
          h.quantidade_por_unidade_compra,h.unidade_base,h.quantidade_total_base,
          h.valor_unitario,h.valor_unitario_base,h.valor_total,h.origem,
          p.id produto_id,p.codigo,p.descricao produto,p.categoria,p.unidade,
          n.status_recebimento
        FROM historico_custos h
        JOIN produtos p ON p.id=h.produto_id
        LEFT JOIN nfe_importacoes n ON n.id=h.recebimento_id
        {where_sql}
        ORDER BY h.data_entrada DESC,h.id DESC
        LIMIT 500
        """,
        tuple(params),
    )
    summary = query(
        f"""
        SELECT
          COALESCE(SUM(h.valor_total),0) total_comprado,
          COALESCE(SUM(h.quantidade),0) quantidade_total,
          COALESCE(SUM(h.quantidade_total_base),0) quantidade_total_base,
          AVG(h.valor_unitario) preco_medio_geral,
          AVG(h.valor_unitario_base) preco_medio_base_geral,
          COUNT(*) total_registros
        FROM historico_custos h
        JOIN produtos p ON p.id=h.produto_id
        LEFT JOIN nfe_importacoes n ON n.id=h.recebimento_id
        {where_sql}
        """,
        tuple(params),
        one=True,
    )
    supplier_most = query(
        f"""
        SELECT h.fornecedor,COUNT(*) total
        FROM historico_custos h
        JOIN produtos p ON p.id=h.produto_id
        LEFT JOIN nfe_importacoes n ON n.id=h.recebimento_id
        {where_sql}
        GROUP BY h.fornecedor
        ORDER BY total DESC
        LIMIT 1
        """,
        tuple(params),
        one=True,
    )
    product_increase = query(
        f"""
        SELECT p.codigo,p.descricao produto,
               MAX(h.valor_unitario)-AVG(h.valor_unitario) aumento
        FROM historico_custos h
        JOIN produtos p ON p.id=h.produto_id
        LEFT JOIN nfe_importacoes n ON n.id=h.recebimento_id
        {where_sql}
        GROUP BY p.id
        HAVING COUNT(*) > 1
        ORDER BY aumento DESC
        LIMIT 1
        """,
        tuple(params),
        one=True,
    )
    grouped = {}
    for entry in rows:
        code = entry["codigo"] or "Nao registrado"
        group = grouped.setdefault(
            code,
            {
                "codigo": code,
                "produto": entry["produto"],
                "categoria": entry["categoria"],
                "unidade": entry["unidade"],
                "unidade_base": entry["unidade_base"] or entry["unidade"],
                "quantidade_por_unidade_compra": decimal_value(entry["quantidade_por_unidade_compra"] or 1),
                "quantidade_total": Decimal("0"),
                "quantidades_compra": {},
                "quantidade_total_base": Decimal("0"),
                "valor_total_comprado": Decimal("0"),
                "entradas_count": 0,
                "ultimo_preco": decimal_value(entry["valor_unitario"]),
                "ultimo_preco_base": decimal_value(entry["valor_unitario_base"]),
                "ultima_compra": entry["data_entrada"],
                "ultimo_fornecedor": entry["fornecedor"],
                "entries": [],
                "suppliers": {},
            },
        )
        qty = decimal_value(entry["quantidade"])
        base_qty = decimal_value(entry["quantidade_total_base"])
        value = decimal_value(entry["valor_total"])
        purchase_unit = entry["unidade_compra"] or entry["unidade"] or "UN"
        group["quantidade_total"] += qty
        group["quantidades_compra"][purchase_unit] = group["quantidades_compra"].get(purchase_unit, Decimal("0")) + qty
        group["quantidade_total_base"] += base_qty
        group["valor_total_comprado"] += value
        group["entradas_count"] += 1
        group["entries"].append(entry)
        supplier = group["suppliers"].setdefault(
            entry["fornecedor"],
            {
                "fornecedor": entry["fornecedor"],
                "quantidade_total": Decimal("0"),
                "quantidades_compra": {},
                "quantidade_total_base": Decimal("0"),
                "valor_total_comprado": Decimal("0"),
                "entradas_count": 0,
                "ultimo_preco": decimal_value(entry["valor_unitario"]),
                "ultimo_preco_base": decimal_value(entry["valor_unitario_base"]),
                "ultima_compra": entry["data_entrada"],
                "menor_preco": decimal_value(entry["valor_unitario"]),
            },
        )
        supplier["quantidade_total"] += qty
        supplier["quantidades_compra"][purchase_unit] = supplier["quantidades_compra"].get(purchase_unit, Decimal("0")) + qty
        supplier["quantidade_total_base"] += base_qty
        supplier["valor_total_comprado"] += value
        supplier["entradas_count"] += 1
        supplier["menor_preco"] = min(supplier["menor_preco"], decimal_value(entry["valor_unitario"]))
    grouped_rows = []
    for group in grouped.values():
        group["quantidades_compra_lista"] = [
            {"unidade": unit, "quantidade": quantity}
            for unit, quantity in sorted(group["quantidades_compra"].items())
        ]
        group["preco_medio_historico"] = (
            group["valor_total_comprado"] / group["quantidade_total"]
            if group["quantidade_total"] > 0
            else Decimal("0")
        )
        group["preco_medio_base_historico"] = (
            group["valor_total_comprado"] / group["quantidade_total_base"]
            if group["quantidade_total_base"] > 0
            else Decimal("0")
        )
        avg_price = group["preco_medio_historico"]
        last_price = group["ultimo_preco"]
        group["variacao_percentual"] = ((last_price - avg_price) / avg_price * 100) if avg_price else Decimal("0")
        supplier_rows = []
        for supplier in group["suppliers"].values():
            supplier["quantidades_compra_lista"] = [
                {"unidade": unit, "quantidade": quantity}
                for unit, quantity in sorted(supplier["quantidades_compra"].items())
            ]
            supplier["preco_medio"] = (
                supplier["valor_total_comprado"] / supplier["quantidade_total"]
                if supplier["quantidade_total"] > 0
                else Decimal("0")
            )
            supplier["preco_medio_base"] = (
                supplier["valor_total_comprado"] / supplier["quantidade_total_base"]
                if supplier["quantidade_total_base"] > 0
                else Decimal("0")
            )
            supplier_rows.append(supplier)
        group["supplier_summary"] = sorted(supplier_rows, key=lambda item: item["preco_medio"])
        grouped_rows.append(group)
    grouped_rows.sort(key=lambda item: (item["ultima_compra"], item["codigo"]), reverse=True)
    product_increase = None
    if grouped_rows:
        product_increase = max(grouped_rows, key=lambda item: item["variacao_percentual"])
        if product_increase["variacao_percentual"] <= 0:
            product_increase = None
    summary["total_codigos"] = len(grouped_rows)
    categorias = query("SELECT nome FROM cadastro_categorias WHERE status='ativo' ORDER BY nome")
    unidades = query("SELECT codigo,nome FROM unidades_medida WHERE status='ativo' ORDER BY nome")
    filter_options = {
        "codigos": query(
            """
            SELECT DISTINCT p.codigo valor
            FROM historico_custos h
            JOIN produtos p ON p.id=h.produto_id
            WHERE TRIM(COALESCE(p.codigo,'')) <> ''
            ORDER BY p.codigo
            """
        ),
        "produtos": query(
            """
            SELECT DISTINCT p.descricao valor
            FROM historico_custos h
            JOIN produtos p ON p.id=h.produto_id
            WHERE TRIM(COALESCE(p.descricao,'')) <> ''
            ORDER BY p.descricao
            """
        ),
        "fornecedores": query(
            """
            SELECT DISTINCT h.fornecedor valor
            FROM historico_custos h
            WHERE TRIM(COALESCE(h.fornecedor,'')) <> ''
            ORDER BY h.fornecedor
            """
        ),
        "documentos": query(
            """
            SELECT DISTINCT h.documento valor
            FROM historico_custos h
            WHERE TRIM(COALESCE(h.documento,'')) <> ''
            ORDER BY h.documento
            """
        ),
        "origens": ["XML", "Manual", "Importacao XLSX"],
        "status": ["Recebido", "Recebido Parcialmente", "Pendente", "Produto Pendente de Cadastro", "Aguardando Ordem de Compra"],
    }
    return render_template(
        "historico_compras.html",
        rows=rows,
        grouped_rows=grouped_rows,
        summary=summary,
        supplier_most=supplier_most,
        product_increase=product_increase,
        filtros=filters,
        categorias=categorias,
        unidades=unidades,
        filter_options=filter_options,
    )


@app.get("/historico-compras/exportar")
def exportar_historico_compras():
    ensure_nfe_schema()
    filters = request_purchase_history_filters()
    where_sql, params = purchase_history_where(filters)
    rows = query(
        f"""
        SELECT
          h.id,h.data_entrada,h.fornecedor,h.documento,h.quantidade,h.unidade_compra,
          h.quantidade_por_unidade_compra,h.unidade_base,h.quantidade_total_base,
          h.valor_unitario,h.valor_unitario_base,h.valor_total,h.origem,
          p.id produto_id,p.codigo,p.descricao produto,p.categoria,p.unidade,
          n.status_recebimento
        FROM historico_custos h
        JOIN produtos p ON p.id=h.produto_id
        LEFT JOIN nfe_importacoes n ON n.id=h.recebimento_id
        {where_sql}
        ORDER BY h.data_entrada DESC,h.id DESC
        """,
        tuple(params),
    )
    grouped = {}
    for entry in rows:
        code = entry["codigo"] or "Nao registrado"
        group = grouped.setdefault(
            code,
            {
                "codigo": code,
                "produto": entry["produto"],
                "categoria": entry["categoria"],
                "unidade_base": entry["unidade_base"] or entry["unidade"],
                "quantidades_compra": {},
                "quantidade_total": Decimal("0"),
                "quantidade_total_base": Decimal("0"),
                "valor_total_comprado": Decimal("0"),
                "entradas_count": 0,
                "ultimo_preco": decimal_value(entry["valor_unitario"]),
                "ultimo_preco_base": decimal_value(entry["valor_unitario_base"]),
                "ultima_compra": entry["data_entrada"],
                "ultimo_fornecedor": entry["fornecedor"],
            },
        )
        qty = decimal_value(entry["quantidade"])
        base_qty = decimal_value(entry["quantidade_total_base"])
        value = decimal_value(entry["valor_total"])
        purchase_unit = entry["unidade_compra"] or entry["unidade"] or "UN"
        group["quantidade_total"] += qty
        group["quantidade_total_base"] += base_qty
        group["valor_total_comprado"] += value
        group["entradas_count"] += 1
        group["quantidades_compra"][purchase_unit] = group["quantidades_compra"].get(purchase_unit, Decimal("0")) + qty
    grouped_rows = []
    for group in grouped.values():
        group["qtd_compra"] = "\n".join(
            f"{quantity} {unit}" for unit, quantity in sorted(group["quantidades_compra"].items())
        )
        group["preco_medio_compra"] = (
            group["valor_total_comprado"] / group["quantidade_total"]
            if group["quantidade_total"] > 0
            else Decimal("0")
        )
        group["preco_medio_base"] = (
            group["valor_total_comprado"] / group["quantidade_total_base"]
            if group["quantidade_total_base"] > 0
            else Decimal("0")
        )
        grouped_rows.append(group)
    grouped_rows.sort(key=lambda item: (item["ultima_compra"], item["codigo"]), reverse=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Consolidado"
    sheet.append([
        "Codigo interno", "Produto", "Categoria", "Qtd. compra", "Qtd. base",
        "Preco medio compra", "Preco medio base", "Ultimo preco compra",
        "Ultimo preco base", "Ultima compra", "Ultimo fornecedor", "Valor total", "Entradas"
    ])
    for row in grouped_rows:
        sheet.append([
            row["codigo"],
            row["produto"],
            row["categoria"],
            row["qtd_compra"],
            float(row["quantidade_total_base"]),
            float(row["preco_medio_compra"]),
            float(row["preco_medio_base"]),
            float(row["ultimo_preco"]),
            float(row["ultimo_preco_base"]),
            row["ultima_compra"],
            row["ultimo_fornecedor"],
            float(row["valor_total_comprado"]),
            row["entradas_count"],
        ])

    detail = workbook.create_sheet("Entradas")
    detail.append([
        "Data", "Codigo interno", "Produto", "Categoria", "Fornecedor", "NF/Documento",
        "Qtd. compra", "Unidade compra", "Qtd. por unidade", "Unidade base",
        "Qtd. base", "Valor unidade compra", "Valor base", "Valor total", "Origem", "Status"
    ])
    for row in rows:
        detail.append([
            row["data_entrada"],
            row["codigo"] or "Nao registrado",
            row["produto"],
            row["categoria"],
            row["fornecedor"],
            row["documento"],
            float(decimal_value(row["quantidade"])),
            row["unidade_compra"] or row["unidade"],
            float(decimal_value(row["quantidade_por_unidade_compra"] or 1)),
            row["unidade_base"] or row["unidade"],
            float(decimal_value(row["quantidade_total_base"])),
            float(decimal_value(row["valor_unitario"])),
            float(decimal_value(row["valor_unitario_base"])),
            float(decimal_value(row["valor_total"])),
            row["origem"],
            row["status_recebimento"] or "Recebido",
        ])

    for ws in workbook.worksheets:
        ws.freeze_panes = "A2"
        for column_cells in ws.columns:
            letter = column_cells[0].column_letter
            max_length = max(len(str(cell.value or "")) for cell in column_cells[:100])
            ws.column_dimensions[letter].width = min(max(max_length + 2, 12), 42)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=f"historico_compras_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/autorizacoes", methods=["GET", "POST"])
def autorizacoes():
    ensure_order_schema()
    ensure_monthly_budget_schema()
    if not using_postgres() and not column_exists("autorizacoes", "decidido_em"):
        execute("ALTER TABLE autorizacoes ADD COLUMN decidido_em DATETIME NULL AFTER data_aprovacao")
    if request.method == "POST":
        form = request.form
        status = form.get("status")
        motivo = (form.get("observacao") or "").strip()
        if status == "reprovada" and not motivo:
            flash("Informe o motivo da reprovacao.", "danger")
            return redirect(url_for("autorizacoes"))
        if status not in ("aprovada", "reprovada"):
            flash("Decisao de autorizacao invalida.", "danger")
            return redirect(url_for("autorizacoes"))
        if using_postgres():
            execute(
                """
                UPDATE autorizacoes
                SET autorizado_por=%s,data_aprovacao=CURRENT_DATE,decidido_em=CURRENT_TIMESTAMP,status=%s,observacao=%s
                WHERE id=%s
                """,
                (form.get("autorizado_por") or "Equipe de Compras", status, motivo or None, form["id"]),
            )
            status_ordem = {"aprovada": "aprovada", "reprovada": "reprovada"}[status]
            execute(
                """
                UPDATE ordens_compra o
                SET status_compra=%s
                FROM autorizacoes a
                WHERE a.ordem_id=o.id AND a.id=%s
                """,
                (status_ordem, form["id"]),
            )
        else:
            execute(
                """
                UPDATE autorizacoes
                SET autorizado_por=%s,data_aprovacao=CURDATE(),decidido_em=NOW(),status=%s,observacao=%s
                WHERE id=%s
                """,
                (form.get("autorizado_por") or "Equipe de Compras", status, motivo or None, form["id"]),
            )
            status_ordem = {"aprovada": "aprovada", "reprovada": "reprovada"}[status]
            execute("UPDATE ordens_compra o JOIN autorizacoes a ON a.ordem_id=o.id SET o.status_compra=%s WHERE a.id=%s", (status_ordem, form["id"]))
        flash("Autorizacao atualizada.", "success")
        return redirect(url_for("autorizacoes"))
    if using_postgres():
        rows = query("""SELECT a.*,o.numero_oc,o.valor_total,o.data_entrega,o.data_preenchimento,
                               f.nome fornecedor,
                               COALESCE(SUM(i.quantidade),o.quantidade) quantidade,
                               MIN(COALESCE(i.unidade_compra,p.unidade)) unidade,
                               1 quantidade_por_unidade_compra,
                               MIN(COALESCE(i.unidade_base,p.unidade_base,p.unidade)) unidade_base
                        FROM autorizacoes a
                        JOIN ordens_compra o ON o.id=a.ordem_id
                        JOIN fornecedores f ON f.id=o.fornecedor_id
                        LEFT JOIN ordem_compra_itens i ON i.ordem_id=o.id
                        LEFT JOIN produtos p ON p.id=i.produto_id
                        GROUP BY a.id,o.id,f.nome
                        ORDER BY a.id DESC""")
    else:
        rows = query("""SELECT a.*,o.numero_oc,o.valor_total,o.data_entrega,o.data_preenchimento,
                               f.nome fornecedor,
                               COALESCE(SUM(i.quantidade),o.quantidade) quantidade,
                               MIN(COALESCE(i.unidade_compra,p.unidade)) unidade,
                               1 quantidade_por_unidade_compra,
                               MIN(COALESCE(i.unidade_base,p.unidade_base,p.unidade)) unidade_base
                        FROM autorizacoes a
                        JOIN ordens_compra o ON o.id=a.ordem_id
                        JOIN fornecedores f ON f.id=o.fornecedor_id
                        LEFT JOIN ordem_compra_itens i ON i.ordem_id=o.id
                        LEFT JOIN produtos p ON p.id=i.produto_id
                        GROUP BY a.id
                        ORDER BY a.id DESC""")
    today = date.today()
    for row in rows:
        items = query(
            """
            SELECT i.*,p.descricao produto,p.codigo,p.categoria,
                   COALESCE(i.unidade_compra,p.unidade,'UN') unidade
            FROM ordem_compra_itens i
            JOIN produtos p ON p.id=i.produto_id
            WHERE i.ordem_id=%s
            ORDER BY i.id
            """,
            (row["ordem_id"],),
        )
        budget_by_category = []
        categories = {}
        for item in items:
            category = item["categoria"] or "Sem categoria"
            categories[category] = categories.get(category, Decimal("0")) + decimal_value(item["valor_total_item"])
        for category, order_value in categories.items():
            budget = query(
                """
                SELECT o.valor_orcamento
                FROM orcamentos o
                JOIN cadastro_categorias c ON c.id=o.categoria_id
                WHERE c.nome=%s AND o.ano=%s AND o.mes=%s
                LIMIT 1
                """,
                (category, today.year, today.month),
                one=True,
            )
            if using_postgres():
                used_history = query(
                    """
                    SELECT COALESCE(SUM(h.valor_total),0) total
                    FROM historico_custos h
                    JOIN produtos p ON p.id=h.produto_id
                    WHERE p.categoria=%s
                      AND EXTRACT(YEAR FROM h.data_entrada)=%s
                      AND EXTRACT(MONTH FROM h.data_entrada)=%s
                    """,
                    (category, today.year, today.month),
                    one=True,
                )
                used_orders = query(
                    """
                    SELECT COALESCE(SUM(i.valor_total_item),0) total
                    FROM ordem_compra_itens i
                    JOIN ordens_compra o ON o.id=i.ordem_id
                    JOIN produtos p ON p.id=i.produto_id
                    WHERE p.categoria=%s
                      AND o.id<>%s
                      AND EXTRACT(YEAR FROM o.data_preenchimento)=%s
                      AND EXTRACT(MONTH FROM o.data_preenchimento)=%s
                      AND o.status_compra IN ('aguardando autorizacao','aprovada','em compra')
                    """,
                    (category, row["ordem_id"], today.year, today.month),
                    one=True,
                )
            else:
                used_history = query(
                    """
                    SELECT COALESCE(SUM(h.valor_total),0) total
                    FROM historico_custos h
                    JOIN produtos p ON p.id=h.produto_id
                    WHERE p.categoria=%s
                      AND YEAR(h.data_entrada)=%s
                      AND MONTH(h.data_entrada)=%s
                    """,
                    (category, today.year, today.month),
                    one=True,
                )
                used_orders = query(
                    """
                    SELECT COALESCE(SUM(i.valor_total_item),0) total
                    FROM ordem_compra_itens i
                    JOIN ordens_compra o ON o.id=i.ordem_id
                    JOIN produtos p ON p.id=i.produto_id
                    WHERE p.categoria=%s
                      AND o.id<>%s
                      AND YEAR(o.data_preenchimento)=%s
                      AND MONTH(o.data_preenchimento)=%s
                      AND o.status_compra IN ('aguardando autorizacao','aprovada','em compra')
                    """,
                    (category, row["ordem_id"], today.year, today.month),
                    one=True,
                )
            budget_value = decimal_value(budget["valor_orcamento"]) if budget else Decimal("0")
            used_value = decimal_value(used_history["total"] if used_history else 0) + decimal_value(used_orders["total"] if used_orders else 0)
            balance_after = budget_value - used_value - order_value
            budget_by_category.append(
                {
                    "categoria": category,
                    "tem_orcamento": bool(budget),
                    "orcamento": float(budget_value),
                    "utilizado": float(used_value),
                    "valor_compra": float(order_value),
                    "saldo_apos": float(balance_after),
                    "estouro": float(abs(balance_after)) if budget and balance_after < 0 else 0,
                }
            )
        row["resumo_json"] = {
            "id": row["id"],
            "numero_oc": row["numero_oc"],
            "valor_total": float(decimal_value(row["valor_total"])),
            "fornecedor": row["fornecedor"],
            "data_entrega": row["data_entrega"].strftime("%d/%m/%Y") if row["data_entrega"] else "-",
            "produtos": [
                {
                    "produto": item["produto"],
                    "quantidade": float(decimal_value(item["quantidade"])),
                    "unidade": item["unidade"],
                    "valor_unitario": float(decimal_value(item["preco_negociado"])),
                    "valor_total": float(decimal_value(item["valor_total_item"])),
                }
                for item in items
            ],
            "orcamentos": budget_by_category,
        }
    return render_template("autorizacoes.html", autorizacoes=rows)


@app.route("/recebimentos", methods=["GET", "POST"])
def recebimentos():
    ensure_auxiliary_registries()
    ensure_product_schema()
    ensure_nfe_schema()
    if request.method == "POST":
        form = request.form
        try:
            real = normalize_iso_date(form.get("data_real"), "Data real")
        except ValueError as error:
            flash(str(error), "danger")
            return redirect(url_for("recebimentos"))
        status = form["status"]
        execute("UPDATE recebimentos SET data_real=%s,status=%s WHERE id=%s", (real, status, form["id"]))
        if status == "entregue completo":
            if using_postgres():
                execute(
                    """
                    UPDATE ordens_compra o
                    SET status_compra='recebida'
                    FROM recebimentos r
                    WHERE r.ordem_id=o.id AND r.id=%s
                    """,
                    (form["id"],),
                )
            else:
                execute("UPDATE ordens_compra o JOIN recebimentos r ON r.ordem_id=o.id SET o.status_compra='recebida' WHERE r.id=%s", (form["id"],))
        flash("Recebimento atualizado.", "success")
        return redirect(url_for("recebimentos"))
    if using_postgres():
        rows = query("""SELECT r.*,o.numero_oc,
            CASE WHEN r.data_real IS NOT NULL THEN GREATEST((r.data_real::date-r.data_prevista::date),0)
                 WHEN r.data_prevista<CURRENT_DATE AND r.status NOT IN ('entregue parcial','entregue completo') THEN (CURRENT_DATE-r.data_prevista::date) ELSE 0 END dias_atraso,
            CASE WHEN r.data_prevista<CURRENT_DATE AND r.data_real IS NULL AND r.status='aguardando' THEN 'atrasado' ELSE r.status END status_exibicao
            FROM recebimentos r JOIN ordens_compra o ON o.id=r.ordem_id ORDER BY r.data_prevista""")
    else:
        rows = query("""SELECT r.*,o.numero_oc,
            CASE WHEN r.data_real IS NOT NULL THEN GREATEST(DATEDIFF(r.data_real,r.data_prevista),0)
                 WHEN r.data_prevista<CURDATE() AND r.status NOT IN ('entregue parcial','entregue completo') THEN DATEDIFF(CURDATE(),r.data_prevista) ELSE 0 END dias_atraso,
            CASE WHEN r.data_prevista<CURDATE() AND r.data_real IS NULL AND r.status='aguardando' THEN 'atrasado' ELSE r.status END status_exibicao
            FROM recebimentos r JOIN ordens_compra o ON o.id=r.ordem_id ORDER BY r.data_prevista""")
    filters = {
        "origem": request.args.get("origem") or "",
        "fornecedor": request.args.get("fornecedor") or "",
        "produto": request.args.get("produto") or "",
        "documento": request.args.get("documento") or "",
        "status": request.args.get("status") or "",
        "data_inicio": request.args.get("data_inicio") or "",
        "data_fim": request.args.get("data_fim") or "",
    }
    refresh_receipt_statuses()
    where = []
    params = []
    if filters["origem"]:
        where.append("n.origem=%s")
        params.append(filters["origem"])
    if filters["fornecedor"]:
        where.append("n.fornecedor LIKE %s")
        params.append(f"%{filters['fornecedor']}%")
    if filters["documento"]:
        where.append("n.numero_nf LIKE %s")
        params.append(f"%{filters['documento']}%")
    if filters["status"]:
        where.append("n.status_recebimento=%s")
        params.append(filters["status"])
    if filters["data_inicio"]:
        where.append("n.data_entrada >= %s")
        params.append(filters["data_inicio"])
    if filters["data_fim"]:
        where.append("n.data_entrada <= %s")
        params.append(filters["data_fim"])
    if filters["produto"]:
        where.append(
            """
            EXISTS (
              SELECT 1 FROM nfe_importacao_itens ix
              LEFT JOIN produtos px ON px.id=ix.produto_id
              WHERE ix.nfe_importacao_id=n.id
                AND (ix.descricao LIKE %s OR px.descricao LIKE %s OR px.codigo LIKE %s)
            )
            """
        )
        term = f"%{filters['produto']}%"
        params.extend([term, term, term])
    where_sql = "WHERE " + " AND ".join(where) if where else ""
    nfe_rows = query(
        f"""
        SELECT n.*,o.numero_oc,
          (SELECT COUNT(*) FROM nfe_importacao_itens i WHERE i.nfe_importacao_id=n.id) total_itens,
          (SELECT COUNT(*) FROM nfe_importacao_itens i WHERE i.nfe_importacao_id=n.id AND i.status='pendente') itens_pendentes
        FROM nfe_importacoes n
        LEFT JOIN ordens_compra o ON o.id=n.ordem_id
        {where_sql}
        ORDER BY n.criado_em DESC, n.id DESC
        """,
        tuple(params),
    )
    produtos_ativos = query("SELECT id,codigo,descricao FROM produtos WHERE status='ativo' ORDER BY descricao")
    if using_postgres():
        ordens_disponiveis = query(
            """
            SELECT o.id,o.numero_oc,o.data_entrega,o.valor_total,
                   f.id fornecedor_id,f.nome fornecedor,
                   COUNT(i.id) total_itens,
                   STRING_AGG((p.codigo || ' - ' || p.descricao), ' | ' ORDER BY i.id) produtos_resumo
            FROM ordens_compra o
            JOIN fornecedores f ON f.id=o.fornecedor_id
            LEFT JOIN ordem_compra_itens i ON i.ordem_id=o.id
            LEFT JOIN produtos p ON p.id=i.produto_id
            WHERE o.status_compra NOT IN ('recebida','entregue completo','entrega parcial recebida','cancelada','reprovada')
            GROUP BY o.id,f.id,f.nome
            ORDER BY o.id DESC
            """
        )
    else:
        ordens_disponiveis = query(
            """
            SELECT o.id,o.numero_oc,o.data_entrega,o.valor_total,
                   f.id fornecedor_id,f.nome fornecedor,
                   COUNT(i.id) total_itens,
                   GROUP_CONCAT(CONCAT(p.codigo,' - ',p.descricao) ORDER BY i.id SEPARATOR ' | ') produtos_resumo
            FROM ordens_compra o
            JOIN fornecedores f ON f.id=o.fornecedor_id
            LEFT JOIN ordem_compra_itens i ON i.ordem_id=o.id
            LEFT JOIN produtos p ON p.id=i.produto_id
            WHERE o.status_compra NOT IN ('recebida','entregue completo','entrega parcial recebida','cancelada','reprovada')
            GROUP BY o.id
            ORDER BY o.id DESC
            """
        )
    order_import_summaries = {}
    for ordem in ordens_disponiveis:
        order_items = query(
            """
            SELECT i.id item_id,i.produto_id,p.descricao produto,COALESCE(i.unidade_compra,p.unidade,'UN') unidade,
                   i.quantidade,i.preco_negociado,i.valor_total_item
            FROM ordem_compra_itens i
            JOIN produtos p ON p.id=i.produto_id
            WHERE i.ordem_id=%s
            ORDER BY i.id
            """,
            (ordem["id"],),
        )
        order_import_summaries[str(ordem["id"])] = {
            "fornecedor": ordem["fornecedor"],
            "valor_total": float(decimal_value(ordem["valor_total"])),
            "itens": [
                {
                    "produto": item["produto"],
                    "item_id": item["item_id"],
                    "produto_id": item["produto_id"],
                    "quantidade": float(decimal_value(item["quantidade"])),
                    "unidade": item["unidade"],
                    "valor_unitario": float(decimal_value(item["preco_negociado"])),
                    "valor_total": float(decimal_value(item["valor_total_item"])),
                }
                for item in order_items
            ],
        }
    fornecedores_ativos = query("SELECT id,nome,cnpj FROM fornecedores WHERE status='ativo' ORDER BY nome")
    unidades = query("SELECT codigo,nome FROM unidades_medida WHERE status='ativo' ORDER BY nome")
    categorias = query("SELECT nome FROM cadastro_categorias WHERE status='ativo' ORDER BY nome")
    tipos_pagamento = query("SELECT nome FROM cadastro_tipos_pagamento WHERE status='ativo' ORDER BY nome")
    manual_products = query("SELECT id,codigo,descricao,unidade,quantidade_por_unidade_compra,unidade_base,fornecedor_id FROM produtos WHERE status<>'inativo' ORDER BY descricao")
    nfe_items_by_id = {}
    order_summary_by_nfe_id = {}
    for nfe in nfe_rows:
        nfe_items_by_id[nfe["id"]] = query(
            """
            SELECT i.*,p.codigo codigo_interno,p.descricao descricao_interna,p.status produto_status,
                   p.quantidade_por_unidade_compra,p.unidade_base,n.fornecedor,n.numero_nf,n.data_entrada
            FROM nfe_importacao_itens i
            JOIN nfe_importacoes n ON n.id=i.nfe_importacao_id
            LEFT JOIN produtos p ON p.id=i.produto_id
            WHERE i.nfe_importacao_id=%s
            ORDER BY i.id
            """,
            (nfe["id"],),
        )
        for item in nfe_items_by_id[nfe["id"]]:
            factor = decimal_value(item.get("quantidade_por_unidade_compra") or 1)
            if factor <= 0:
                factor = Decimal("1")
            base_qty = decimal_value(item["quantidade"]) * factor
            item["quantidade_total_base"] = base_qty
            item["valor_unitario_base"] = decimal_value(item["valor_total"]) / base_qty if base_qty > 0 else decimal_value(item["valor_unitario"])
            item["unidade_base"] = item.get("unidade_base") or item.get("unidade") or "UN"
            purchase_unit = (item.get("unidade") or "").upper()
            item["alerta_conversao"] = item.get("produto_id") and factor == Decimal("1") and purchase_unit in ("PALLET", "CX", "CAIXA", "ROLO")
        if nfe.get("ordem_id"):
            order = query(
                """
                SELECT o.id,o.numero_oc,o.valor_total,o.data_entrega,f.nome fornecedor
                FROM ordens_compra o
                JOIN fornecedores f ON f.id=o.fornecedor_id
                WHERE o.id=%s
                """,
                (nfe["ordem_id"],),
                one=True,
            )
            if order:
                order_items = query(
                    """
                    SELECT p.descricao produto,COALESCE(i.unidade_compra,p.unidade,'UN') unidade,
                           i.quantidade,i.valor_total_item
                    FROM ordem_compra_itens i
                    JOIN produtos p ON p.id=i.produto_id
                    WHERE i.ordem_id=%s
                    ORDER BY i.id
                    """,
                    (nfe["ordem_id"],),
                )
                order["itens"] = order_items
                order_summary_by_nfe_id[nfe["id"]] = order
    return render_template(
        "recebimentos.html",
        recebimentos=rows,
        nfe_rows=nfe_rows,
        nfe_items_by_id=nfe_items_by_id,
        order_summary_by_nfe_id=order_summary_by_nfe_id,
        produtos_ativos=produtos_ativos,
        ordens_disponiveis=ordens_disponiveis,
        order_import_summaries=order_import_summaries,
        fornecedores=fornecedores_ativos,
        unidades=unidades,
        categorias=categorias,
        tipos_pagamento=tipos_pagamento,
        manual_products=manual_products,
        filtros=filters,
    )


@app.post("/recebimentos/<int:nfe_id>/excluir")
def excluir_recebimento(nfe_id):
    ensure_nfe_schema()
    next_url = request.form.get("next") or url_for("recebimentos")
    if not next_url.startswith("/recebimentos"):
        next_url = url_for("recebimentos")
    connection = get_db()
    cursor = connection.cursor() if using_postgres() else connection.cursor(dictionary=True)
    try:
        cursor.execute("SELECT * FROM nfe_importacoes WHERE id=%s", (nfe_id,))
        receipt = cursor.fetchone()
        if not receipt:
            flash("Recebimento nao encontrado.", "warning")
            return redirect(next_url)
        cursor.execute(
            """
            SELECT DISTINCT produto_id
            FROM nfe_importacao_itens
            WHERE nfe_importacao_id=%s AND produto_id IS NOT NULL
            """,
            (nfe_id,),
        )
        product_ids = [row["produto_id"] for row in cursor.fetchall()]
        cursor.execute("DELETE FROM movimentacoes_estoque WHERE recebimento_id=%s", (nfe_id,))
        cursor.execute("DELETE FROM historico_custos WHERE recebimento_id=%s", (nfe_id,))
        cursor.execute("DELETE FROM recebimento_auditoria WHERE recebimento_id=%s", (nfe_id,))
        cursor.execute("DELETE FROM nfe_importacao_itens WHERE nfe_importacao_id=%s", (nfe_id,))
        cursor.execute("DELETE FROM nfe_importacoes WHERE id=%s", (nfe_id,))
        recalculate_products_after_receipt_delete(cursor, product_ids)
        if receipt.get("ordem_id"):
            cursor.execute(
                """
                UPDATE ordens_compra
                SET status_compra='aprovada'
                WHERE id=%s AND status_compra='recebida'
                """,
                (receipt["ordem_id"],),
            )
        connection.commit()
        flash("Recebimento excluido com sucesso.", "success")
    except Exception:
        connection.rollback()
        flash("Nao foi possivel excluir este recebimento. Tente novamente.", "danger")
    finally:
        cursor.close()
        connection.close()
    return redirect(next_url)


def import_nfe_file(file_storage, ordem_id=None, observacoes=None, data_entrada=None, data_entrega=None, tipo_entrega=None, delivered_quantities=None, data_entrega_restante=None):
    ensure_auxiliary_registries()
    ensure_product_schema()
    ensure_nfe_schema()
    ensure_order_schema()
    data = parse_nfe_xml(file_storage)
    data_entrada = normalize_iso_date(data_entrada, "Data de entrada", default_today=True)
    data_entrega = normalize_iso_date(data_entrega, "Data de entrega")
    data_entrega_restante = normalize_iso_date(data_entrega_restante, "Data prevista para entrega restante")
    exists = query("SELECT id FROM nfe_importacoes WHERE chave_nfe=%s", (data["chave_nfe"],), one=True)
    if exists:
        return {"status": "duplicado", "nfe_id": exists["id"], "pending": 0, "products_created": [], "supplier_created": None}
    if not data_entrega:
        raise ValueError("Informe a data de entrega para concluir a entrada do XML.")
    if ordem_id and tipo_entrega not in ("completa", "parcial"):
        raise ValueError("Informe se a entrega e completa ou parcial.")

    supplier = get_or_create_nfe_supplier(
        data["fornecedor"],
        data["fornecedor_cnpj"],
        data.get("fornecedor_telefone"),
    )
    fornecedor_id = supplier["id"]
    identified_items = []
    pending = 0
    created_products = []
    for item in data["items"]:
        product = identify_nfe_product(
            data["fornecedor"],
            data["fornecedor_cnpj"],
            item["codigo_fornecedor"],
            item["descricao"],
        )
        if not product:
            product = create_pending_product_from_nfe_item(item, fornecedor_id)
            pending += 1
            created_products.append(product)
        register_product_supplier_link(data["fornecedor"], data["fornecedor_cnpj"], item, product["id"])
        identified_items.append((item, product, product.get("created", False)))

    if ordem_id:
        order = query("SELECT id FROM ordens_compra WHERE id=%s", (ordem_id,), one=True)
        if not order:
            raise ValueError("Pedido de compra informado nao existe.")

    connection = get_db()
    cursor = connection.cursor()
    try:
        cursor.execute(
            """
            INSERT INTO nfe_importacoes
            (ordem_id,origem,fornecedor,fornecedor_cnpj,numero_nf,tipo_documento,serie,chave_nfe,data_emissao,data_entrada,data_entrega,tipo_entrega,valor_total,observacoes,usuario_responsavel,status_recebimento,confirmado_em)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,NOW())
            """ + (" RETURNING id" if using_postgres() else ""),
            (
                ordem_id,
                "XML",
                data["fornecedor"],
                data["fornecedor_cnpj"],
                data["numero_nf"],
                "Nota Fiscal",
                data["serie"],
                data["chave_nfe"],
                data["data_emissao"],
                data_entrada,
                data_entrega,
                tipo_entrega if ordem_id else None,
                data["valor_total"],
                observacoes or None,
                "Sistema",
                "Produto Pendente de Cadastro" if pending else ("Recebido Parcialmente" if ordem_id and tipo_entrega == "parcial" else ("Recebido" if ordem_id else "Aguardando Ordem de Compra")),
            ),
        )
        if using_postgres():
            nfe_id = cursor.fetchone()["id"]
        else:
            nfe_id = cursor.lastrowid
        for item, product, product_created in identified_items:
            cursor.execute(
                """
                INSERT INTO nfe_importacao_itens
                (nfe_importacao_id,produto_id,codigo_fornecedor,descricao,ncm,cfop,cest,unidade,quantidade,valor_unitario,desconto,valor_total,ean,status)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """,
                (
                    nfe_id,
                    product["id"] if product else None,
                    item["codigo_fornecedor"],
                    item["descricao"],
                    item["ncm"],
                    item["cfop"],
                    item["cest"],
                    item["unidade"],
                    item["quantidade"],
                    item["valor_unitario"],
                    0,
                    item["valor_total"],
                    item["ean"],
                    "pendente" if product_created else "vinculado",
                ),
            )
        connection.commit()
        audit_receipt(nfe_id, "Importacao XML", f"Documento {data['numero_nf']} importado.", "Sistema", "XML")
        apply_receipt_effects(nfe_id)
        if ordem_id:
            delivery_result = process_order_delivery(ordem_id, nfe_id, tipo_entrega, data_entrada, delivered_quantities, data_entrega_restante, data["numero_nf"])
            if delivery_result and delivery_result.get("partial_order"):
                execute("UPDATE nfe_importacoes SET ordem_parcial_id=%s WHERE id=%s", (delivery_result["partial_order"]["id"], nfe_id))
        return {
            "status": "importado",
            "nfe_id": nfe_id,
            "pending": pending,
            "products_created": created_products,
            "supplier_created": supplier if supplier["created"] else None,
        }
    except Exception:
        connection.rollback()
        raise
    finally:
        cursor.close()
        connection.close()


@app.post("/recebimentos/importar_xml")
def importar_xml_nfe():
    ensure_nfe_schema()
    ensure_order_schema()
    files = [file for file in request.files.getlist("xmls") + request.files.getlist("xml") if file and file.filename]
    ordem_id = request.form.get("ordem_id") or None
    observacoes = (request.form.get("observacoes") or "").strip()
    try:
        data_entrada = normalize_iso_date(request.form.get("data_entrada"), "Data de entrada", default_today=True)
        data_entrega = normalize_iso_date(request.form.get("data_entrega"), "Data de entrega")
        data_entrega_restante = normalize_iso_date(request.form.get("data_entrega_restante"), "Data prevista para entrega restante")
    except ValueError as error:
        flash(str(error), "danger")
        return redirect(url_for("recebimentos"))
    tipo_entrega = request.form.get("tipo_entrega") if ordem_id else None
    delivered_quantities = delivered_quantities_from_form(request.form)
    if not data_entrega:
        flash("Informe a data de entrega para concluir a entrada do XML.", "danger")
        return redirect(url_for("recebimentos"))
    if ordem_id and not query("SELECT id FROM ordens_compra WHERE id=%s", (ordem_id,), one=True):
        flash("Pedido de compra informado nao existe.", "danger")
        return redirect(url_for("recebimentos"))
    if ordem_id and tipo_entrega not in ("completa", "parcial"):
        flash("Informe se a entrega e completa ou parcial.", "danger")
        return redirect(url_for("recebimentos"))
    if ordem_id and tipo_entrega == "parcial" and not data_entrega_restante:
        flash("Informe a nova data de entrega do restante da ordem.", "danger")
        return redirect(url_for("recebimentos"))
    if not files:
        flash("Selecione ao menos um arquivo XML de NF-e.", "danger")
        return redirect(url_for("recebimentos"))

    imported = errors = duplicates = pending_products = 0
    auto_products = []
    auto_suppliers = []
    error_messages = []
    last_nfe_id = None
    for file in files:
        if not file.filename.lower().endswith(".xml"):
            errors += 1
            error_messages.append(f"{file.filename}: formato invalido")
            continue
        try:
            result = import_nfe_file(
                file,
                ordem_id=ordem_id,
                observacoes=observacoes,
                data_entrada=data_entrada,
                data_entrega=data_entrega,
                tipo_entrega=tipo_entrega,
                delivered_quantities=delivered_quantities,
                data_entrega_restante=data_entrega_restante,
            )
            if result["status"] == "duplicado":
                duplicates += 1
                last_nfe_id = result["nfe_id"]
            else:
                imported += 1
                pending_products += result["pending"]
                auto_products.extend(result.get("products_created", []))
                if result.get("supplier_created"):
                    auto_suppliers.append(result["supplier_created"])
                last_nfe_id = result["nfe_id"]
        except Exception as error:
            errors += 1
            error_messages.append(f"{file.filename}: {error}")

    summary = (
        f"Importacao em massa concluida. XMLs importados com sucesso: {imported}. "
        f"XMLs com erro: {errors}. Notas cadastradas: {imported}. "
        f"Produtos cadastrados automaticamente para revisao: {len(auto_products)}. "
        f"Fornecedores cadastrados automaticamente: {len(auto_suppliers)}. "
        f"Duplicadas ignoradas: {duplicates}."
    )
    if auto_suppliers:
        names = ", ".join(sorted({supplier["nome"] for supplier in auto_suppliers})[:5])
        summary += f" Fornecedores criados: {names}."
    if auto_products:
        names = ", ".join(product["descricao"] for product in auto_products[:5])
        summary += f" Produtos pendentes: {names}."
    if error_messages:
        summary += " " + " | ".join(error_messages[:5])
        if len(error_messages) > 5:
            summary += f" | Mais {len(error_messages) - 5} erro(s)."
    flash(summary, "success" if imported else "warning")
    if last_nfe_id:
        return redirect(url_for("recebimentos"))
    return redirect(url_for("recebimentos"))


@app.post("/recebimentos/manual")
def salvar_recebimento_manual():
    ensure_auxiliary_registries()
    ensure_product_schema()
    ensure_nfe_schema()
    ensure_order_schema()
    form = request.form
    fornecedor_id = form.get("fornecedor_id")
    supplier = query("SELECT * FROM fornecedores WHERE id=%s", (fornecedor_id,), one=True)
    if not supplier:
        flash("Fornecedor obrigatorio para entrada manual.", "danger")
        return redirect(url_for("recebimentos"))
    numero_doc = (form.get("numero_nf") or "").strip() or f"MANUAL-{datetime.now().strftime('%Y%m%d%H%M%S')}"
    tipo_documento = form.get("tipo_documento") or "Sem Documento"
    try:
        data_entrada = normalize_iso_date(form.get("data_entrada"), "Data de entrada", default_today=True)
        data_compra = normalize_iso_date(form.get("data_compra") or data_entrada, "Data da compra", default_today=True)
        data_entrega_restante = normalize_iso_date(form.get("data_entrega_restante"), "Data prevista para entrega restante")
    except ValueError as error:
        flash(str(error), "danger")
        return redirect(url_for("recebimentos"))
    ordem_id = form.get("ordem_id") or None
    tipo_entrega = form.get("tipo_entrega") if ordem_id else None
    delivered_quantities = delivered_quantities_from_form(form)
    if ordem_id and not query("SELECT id FROM ordens_compra WHERE id=%s", (ordem_id,), one=True):
        flash("Pedido de compra informado nao existe.", "danger")
        return redirect(url_for("recebimentos"))
    if ordem_id and tipo_entrega not in ("completa", "parcial"):
        flash("Informe se a entrega e completa ou parcial.", "danger")
        return redirect(url_for("recebimentos"))
    if ordem_id and tipo_entrega == "parcial" and not data_entrega_restante:
        flash("Informe a nova data de entrega do restante da ordem.", "danger")
        return redirect(url_for("recebimentos"))
    codigo_fornecedor_list = form.getlist("codigo_fornecedor[]")
    descricao_list = form.getlist("descricao[]")
    produto_id_list = form.getlist("produto_id[]")
    codigo_interno_list = form.getlist("codigo_interno[]")
    unidade_list = form.getlist("unidade[]")
    quantidade_list = form.getlist("quantidade[]")
    valor_unitario_list = form.getlist("valor_unitario[]")
    desconto_list = form.getlist("desconto[]")
    items = []
    for index, descricao in enumerate(descricao_list):
        descricao = (descricao or "").strip()
        qty = decimal_value(quantidade_list[index] if index < len(quantidade_list) else 0)
        unit = money_value(valor_unitario_list[index] if index < len(valor_unitario_list) else 0)
        desconto = money_value(desconto_list[index] if index < len(desconto_list) else 0)
        if not descricao and qty == 0:
            continue
        if not descricao or qty <= 0 or unit < 0 or desconto < 0:
            flash("Revise os produtos da entrada manual. Descricao, quantidade e valores devem ser validos.", "danger")
            return redirect(url_for("recebimentos"))
        total = (qty * unit) - desconto
        if total < 0:
            flash("O desconto nao pode ser maior que o total do item.", "danger")
            return redirect(url_for("recebimentos"))
        product = None
        product_id = produto_id_list[index] if index < len(produto_id_list) else ""
        codigo_interno = (codigo_interno_list[index] if index < len(codigo_interno_list) else "").strip()
        if product_id:
            product = query("SELECT id,codigo FROM produtos WHERE id=%s", (product_id,), one=True)
        if not product and codigo_interno:
            product = query("SELECT id,codigo FROM produtos WHERE codigo=%s AND status<>'inativo' ORDER BY id DESC LIMIT 1", (codigo_interno,), one=True)
        product_created = False
        item_data = {
            "codigo_fornecedor": (codigo_fornecedor_list[index] if index < len(codigo_fornecedor_list) else "").strip(),
            "descricao": descricao,
            "unidade": unidade_list[index] if index < len(unidade_list) else "UN",
            "quantidade": qty,
            "valor_unitario": unit,
            "desconto": desconto,
            "valor_total": total,
        }
        if not product:
            product = create_pending_product_from_nfe_item(item_data, supplier["id"])
            product_created = True
        register_product_supplier_link(supplier["nome"], supplier.get("cnpj"), item_data, product["id"])
        items.append((item_data, product, product_created))
    if not items:
        flash("Adicione ao menos um produto para salvar a entrada manual.", "danger")
        return redirect(url_for("recebimentos"))

    valor_total = sum(item[0]["valor_total"] for item in items)
    connection = get_db()
    cursor = connection.cursor()
    try:
        cursor.execute(
            """
            INSERT INTO nfe_importacoes
            (ordem_id,origem,fornecedor,fornecedor_cnpj,numero_nf,tipo_documento,serie,chave_nfe,data_emissao,data_entrada,data_entrega,tipo_entrega,valor_total,
             condicao_pagamento,observacoes,usuario_responsavel,status_recebimento,confirmado_em)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,NOW())
            """ + (" RETURNING id" if using_postgres() else ""),
            (
                ordem_id,
                "Manual",
                supplier["nome"],
                supplier.get("cnpj") or "Nao informado",
                numero_doc,
                tipo_documento,
                "",
                f"MANUAL-{datetime.now().strftime('%Y%m%d%H%M%S%f')}",
                data_compra,
                data_entrada,
                data_entrada,
                tipo_entrega if ordem_id else None,
                valor_total,
                form.get("condicao_pagamento"),
                form.get("observacoes"),
                "Sistema",
                "Produto Pendente de Cadastro" if any(product_created for _, _, product_created in items) else ("Recebido Parcialmente" if ordem_id and tipo_entrega == "parcial" else "Recebido"),
            ),
        )
        if using_postgres():
            recebimento_id = cursor.fetchone()["id"]
        else:
            recebimento_id = cursor.lastrowid
        for item_data, product, product_created in items:
            cursor.execute(
                """
                INSERT INTO nfe_importacao_itens
                (nfe_importacao_id,produto_id,codigo_fornecedor,descricao,unidade,quantidade,valor_unitario,desconto,valor_total,status)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """,
                (
                    recebimento_id,
                    product["id"],
                    item_data["codigo_fornecedor"],
                    item_data["descricao"],
                    item_data["unidade"],
                    item_data["quantidade"],
                    item_data["valor_unitario"],
                    item_data["desconto"],
                    item_data["valor_total"],
                    "pendente" if product_created else "vinculado",
                ),
            )
        connection.commit()
    except Exception:
        connection.rollback()
        raise
    finally:
        cursor.close()
        connection.close()

    audit_receipt(recebimento_id, "Entrada manual", f"Documento {numero_doc} registrado manualmente.", "Sistema", "Manual")
    apply_receipt_effects(recebimento_id)
    if ordem_id:
        delivery_result = process_order_delivery(ordem_id, recebimento_id, tipo_entrega, data_entrada, delivered_quantities, data_entrega_restante, numero_doc)
        if delivery_result and delivery_result.get("partial_order"):
            execute("UPDATE nfe_importacoes SET ordem_parcial_id=%s WHERE id=%s", (delivery_result["partial_order"]["id"], recebimento_id))
    flash("Recebimento registrado com sucesso.", "success")
    return redirect(url_for("recebimentos"))


@app.post("/recebimentos/produto_rapido")
def cadastrar_produto_rapido_recebimento():
    ensure_auxiliary_registries()
    ensure_product_schema()
    form = request.form
    fornecedor_id = form.get("fornecedor_id")
    supplier = query("SELECT * FROM fornecedores WHERE id=%s", (fornecedor_id,), one=True)
    if not supplier:
        flash("Fornecedor obrigatorio para cadastro rapido.", "danger")
        return redirect(url_for("recebimentos"))
    quantidade_por_unidade = decimal_value(form.get("quantidade_por_unidade_compra") or 1)
    if quantidade_por_unidade <= 0:
        quantidade_por_unidade = Decimal("1")
    unidade = form.get("unidade") or "UN"
    produto_id = execute(
        """
        INSERT INTO produtos (codigo,descricao,categoria,unidade,quantidade_por_unidade_compra,unidade_base,fornecedor_id,estoque_seguranca,status)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """,
        (
            (form.get("codigo") or "Nao registrado").strip(),
            form.get("descricao"),
            form.get("categoria"),
            unidade,
            quantidade_por_unidade,
            form.get("unidade_base") or unidade,
            fornecedor_id,
            0,
            "ativo",
        ),
    )
    item_data = {"codigo_fornecedor": form.get("codigo_fornecedor") or "", "descricao": form.get("descricao") or ""}
    register_product_supplier_link(supplier["nome"], supplier.get("cnpj"), item_data, produto_id)
    flash("Produto cadastrado rapidamente. Selecione-o na entrada manual.", "success")
    return redirect(url_for("recebimentos"))


@app.get("/api/ordens/<int:ordem_id>/itens")
def api_ordem_itens(ordem_id):
    ordem = query(
        """
        SELECT o.*,f.nome fornecedor
        FROM ordens_compra o
        JOIN fornecedores f ON f.id=o.fornecedor_id
        WHERE o.id=%s
        """,
        (ordem_id,),
        one=True,
    )
    if not ordem:
        return jsonify({"items": []})
    items = query(
        """
        SELECT i.*,p.codigo produto_codigo,p.descricao produto_descricao,
               COALESCE(i.unidade_compra,p.unidade) produto_unidade
        FROM ordem_compra_itens i
        JOIN produtos p ON p.id=i.produto_id
        WHERE i.ordem_id=%s
        ORDER BY i.id
        """,
        (ordem_id,),
    )
    return jsonify(
        {
            "fornecedor_id": ordem["fornecedor_id"],
            "items": [
                {
                    "produto_id": item["produto_id"],
                    "codigo": item["produto_codigo"],
                    "descricao": item["produto_descricao"],
                    "unidade": item["produto_unidade"],
                    "quantidade": str(item["quantidade"]),
                    "valor_unitario": str(item["preco_negociado"]),
                    "frete": str(item["frete"]),
                }
                for item in items
            ],
        }
    )


@app.post("/recebimentos/nfe_itens/<int:item_id>/vincular")
def vincular_item_nfe(item_id):
    ensure_nfe_schema()
    produto_id = request.form.get("produto_id")
    if not produto_id:
        flash("Selecione um codigo interno para vincular.", "danger")
        return redirect(url_for("recebimentos"))

    item = query(
        """
        SELECT i.*,n.fornecedor,n.fornecedor_cnpj,n.id nfe_id
        FROM nfe_importacao_itens i
        JOIN nfe_importacoes n ON n.id=i.nfe_importacao_id
        WHERE i.id=%s
        """,
        (item_id,),
        one=True,
    )
    product = query("SELECT id,codigo FROM produtos WHERE id=%s", (produto_id,), one=True)
    if not item or not product:
        flash("Item ou produto nao encontrado.", "danger")
        return redirect(url_for("recebimentos"))

    register_product_supplier_link(
        item["fornecedor"],
        item["fornecedor_cnpj"],
        {"codigo_fornecedor": item["codigo_fornecedor"], "descricao": item["descricao"]},
        produto_id,
    )
    execute("UPDATE nfe_importacao_itens SET produto_id=%s,status='vinculado' WHERE id=%s", (produto_id, item_id))
    refresh_receipt_status(item["nfe_id"])
    apply_receipt_effects(item["nfe_id"])
    flash("Vinculo salvo. Proximas importacoes reconhecerÃ£o este produto automaticamente.", "success")
    return redirect(url_for("recebimentos", nfe_id=item["nfe_id"]))


@app.post("/recebimentos/nfe/<int:nfe_id>/vincular_ordem")
def vincular_ordem_nfe(nfe_id):
    ensure_nfe_schema()
    ordem_id = request.form.get("ordem_id")
    if not ordem_id:
        flash("Selecione uma Ordem de Compra para vincular a NF-e.", "danger")
        return redirect(url_for("recebimentos", nfe_id=nfe_id))
    nfe = query("SELECT id,data_entrada,data_entrega FROM nfe_importacoes WHERE id=%s", (nfe_id,), one=True)
    ordem = query("SELECT id FROM ordens_compra WHERE id=%s", (ordem_id,), one=True)
    if not nfe or not ordem:
        flash("NF-e ou Ordem de Compra nao encontrada.", "danger")
        return redirect(url_for("recebimentos", nfe_id=nfe_id))
    execute("UPDATE nfe_importacoes SET ordem_id=%s WHERE id=%s", (ordem_id, nfe_id))
    refresh_receipt_status(nfe_id)
    execute("UPDATE recebimentos SET data_real=%s,status='entregue completo' WHERE ordem_id=%s", (nfe["data_entrega"] or nfe["data_entrada"], ordem_id))
    execute("UPDATE ordens_compra SET status_compra='recebida' WHERE id=%s", (ordem_id,))
    audit_receipt(nfe_id, "Vinculo de OC", f"Ordem {ordem_id} vinculada.", "Sistema")
    flash("NF-e vinculada a Ordem de Compra.", "success")
    return redirect(url_for("recebimentos", nfe_id=nfe_id))


@app.route("/orcamentos", methods=["GET", "POST"])
def orcamentos():
    return redirect(url_for("cadastros", modulo="orcamentos"))


@app.post("/orcamentos/<int:item_id>/excluir")
def excluir_orcamento(item_id):
    execute("DELETE FROM orcamentos WHERE id=%s", (item_id,))
    flash("Orcamento excluido.", "success")
    return redirect(url_for("cadastros", modulo="orcamentos"))


if __name__ == "__main__":
    app.run(debug=True)















